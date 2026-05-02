# Programador: Oscar Amaya Romero
from __future__ import annotations

from datetime import date, datetime
import csv
import json
import html
from io import BytesIO, StringIO

import psycopg2
from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from werkzeug.security import generate_password_hash

from azdigital.decorators import admin_required, rol_requerido, superadmin_required
from azdigital.services.auth_service import verificar_password
from azdigital.repositories import (
    actividades_repo,
    cierre_caja_repo,
    clientes_repo,
    historial_usuarios_repo,
    compras_repo,
    compras_reports_repo,
    empresas_repo,
    inventario_reports_repo,
    kardex_repo,
    lista_compras_repo,
    mh_unidades_repo,
    presentaciones_repo,
    productos_repo,
    promociones_repo,
    proveedores_repo,
    sucursales_repo,
    usuarios_repo,
    ventas_repo,
    ventas_reports_repo,
)
from azdigital.utils.historial_helper import registrar_accion
from azdigital.utils.mh_cat003_unidades import catalogo_para_select_optgroups, normalizar_codigo_mh
from azdigital.utils.precio_umb_desde_caja import (
    aplicar_derivacion_desde_presentacion,
    presentacion_tiene_monto_derivable,
)
from database import ConexionDB

bp = Blueprint("admin", __name__)


def _empresa_id():
    return session.get("empresa_id", 1)


def _sucursal_id_session() -> int | None:
    """sucursal_id en sesión puede ser str corrupto desde el cliente; evita ValueError → 500."""
    raw = session.get("sucursal_id")
    if raw is None or raw == "":
        return None
    try:
        n = int(raw)
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _obtener_ventas_periodo(inicio: str, fin: str, empresa_id: int = None):
    db = ConexionDB()
    emp = empresa_id or _empresa_id()
    exp = ventas_repo._sql_etiqueta_cliente_venta().strip()
    sql_detalle = f"""
        SELECT v.id, TO_CHAR(v.fecha_registro, 'DD/MM/YYYY HH:MI AM'), ({exp}), v.tipo_pago, v.total_pagar
        FROM ventas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.fecha_registro::date BETWEEN %s AND %s AND (v.empresa_id IS NULL OR v.empresa_id = %s)
          AND COALESCE(v.estado, 'ACTIVO') = 'ACTIVO'
        ORDER BY v.id DESC
    """
    return db.ejecutar_sql(sql_detalle, (inicio, fin, emp), es_select=True) or []


@bp.route("/reporte")
@rol_requerido("GERENTE", "CONTADOR")
def reporte():
    db = ConexionDB()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    ventas_detalle = _obtener_ventas_periodo(inicio, fin)
    emp = _empresa_id()
    sql_evol = """
        SELECT fecha_registro::date, SUM(total_pagar)
        FROM ventas
        WHERE fecha_registro::date BETWEEN %s AND %s AND (empresa_id IS NULL OR empresa_id = %s)
          AND COALESCE(estado, 'ACTIVO') = 'ACTIVO'
        GROUP BY 1
        ORDER BY 1 ASC
    """
    res_evol = db.ejecutar_sql(sql_evol, (inicio, fin, emp), es_select=True) or []
    sql_top = """
        SELECT p.nombre, COALESCE(SUM(dv.cantidad), 0)::int
        FROM venta_detalles dv
        JOIN productos p ON dv.producto_id = p.id
        JOIN ventas v ON dv.venta_id = v.id
        WHERE v.fecha_registro::date BETWEEN %s AND %s
          AND (v.empresa_id IS NULL OR v.empresa_id = %s)
          AND COALESCE(v.estado, 'ACTIVO') = 'ACTIVO'
        GROUP BY p.id, p.nombre
        ORDER BY 2 DESC
        LIMIT 10
    """
    top_productos = db.ejecutar_sql(sql_top, (inicio, fin, emp), es_select=True) or []
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        sugerencia_compra = productos_repo.productos_stock_bajo(cur, umbral=10, empresa_id=emp) or []
        vendido_periodo = {}
        if sugerencia_compra:
            ids = [r[0] for r in sugerencia_compra]
            placeholders = ",".join(["%s"] * len(ids))
            cur.execute(
                f"""
                SELECT dv.producto_id, SUM(dv.cantidad)
                FROM venta_detalles dv
                JOIN ventas v ON v.id = dv.venta_id
                WHERE v.fecha_registro::date BETWEEN %s AND %s
                  AND (v.empresa_id IS NULL OR v.empresa_id = %s)
                  AND COALESCE(v.estado, 'ACTIVO') = 'ACTIVO'
                  AND dv.producto_id IN ({placeholders})
                GROUP BY dv.producto_id
                """,
                (inicio, fin, emp, *ids),
            )
            for r in cur.fetchall() or []:
                vendido_periodo[r[0]] = float(r[1] or 0)
        sugerencia_con_ventas = [
            (r[0], r[1], r[2], vendido_periodo.get(r[0], 0))
            for r in sugerencia_compra
        ]
        sugerencia_con_ventas.sort(key=lambda x: (-x[3], x[2]))
    except Exception:
        sugerencia_con_ventas = []
    finally:
        cur.close()
        conn.close()
    total_periodo = 0.0
    count_efectivo = 0
    count_tarjeta = 0
    for v in ventas_detalle or []:
        if len(v) > 4:
            try:
                total_periodo += float(v[4])
            except (TypeError, ValueError):
                pass
        if len(v) > 3 and str(v[3] or "").strip().upper() == "EFECTIVO":
            count_efectivo += 1
        elif len(v) > 3:
            count_tarjeta += 1

    dias_periodo = 1
    try:
        d_inicio = date.fromisoformat(str(inicio)[:10])
        d_fin = date.fromisoformat(str(fin)[:10])
        dias_periodo = max(1, (d_fin - d_inicio).days + 1)
    except (ValueError, TypeError):
        pass
    promedio_diario = total_periodo / dias_periodo if dias_periodo else 0
    promedio_mensual = promedio_diario * 30 if dias_periodo else 0

    return render_template(
        "reporte.html",
        ventas_detalle=ventas_detalle or [],
        inicio=inicio,
        fin=fin,
        total_periodo=total_periodo,
        promedio_diario=promedio_diario,
        promedio_mensual=promedio_mensual,
        dias_periodo=dias_periodo,
        count_efectivo=count_efectivo,
        count_tarjeta=count_tarjeta,
        labels_barras=[str(r[0]) for r in res_evol] if res_evol else [],
        valores_barras=[float(r[1]) for r in res_evol] if res_evol else [],
        top_productos=top_productos,
        nombres_top_productos=[(r[0] or "")[:28] for r in top_productos] if top_productos else [],
        cantidades_top_productos=[r[1] for r in top_productos] if top_productos else [],
        sugerencia_compra=sugerencia_con_ventas,
    )


def _formatear_fecha_pdf(iso: str) -> str:
    try:
        return date.fromisoformat(iso.strip()[:10]).strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return str(iso or "")


def _datos_cabecera_reporte_ventas(inicio: str, fin: str) -> dict[str, str]:
    """Empresa, sucursal, período y fecha de generación (PDF/Excel)."""
    emp_id = _empresa_id()
    nombre_empresa = (session.get("empresa_nombre") or "Empresa").strip() or "Empresa"
    nombre_sucursal = "Todas las sucursales (consolidado empresa)"
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        emp = empresas_repo.get_empresa(cur, emp_id)
        if emp:
            nombre_empresa = (emp[1] or (emp[9] if len(emp) > 9 else None) or nombre_empresa) or nombre_empresa
        sid = session.get("sucursal_id")
        if sid is not None and str(sid).strip().isdigit():
            suc = sucursales_repo.get_sucursal(cur, int(sid))
            if suc and len(suc) > 1 and suc[1]:
                nombre_sucursal = str(suc[1])
            elif sid:
                nombre_sucursal = f"Sucursal #{sid}"
    finally:
        cur.close()
        conn.close()
    generado = datetime.now().strftime("%d/%m/%Y %H:%M")
    periodo_txt = _formatear_fecha_pdf(inicio)
    if inicio != fin:
        periodo_txt = f"{periodo_txt} al {_formatear_fecha_pdf(fin)}"
    return {
        "empresa": str(nombre_empresa),
        "sucursal": str(nombre_sucursal),
        "periodo": periodo_txt,
        "generado": generado,
    }


def _datos_cabecera_inventario(periodo_txt: str = "", sucursal_nombre: str = None, emp_id: int = None) -> dict[str, str]:
    """Cabecera para reportes de inventario (Excel/PDF)."""
    emp_id = emp_id or _empresa_id()
    nombre_empresa = (session.get("empresa_nombre") or "Empresa").strip() or "Empresa"
    nombre_sucursal = sucursal_nombre if sucursal_nombre else "Todas las sucursales"
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        emp = empresas_repo.get_empresa(cur, emp_id)
        if emp:
            nombre_empresa = (emp[1] or (emp[9] if len(emp) > 9 else None) or nombre_empresa) or nombre_empresa
        if not sucursal_nombre and session.get("sucursal_id"):
            sid = session.get("sucursal_id")
            suc = sucursales_repo.get_sucursal(cur, int(sid)) if str(sid).strip().isdigit() else None
            if suc and len(suc) > 1 and suc[1]:
                nombre_sucursal = str(suc[1])
    finally:
        cur.close()
        conn.close()
    return {
        "empresa": str(nombre_empresa),
        "sucursal": str(nombre_sucursal),
        "periodo": periodo_txt,
        "generado": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }


def _estilos_excel_elegante():
    """Estilos reutilizables para Excel (mismo look que ventas)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="FFCBD5E1")
    return {
        "border_all": Border(left=thin, right=thin, top=thin, bottom=thin),
        "fill_header": PatternFill(start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid"),
        "fill_total": PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid"),
        "fill_alt": PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid"),
        "font_header": Font(bold=True, color="FFFFFFFF", size=10),
        "font_title": Font(bold=True, size=15, color="FF1E293B"),
        "font_meta": Font(size=10, color="FF475569"),
        "font_total": Font(bold=True, size=10, color="FF0F172A"),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "align_left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "align_right": Alignment(horizontal="right", vertical="center"),
    }


def _estilos_pdf_elegante():
    """Estilos reutilizables para PDF (reportlab)."""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="RptMeta", parent=styles["Normal"], fontSize=10, leading=14, textColor=colors.HexColor("#475569")))
    styles.add(ParagraphStyle(name="RptSubtitle", parent=styles["Normal"], fontSize=11, leading=14, textColor=colors.HexColor("#1e293b"), spaceAfter=4))
    return styles, colors


@bp.route("/reporte/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR")
def exportar_reporte_excel():
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    ventas_detalle = _obtener_ventas_periodo(inicio, fin)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500

    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ventas"

    thin = Side(style="thin", color="FFCBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill(start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid")
    fill_total = PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid")
    fill_alt = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFFFF", size=10)
    font_title = Font(bold=True, size=15, color="FF1E293B")
    font_meta = Font(size=10, color="FF475569")
    font_total = Font(bold=True, size=10, color="FF0F172A")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    tcell = ws.cell(r, 1, "Reporte de ventas")
    tcell.font = font_title
    tcell.alignment = align_center
    ws.row_dimensions[r].height = 26

    for texto in (
        f"Período: {cab['periodo']}",
        f"Empresa: {cab['empresa']}",
        f"Sucursal: {cab['sucursal']}",
        f"Generado: {cab['generado']}",
    ):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        m = ws.cell(r, 1, texto)
        m.font = font_meta
        m.alignment = align_left

    r += 1
    ws.row_dimensions[r].height = 6

    hdr = r + 1
    headers = ["ID", "FECHA/HORA", "CLIENTE", "MÉTODO", "TOTAL"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(hdr, col, h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
    ws.row_dimensions[hdr].height = 22

    data_start = hdr + 1
    for i, v in enumerate(ventas_detalle):
        rr = data_start + i
        ws.cell(rr, 1, v[0])
        ws.cell(rr, 2, str(v[1]))
        ws.cell(rr, 3, str(v[2]))
        ws.cell(rr, 4, str(v[3]))
        ce = ws.cell(rr, 5, float(v[4]))
        ce.number_format = '"$"#,##0.00'
        ce.alignment = align_right
        ws.cell(rr, 1).alignment = align_center
        ws.cell(rr, 2).alignment = align_left
        ws.cell(rr, 3).alignment = align_left
        ws.cell(rr, 4).alignment = align_center
        for col in range(1, 6):
            cell = ws.cell(rr, col)
            cell.border = border_all
            if i % 2:
                cell.fill = fill_alt

    total = sum(float(v[4]) for v in ventas_detalle)
    tot_row = data_start + len(ventas_detalle)
    ws.cell(tot_row, 1, "")
    ws.cell(tot_row, 2, "")
    ws.cell(tot_row, 3, "")
    ws.cell(tot_row, 4, "TOTAL")
    te = ws.cell(tot_row, 5, total)
    te.number_format = '"$"#,##0.00'
    te.alignment = align_right
    ws.cell(tot_row, 4).alignment = align_right
    for col in range(1, 6):
        c = ws.cell(tot_row, col)
        c.border = border_all
        c.fill = fill_total
    ws.cell(tot_row, 4).font = font_total
    te.font = font_total

    foot = tot_row + 2
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=5)
    u = (session.get("username") or "").strip()
    fc = ws.cell(foot, 1, f"AZ DIGITAL — {u}" if u else "AZ DIGITAL")
    fc.font = Font(size=9, italic=True, color="FF94A3B8")
    fc.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"reporte_ventas_{inicio}_{fin}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR")
def exportar_reporte_pdf():
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    ventas_detalle = _obtener_ventas_periodo(inicio, fin)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return "<p>Instala reportlab: pip install reportlab</p>", 500

    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    nombre_empresa = cab["empresa"]
    nombre_sucursal = cab["sucursal"]
    generado = cab["generado"]
    periodo_txt = cab["periodo"]
    buf = BytesIO()
    doc_w, doc_h = letter
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        rightMargin=0.65 * inch,
        leftMargin=0.65 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
    )
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="RptMeta",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#475569"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="RptSubtitle",
            parent=styles["Normal"],
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#1e293b"),
            spaceAfter=4,
        )
    )
    elements = []
    elements.append(Paragraph("Reporte de ventas", styles["Title"]))
    elements.append(Paragraph(f"<b>Período:</b> {html.escape(periodo_txt)}", styles["RptSubtitle"]))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(
        Paragraph(f"<b>Empresa:</b> {html.escape(str(nombre_empresa))}", styles["RptMeta"])
    )
    elements.append(
        Paragraph(f"<b>Sucursal:</b> {html.escape(str(nombre_sucursal))}", styles["RptMeta"])
    )
    elements.append(
        Paragraph(f"<b>Generado:</b> {html.escape(generado)}", styles["RptMeta"])
    )
    elements.append(Spacer(1, 0.28 * inch))

    data = [["ID", "FECHA/HORA", "CLIENTE", "MÉTODO", "TOTAL"]]
    for v in ventas_detalle:
        data.append([str(v[0]), str(v[1]), str(v[2]), str(v[3]), f"${float(v[4]):,.2f}"])
    total = sum(float(v[4]) for v in ventas_detalle)
    data.append(["", "", "", "TOTAL", f"${total:,.2f}"])

    usable_w = doc_w - doc.leftMargin - doc.rightMargin
    col_fracs = [0.08, 0.20, 0.34, 0.16, 0.22]
    cw = [usable_w * c for c in col_fracs]
    t = Table(data, colWidths=cw, repeatRows=1)
    last_r = len(data) - 1
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ALIGN", (1, 1), (3, last_r - 1), "LEFT"),
        ("ALIGN", (1, 0), (3, 0), "CENTER"),
        ("BACKGROUND", (0, last_r), (-1, last_r), colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, last_r), (-1, last_r), "Helvetica-Bold"),
        ("FONTSIZE", (0, last_r), (-1, last_r), 9),
        ("ALIGN", (3, last_r), (3, last_r), "RIGHT"),
    ]
    if last_r > 1:
        style_cmds.append(
            ("ROWBACKGROUNDS", (0, 1), (-1, last_r - 1), [colors.white, colors.HexColor("#f8fafc")]),
        )
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(
        Paragraph(
            f"<i>AZ DIGITAL — {html.escape(str(session.get('username') or ''))}</i>",
            styles["RptMeta"],
        )
    )
    doc.build(elements)
    buf.seek(0)
    fn = f"reporte_ventas_{inicio}_{fin}.pdf"
    return send_file(buf, as_attachment=True, download_name=fn, mimetype="application/pdf")


# --- Reportes de Inventario (Art. 142, F-983, NIC 2) ---


@bp.route("/reporte/inventario")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_inventario_index():
    return redirect(url_for("admin.reporte"))


def _get_productos_para_filtro(cur, empresa_id: int):
    cur.execute(
        "SELECT p.id, p.nombre, p.codigo_barra FROM productos p WHERE p.empresa_id = %s ORDER BY p.nombre LIMIT 500",
        (empresa_id,),
    )
    return cur.fetchall() or []


def _reporte_f983_excel_elegante(filas, ejercicio):
    """Genera Excel F-983 con cabecera elegante (mantiene formato DGI)."""
    from openpyxl import Workbook
    cab = _datos_cabecera_inventario(f"Ejercicio fiscal {ejercicio}")
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = "F983"
    r, num_cols = 1, 11
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    ws.cell(r, 1, "Informe F-983 — Inventario Físico Anual").font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    ws.row_dimensions[r].height = 26
    for txt in (f"Ejercicio: {ejercicio}", f"Empresa: {cab['empresa']}", f"Sucursal: {cab['sucursal']}", f"Generado: {cab['generado']}"):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(r, 1, txt).font = st["font_meta"]
    r += 2
    headers = ["Denominación del Bien", "Código Inventario", "Unidad", "Inv. Inicial", "Compras", "Ventas", "Inv. Final", "Costo Unit. sin IVA", "Categoría", "Ref. Libros", "Ejercicio"]
    hdr = r + 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(hdr, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    for i, row in enumerate(filas):
        rr = hdr + 1 + i
        vals = [str(row[0] or "")[:50], str(row[1] or "")[:25], str(row[2] or "UNI")[:5], round(float(row[3] or 0), 10), round(float(row[4] or 0), 10), round(float(row[5] or 0), 10), round(float(row[6] or 0), 10), round(float(row[7] or 0.01), 10), 1, 1, ejercicio]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(rr, col, val)
            cell.border = st["border_all"]
            if col in (4, 5, 6, 7, 8):
                cell.number_format = '#,##0.00' if col != 8 else '"$"#,##0.00'
                cell.alignment = st["align_right"]
            if i % 2:
                cell.fill = st["fill_alt"]
    col_letters = "ABCDEFGHIJK"
    for col, w in enumerate([45, 22, 8, 12, 12, 12, 12, 14, 10, 10, 10]):
        if col < len(col_letters):
            ws.column_dimensions[col_letters[col]].width = w
    return wb


@bp.route("/reporte/inventario/kardex")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_kardex_detallado():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        producto_id = request.args.get("producto_id", "").strip()
        sucursal_id = request.args.get("sucursal_id", "").strip()
        inicio = request.args.get("inicio", date.today().replace(month=1, day=1).strftime("%Y-%m-%d"))
        fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
        prod_id = int(producto_id) if producto_id.isdigit() else None
        suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
        if not es_super:
            suc_id = suc_id if _sucursal_valida_para_empresa(cur, suc_id, emp_id) else None
        filas = []
        try:
            filas = inventario_reports_repo.listar_kardex_detallado(
                cur, emp_id, producto_id=prod_id, sucursal_id=suc_id, fecha_inicio=inicio, fecha_fin=fin
            )
        except Exception as e:
            if "costo_unitario" in str(e) or "unidad_medida" in str(e):
                flash("Ejecute: python scripts/alter_productos_costo_valuacion.py", "warning")
            else:
                flash(str(e), "danger")
        productos = _get_productos_para_filtro(cur, emp_id)
        sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        return render_template(
            "reporte_kardex_detallado.html",
            filas=filas,
            productos=productos,
            sucursales=sucursales,
            producto_id=prod_id,
            sucursal_id=suc_id,
            inicio=inicio,
            fin=fin,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/inventario/f983")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_f983():
    emp_id = _empresa_id()
    ejercicio = int(request.args.get("ejercicio", date.today().year))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = inventario_reports_repo.listar_productos_para_f983(cur, emp_id, ejercicio)
        except Exception as e:
            if "costo_unitario" in str(e) or "unidad_medida" in str(e):
                flash("Ejecute: python scripts/alter_productos_costo_valuacion.py", "warning")
            else:
                flash(str(e), "danger")
        return render_template(
            "reporte_f983.html",
            filas=filas,
            ejercicio=ejercicio,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/inventario/f983/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_f983_exportar_excel():
    emp_id = _empresa_id()
    ejercicio = int(request.args.get("ejercicio", date.today().year))
    try:
        from openpyxl import Workbook
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = inventario_reports_repo.listar_productos_para_f983(cur, emp_id, ejercicio)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    wb = _reporte_f983_excel_elegante(filas, ejercicio)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"F983_{ejercicio}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/inventario/valuacion")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_valuacion():
    emp_id = _empresa_id()
    sucursal_id_raw = request.args.get("sucursal_id", "").strip()
    suc_id = int(sucursal_id_raw) if sucursal_id_raw.isdigit() else None
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = inventario_reports_repo.listar_valuacion_inventario(cur, emp_id, sucursal_id=suc_id)
        except Exception as e:
            if "costo_unitario" in str(e):
                flash("Ejecute: python scripts/alter_productos_costo_valuacion.py", "warning")
            else:
                flash(str(e), "danger")
        sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        total_valor = sum(float(r[5] or 0) for r in filas)
        return render_template(
            "reporte_valuacion.html",
            filas=filas,
            sucursales=sucursales,
            sucursal_id=suc_id,
            total_valor=total_valor,
        )
    finally:
        cur.close()
        conn.close()


def _parsear_notas_ajuste(notas: str, referencia: str = None) -> tuple[float | None, float | None]:
    """Extrae Sistema y Físico de notas 'Sistema: X, Físico: Y'. Busca en notas y referencia."""
    import re
    textos = [t for t in (notas, referencia) if t and isinstance(t, str)]
    texto = " ".join(textos) if textos else ""
    if not texto:
        return (None, None)
    m_s = re.search(r"Sistema:\s*([0-9]+[.,]?[0-9]*)", texto, re.I)
    m_f = re.search(r"F[ií]sico:\s*([0-9]+[.,]?[0-9]*)", texto, re.I)
    try:
        def to_float(s):
            if not s:
                return None
            return float(str(s).strip().replace(",", "."))
        sist = to_float(m_s.group(1)) if m_s else None
        fis = to_float(m_f.group(1)) if m_f else None
        return (sist, fis)
    except (ValueError, TypeError):
        return (None, None)


@bp.route("/reporte/inventario/lista_compras", methods=["GET", "POST"])
@rol_requerido("GERENTE")
def reporte_lista_compras():
    """Lista de productos a comprar (stock bajo) y sugerencias de precios justos."""
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or request.form.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        umbral = 10
        try:
            u = (request.args.get("umbral") or request.form.get("umbral") or "10").strip()
            if u.replace(".", "").isdigit():
                umbral = max(1, float(u))
        except (ValueError, TypeError):
            pass
        margen_sugerido = 25.0
        try:
            m = (request.args.get("margen") or request.form.get("margen") or "25").strip()
            if m.replace(".", "").replace("-", "").isdigit():
                margen_sugerido = max(0, min(100, float(m)))
        except (ValueError, TypeError):
            pass

        hoy = date.today()
        inicio = (request.args.get("inicio") or request.form.get("inicio") or "").strip()
        fin = (request.args.get("fin") or request.form.get("fin") or "").strip()
        if not inicio:
            primer_dia_mes = hoy.replace(day=1)
            inicio = primer_dia_mes.strftime("%Y-%m-%d")
        if not fin:
            fin = hoy.strftime("%Y-%m-%d")

        ventas_por_prod = ventas_reports_repo.ventas_cantidad_por_producto(cur, emp_id, inicio, fin)

        if request.method == "POST":
            accion = (request.form.get("accion") or "").strip()
            if accion == "actualizar_precio":
                prod_id = int((request.form.get("producto_id") or "0"))
                precio_new = None
                costo_new = None
                try:
                    pv = (request.form.get("precio_nuevo") or "").strip().replace(",", ".")
                    if pv and float(pv) >= 0:
                        precio_new = float(pv)
                except (ValueError, TypeError):
                    pass
                try:
                    cv = (request.form.get("costo_nuevo") or "").strip().replace(",", ".")
                    if cv and float(cv) >= 0:
                        costo_new = float(cv)
                except (ValueError, TypeError):
                    pass
                if prod_id and (precio_new is not None or costo_new is not None):
                    if lista_compras_repo.actualizar_costo_y_precio_producto(cur, prod_id, emp_id, costo_new, precio_new):
                        if costo_new is not None and lista_compras_repo.tabla_costos_existe(cur):
                            lista_compras_repo.registrar_costo_compra(cur, prod_id, costo_new, 1, session.get("user_id"), "Actualización lista compras")
                        registrar_accion(cur, historial_usuarios_repo.EVENTO_PRODUCTO_EDITADO, f"Producto #{prod_id} (lista compras)")
                        conn.commit()
                        flash("Producto actualizado.", "success")
                    else:
                        conn.rollback()
                        flash("No se pudo actualizar el producto.", "warning")
                q = {"umbral": umbral, "margen": margen_sugerido, "inicio": inicio, "fin": fin}
                if es_super and emp_id:
                    q["empresa_id"] = emp_id
                return redirect(url_for("admin.reporte_lista_compras", **q))
            if accion == "registrar_compra":
                prod_id = int((request.form.get("producto_id") or "0"))
                costo_compra = None
                cantidad_compra = 1.0
                try:
                    cc = (request.form.get("costo_compra") or "").strip().replace(",", ".")
                    if cc and float(cc) > 0:
                        costo_compra = float(cc)
                except (ValueError, TypeError):
                    pass
                try:
                    cq = (request.form.get("cantidad_compra") or "1").strip().replace(",", ".")
                    if cq and float(cq) > 0:
                        cantidad_compra = float(cq)
                except (ValueError, TypeError):
                    pass
                if prod_id and costo_compra:
                    lista_compras_repo.registrar_costo_compra(cur, prod_id, costo_compra, cantidad_compra, session.get("user_id"), "Compra registrada")
                    lista_compras_repo.actualizar_costo_y_precio_producto(cur, prod_id, emp_id, costo_compra, None)
                    if cantidad_compra and float(cantidad_compra) > 0:
                        try:
                            cur.execute("SELECT sucursal_id FROM productos WHERE id = %s", (prod_id,))
                            rp = cur.fetchone()
                            suc_id = int(rp[0]) if rp and rp[0] else kardex_repo.primera_sucursal_empresa(cur, emp_id)
                            if suc_id:
                                kardex_repo.registrar_entrada(cur, prod_id, suc_id, float(cantidad_compra), session.get("user_id"), f"Compra lista: {cantidad_compra} uds @ ${costo_compra:.2f}")
                            else:
                                productos_repo.incrementar_stock(cur, prod_id, float(cantidad_compra))
                        except Exception:
                            productos_repo.incrementar_stock(cur, prod_id, float(cantidad_compra))
                    registrar_accion(cur, historial_usuarios_repo.EVENTO_COMPRA_REGISTRADA, f"Lista compras: costo y stock producto #{prod_id}")
                    conn.commit()
                    flash("Compra registrada. Costo y stock actualizados.", "success")
                else:
                    flash("Ingrese costo de compra válido.", "warning")
                q = {"umbral": umbral, "margen": margen_sugerido, "inicio": inicio, "fin": fin}
                if es_super and emp_id:
                    q["empresa_id"] = emp_id
                return redirect(url_for("admin.reporte_lista_compras", **q))

        filas_raw = lista_compras_repo.listar_productos_para_compra(cur, emp_id, umbral=umbral)
        filas = []
        for r in filas_raw:
            pid, codigo, nombre, stock, costo_actual, precio_actual = r[0], r[1], r[2], float(r[3] or 0), float(r[4] or 0), float(r[5] or 0)
            costo_ultimo = lista_compras_repo.ultimo_costo_desde_historial(cur, pid) or costo_actual
            costo_prom = lista_compras_repo.costo_promedio_historico(cur, pid) or costo_actual
            margen_actual = ((precio_actual - costo_actual) / costo_actual * 100) if costo_actual and costo_actual > 0 else 0
            precio_sugerido = costo_prom * (1 + margen_sugerido / 100) if costo_prom else precio_actual
            ventas_periodo = ventas_por_prod.get(pid, 0.0)
            cant_min_umbral = max(0, umbral - stock) if stock < umbral else 0
            cant_sugerida = max(ventas_periodo, cant_min_umbral)
            filas.append({
                "id": pid,
                "codigo": codigo,
                "nombre": nombre,
                "stock": stock,
                "ventas_periodo": ventas_periodo,
                "cant_sugerida": cant_sugerida,
                "costo_actual": costo_actual,
                "costo_ultimo": costo_ultimo,
                "costo_promedio": costo_prom,
                "precio_actual": precio_actual,
                "margen_actual": margen_actual,
                "precio_sugerido": precio_sugerido,
                "margen_sugerido": margen_sugerido,
            })
        empresas = []
        if es_super:
            try:
                empresas = empresas_repo.listar_empresas(cur) or []
            except Exception:
                empresas = []
        return render_template(
            "reporte_lista_compras.html",
            filas=filas,
            umbral=umbral,
            margen_sugerido=margen_sugerido,
            inicio=inicio,
            fin=fin,
            empresas=empresas,
            emp_id=emp_id,
            es_super=es_super,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/inventario/movimientos", methods=["GET", "POST"])
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_movimientos():
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        emp_id = _empresa_id()
        req = request.args if request.method == "GET" else request.form
        if es_super:
            emp_raw = (req.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)

        if request.method == "POST" and request.form.get("accion") == "agregar_ajuste":
            prod_raw = (request.form.get("producto_id") or "").strip()
            suc_raw = (request.form.get("sucursal_ajuste") or "").strip()
            fisico_raw = (request.form.get("cantidad_fisica") or "").strip().replace(",", ".")
            ref = (request.form.get("referencia_ajuste") or "").strip()[:80]
            if prod_raw.isdigit() and fisico_raw:
                try:
                    prod_id_form = int(prod_raw)
                    fisico = float(fisico_raw)
                    suc_id_form = int(suc_raw) if suc_raw.isdigit() else None
                    if not _acceso_producto_inventario(cur, prod_id_form, emp_id, es_super):
                        flash("Producto no encontrado o sin permiso.", "danger")
                    else:
                        stock_sist = inventario_reports_repo.get_stock_producto(cur, prod_id_form, suc_id_form)
                        fecha_txt = (request.form.get("fecha_ajuste") or date.today().isoformat())[:10]
                        referencia = ref or f"Ajuste manual {fecha_txt}"
                        justif = (request.form.get("justificacion_ajuste") or "").strip()
                        min_j = kardex_repo.MIN_CARACTERES_JUSTIFICACION_AJUSTE
                        motivo_aj = (request.form.get("motivo_ajuste") or "").strip().upper()
                        motivo_ap = None
                        ok_go = True
                        if len(justif) < min_j:
                            flash(f"Justificación obligatoria (mín. {min_j} caracteres): detalle qué se encontró.", "warning")
                            ok_go = False
                        elif fisico < stock_sist:
                            if motivo_aj not in kardex_repo.MOTIVOS_AJUSTE_SALIDA:
                                flash(
                                    "Si el conteo físico es menor que el sistema, elija el motivo: "
                                    "merma operativa, avería, ajuste de inventario o faltante.",
                                    "warning",
                                )
                                ok_go = False
                            else:
                                motivo_ap = motivo_aj
                        if ok_go:
                            cur.execute(
                                """
                                SELECT COALESCE(NULLIF(costo_unitario, 0), NULLIF(precio_unitario, 0), 0)
                                FROM productos WHERE id = %s
                                """,
                                (prod_id_form,),
                            )
                            rwc = cur.fetchone()
                            cu = float(rwc[0] or 0) if rwc else 0.0
                            absd = abs(fisico - stock_sist)
                            impacto_est = absd * cu if cu > 0 else 0.0
                            sup_u = (request.form.get("supervisor_inventario_usuario") or "").strip()
                            sup_p = (request.form.get("supervisor_inventario_clave") or "").strip()
                            need_sup = (
                                impacto_est > kardex_repo.UMBRAL_IMPACTO_AJUSTE_USD
                                or absd > kardex_repo.UMBRAL_CANTIDAD_AJUSTE_UMB
                            )
                            if need_sup and not _rol_ajuste_inventario_sin_supervisor():
                                if not _validar_supervisor_ajuste_inventario(cur, sup_u, sup_p):
                                    flash(
                                        "Ajuste elevado: requiere usuario y clave de Gerente o Administrador.",
                                        "danger",
                                    )
                                    ok_go = False
                        if ok_go:
                            if abs(fisico - stock_sist) < 0.0001:
                                flash("No hay diferencia entre el sistema y el conteo físico; no se registró movimiento.", "info")
                            else:
                                kardex_repo.ajustar_por_conteo_fisico(
                                    cur, prod_id_form, stock_sist, fisico, suc_id_form,
                                    session.get("user_id"), referencia,
                                    motivo_ajuste=motivo_ap,
                                    comentario_justificacion=justif,
                                )
                                registrar_accion(
                                    cur, historial_usuarios_repo.EVENTO_AJUSTE_INVENTARIO,
                                    f"Producto #{prod_id_form}: {stock_sist}→{fisico} | {referencia[:40]}",
                                )
                                conn.commit()
                                flash("Ajuste aplicado correctamente. Stock actualizado en Kardex.", "success")
                except ValueError as e:
                    flash(str(e), "danger")
                except Exception as e:
                    conn.rollback()
                    flash(f"Error: {e}", "danger")
            else:
                flash("Seleccione producto e ingrese cantidad física.", "warning")
            from urllib.parse import urlencode
            q = {}
            if es_super and emp_id:
                q["empresa_id"] = emp_id
            if suc_raw and suc_raw.isdigit():
                q["sucursal_id"] = suc_raw
            if req.get("inicio"):
                q["inicio"] = req.get("inicio")
            if req.get("fin"):
                q["fin"] = req.get("fin")
            redir = url_for("admin.reporte_movimientos") + ("?" + urlencode(q) if q else "")
            return redirect(redir)

        producto_id_raw = req.get("producto_id", "").strip()
        sucursal_id_raw = req.get("sucursal_id", "").strip()
        inicio_param = req.get("inicio", "").strip()
        fin_param = req.get("fin", "").strip()
        hoy = date.today()
        primer_dia = hoy.replace(day=1).strftime("%Y-%m-%d")
        inicio = inicio_param if inicio_param else primer_dia
        fin = fin_param if fin_param else hoy.strftime("%Y-%m-%d")

        prod_id = int(producto_id_raw) if producto_id_raw.isdigit() else None
        suc_id = int(sucursal_id_raw) if sucursal_id_raw.isdigit() else None
        if not es_super and suc_id:
            suc_id = suc_id if _sucursal_valida_para_empresa(cur, suc_id, emp_id) else suc_id

        filas_raw = inventario_reports_repo.listar_movimientos_global(
            cur, emp_id, producto_id=prod_id, sucursal_id=suc_id, fecha_inicio=inicio, fecha_fin=fin
        )
        filas = []
        total_cant = total_sist = total_ajus = total_diff = 0.0
        for r in filas_raw:
            sist, ajus = None, None
            if (r[2] or "").upper() in ("AJUSTE_ENTRADA", "AJUSTE_SALIDA"):
                sist, ajus = _parsear_notas_ajuste(
                    r[8] if len(r) > 8 else None,
                    r[9] if len(r) > 9 else None,
                )
            diff = (ajus - sist) if (sist is not None and ajus is not None) else None
            row = list(r)
            if len(row) > 1 and row[1]:
                try:
                    dt = row[1]
                    if hasattr(dt, "strftime"):
                        row[1] = dt.strftime("%d/%m/%Y %H:%M")
                    elif isinstance(dt, str) and len(dt) >= 10:
                        row[1] = dt[:16].replace("T", " ") if "T" in dt else dt[:16]
                except Exception:
                    pass
            filas.append((*row, sist, ajus, diff))
            total_cant += float(r[5] or 0)
            if sist is not None:
                total_sist += sist
            if ajus is not None:
                total_ajus += ajus
            if diff is not None:
                total_diff += diff

        productos = _get_productos_para_filtro(cur, emp_id)
        sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        empresas = empresas_repo.listar_empresas(cur) or [] if es_super else []
        productos_con_stock = inventario_reports_repo.listar_productos_para_conteo(cur, emp_id, suc_id)
        return render_template(
            "reporte_movimientos.html",
            filas=filas,
            productos=productos,
            productos_con_stock=productos_con_stock,
            sucursales=sucursales,
            producto_id=prod_id,
            sucursal_id=suc_id,
            inicio=inicio,
            fin=fin,
            es_superadmin=es_super,
            empresas=empresas,
            empresa_id=emp_id,
            total_cant=total_cant,
            total_sist=total_sist,
            total_ajus=total_ajus,
            total_diff=total_diff,
            fecha_hoy=date.today().isoformat(),
            etiqueta_motivo=kardex_repo.etiqueta_motivo_ajuste,
            umbral_impacto_ajuste_usd=kardex_repo.UMBRAL_IMPACTO_AJUSTE_USD,
            umbral_cantidad_ajuste_umb=kardex_repo.UMBRAL_CANTIDAD_AJUSTE_UMB,
            min_caracteres_justificacion_ajuste=kardex_repo.MIN_CARACTERES_JUSTIFICACION_AJUSTE,
        )
    finally:
        cur.close()
        conn.close()


def _exportar_excel_inventario(titulo: str, headers: list, filas: list, cab: dict, num_cols: int, col_formato_moneda: list = None):
    """Genera Excel con formato elegante (cabecera, estilos, total)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    ws.cell(r, 1, titulo).font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    ws.row_dimensions[r].height = 26
    for txt in (f"Período: {cab['periodo']}", f"Empresa: {cab['empresa']}", f"Sucursal: {cab['sucursal']}", f"Generado: {cab['generado']}"):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(r, 1, txt).font = st["font_meta"]
    r += 2
    hdr = r
    for col, h in enumerate(headers, 1):
        c = ws.cell(hdr, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    for i, row in enumerate(filas):
        rr = hdr + 1 + i
        for col, val in enumerate(row, 1):
            cell = ws.cell(rr, col, val)
            cell.border = st["border_all"]
            if col_formato_moneda and col in col_formato_moneda and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = st["align_right"]
            if i % 2:
                cell.fill = st["fill_alt"]
    for col in range(1, num_cols + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 14
    return wb


def _exportar_pdf_inventario(titulo: str, subtitulo: str, headers: list, filas: list, cab: dict, col_fracs: list):
    """Genera PDF con formato elegante (reportlab)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    styles, colors = _estilos_pdf_elegante()
    buf = BytesIO()
    doc_w = letter[0]
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=0.65*inch, leftMargin=0.65*inch, topMargin=0.55*inch, bottomMargin=0.55*inch)
    elements = [
        Paragraph(html.escape(titulo), styles["Title"]),
        Paragraph(f"<b>{html.escape(subtitulo)}</b>", styles["RptSubtitle"]),
        Spacer(1, 0.12*inch),
        Paragraph(f"<b>Empresa:</b> {html.escape(cab['empresa'])}", styles["RptMeta"]),
        Paragraph(f"<b>Sucursal:</b> {html.escape(cab['sucursal'])}", styles["RptMeta"]),
        Paragraph(f"<b>Generado:</b> {html.escape(cab['generado'])}", styles["RptMeta"]),
        Spacer(1, 0.28*inch),
    ]
    data = [headers] + [[str(x) for x in r] for r in filas]
    usable_w = doc_w - 1.3*inch
    cw = [usable_w * f for f in col_fracs]
    t = Table(data, colWidths=cw, repeatRows=1)
    last_r = len(data) - 1
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
    ]
    if last_r > 1:
        style_cmds.append(("ROWBACKGROUNDS", (0, 1), (-1, last_r), [colors.white, colors.HexColor("#f8fafc")]))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(f"<i>AZ DIGITAL — {html.escape(str(session.get('username') or ''))}</i>", styles["RptMeta"]))
    doc.build(elements)
    buf.seek(0)
    return buf


@bp.route("/reporte/inventario/kardex/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_kardex_exportar_excel():
    emp_id = _empresa_id()
    producto_id = request.args.get("producto_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    inicio = request.args.get("inicio", date.today().replace(month=1, day=1).strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        prod_id = int(producto_id) if producto_id.isdigit() else None
        suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
        filas = inventario_reports_repo.listar_kardex_detallado(cur, emp_id, producto_id=prod_id, sucursal_id=suc_id, fecha_inicio=inicio, fecha_fin=fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500
    periodo_txt = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    cab = _datos_cabecera_inventario(periodo_txt)
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex Detallado"
    r, num_cols = 1, 9
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    ws.cell(r, 1, "Kardex Detallado — Art. 142 y 142-A CT").font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    ws.row_dimensions[r].height = 26
    for txt in (f"Período: {periodo_txt}", f"Empresa: {cab['empresa']}", f"Sucursal: {cab['sucursal']}", f"Generado: {cab['generado']}"):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(r, 1, txt).font = st["font_meta"]
    r += 2
    headers = ["Fecha", "Tipo", "Entrada", "Salida", "Costo Unit.", "Saldo", "Origen", "Destino", "Referencia"]
    hdr = r + 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(hdr, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    for i, row in enumerate(filas):
        rr = hdr + 1 + i
        vals = [str(row[0]), str(row[1]), float(row[3] or 0), float(row[4] or 0), float(row[5] or 0), float(row[6] or 0), str(row[8] or ""), str(row[9] or ""), str(row[7] or "")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(rr, col, val)
            cell.border = st["border_all"]
            if col in (3, 4, 5, 6):
                cell.number_format = '#,##0.00' if col != 5 else '"$"#,##0.00'
                cell.alignment = st["align_right"]
            if i % 2:
                cell.fill = st["fill_alt"]
    widths = [18, 12, 10, 10, 12, 10, 18, 18, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"kardex_detallado_{inicio}_{fin}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/inventario/kardex/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_kardex_exportar_pdf():
    emp_id = _empresa_id()
    producto_id = request.args.get("producto_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    inicio = request.args.get("inicio", date.today().replace(month=1, day=1).strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        prod_id = int(producto_id) if producto_id.isdigit() else None
        suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
        filas = inventario_reports_repo.listar_kardex_detallado(cur, emp_id, producto_id=prod_id, sucursal_id=suc_id, fecha_inicio=inicio, fecha_fin=fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return "<p>Instala reportlab: pip install reportlab</p>", 500
    periodo_txt = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    cab = _datos_cabecera_inventario(periodo_txt)
    headers = ["Fecha", "Tipo", "Ent.", "Sal.", "Costo", "Saldo", "Origen/Dest.", "Ref."]
    data_rows = []
    for row in filas:
        data_rows.append([str(row[0])[:16], str(row[1])[:8], f"{float(row[3] or 0):.1f}", f"{float(row[4] or 0):.1f}", f"${float(row[5] or 0):.2f}", f"{float(row[6] or 0):.1f}", (str(row[8] or "") + " → " + str(row[9] or ""))[:25], (str(row[7] or ""))[:20]])
    buf = _exportar_pdf_inventario("Kardex Detallado", f"Período: {periodo_txt}", headers, data_rows, cab, [0.14, 0.08, 0.07, 0.07, 0.10, 0.08, 0.22, 0.24])
    return send_file(buf, as_attachment=True, download_name=f"kardex_detallado_{inicio}_{fin}.pdf", mimetype="application/pdf")


@bp.route("/reporte/inventario/valuacion/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_valuacion_exportar_excel():
    emp_id = _empresa_id()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = inventario_reports_repo.listar_valuacion_inventario(cur, emp_id, sucursal_id=suc_id)
        sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
    except Exception:
        filas = []
        sucursales = []
    finally:
        cur.close()
        conn.close()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500
    suc_nombre = "Todas (consolidado)"
    if suc_id and sucursales:
        for s in sucursales:
            if s[0] == suc_id:
                suc_nombre = s[1]
                break
    cab = _datos_cabecera_inventario("Valuación al cierre", suc_nombre)
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuacion"
    r, num_cols = 1, 7
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    ws.cell(r, 1, "Valuación de Inventarios — NIC 2").font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    ws.row_dimensions[r].height = 26
    for txt in (f"Sucursal: {cab['sucursal']}", f"Empresa: {cab['empresa']}", f"Generado: {cab['generado']}"):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(r, 1, txt).font = st["font_meta"]
    r += 2
    headers = ["Producto", "Código", "Unidad", "Cantidad", "Costo Unit.", "Valor Total", "Método"]
    hdr = r + 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(hdr, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    total_valor = 0
    for i, row in enumerate(filas):
        rr = hdr + 1 + i
        v5 = float(row[5] or 0)
        total_valor += v5
        vals = [str(row[0] or "")[:50], str(row[1] or ""), str(row[2] or "UNI"), float(row[3] or 0), float(row[4] or 0), v5, str(row[6] or "PROMEDIO")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(rr, col, val)
            cell.border = st["border_all"]
            if col in (4, 5, 6):
                cell.number_format = '#,##0.00' if col == 4 else '"$"#,##0.00'
                cell.alignment = st["align_right"]
            if i % 2:
                cell.fill = st["fill_alt"]
    tot_row = hdr + 1 + len(filas)
    ws.cell(tot_row, 1, "")
    for c in range(2, 6):
        ws.cell(tot_row, c, "")
    ws.cell(tot_row, 5, "TOTAL").font = st["font_total"]
    ws.cell(tot_row, 6, total_valor).number_format = '"$"#,##0.00'
    ws.cell(tot_row, 6).font = st["font_total"]
    ws.cell(tot_row, 7, "")
    for col in range(1, 8):
        cell = ws.cell(tot_row, col)
        cell.border = st["border_all"]
        cell.fill = st["fill_total"]
    widths = [35, 14, 8, 10, 14, 14, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="valuacion_inventario.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/inventario/valuacion/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_valuacion_exportar_pdf():
    emp_id = _empresa_id()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = inventario_reports_repo.listar_valuacion_inventario(cur, emp_id, sucursal_id=suc_id)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    try:
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return "<p>Instala reportlab: pip install reportlab</p>", 500
    db2 = ConexionDB()
    conn2 = psycopg2.connect(**db2.config)
    cur2 = conn2.cursor()
    sucursales = sucursales_repo.listar_sucursales_min(cur2, empresa_id=emp_id) or []
    cur2.close()
    conn2.close()
    suc_nombre = "Todas (consolidado)"
    if suc_id and sucursales:
        for s in sucursales:
            if s[0] == suc_id:
                suc_nombre = s[1]
                break
    cab = _datos_cabecera_inventario("Valuación al cierre", suc_nombre)
    headers = ["Producto", "Código", "Unid.", "Cant.", "Costo", "Valor", "Método"]
    data = [headers]
    total = 0
    for row in filas:
        v5 = float(row[5] or 0)
        total += v5
        data.append([str(row[0] or "")[:30], str(row[1] or "")[:12], str(row[2] or "UNI"), f"{float(row[3] or 0):.1f}", f"${float(row[4] or 0):.2f}", f"${v5:,.2f}", str(row[6] or "PROM")[:4]])
    data.append(["", "", "", "", "TOTAL", f"${total:,.2f}", ""])
    buf = _exportar_pdf_inventario("Valuación de Inventarios", f"Sucursal: {suc_nombre}", headers, data[1:], cab, [0.28, 0.12, 0.08, 0.08, 0.12, 0.14, 0.08])
    return send_file(buf, as_attachment=True, download_name="valuacion_inventario.pdf", mimetype="application/pdf")


@bp.route("/reporte/inventario/movimientos/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_movimientos_exportar_excel():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if _es_superadmin_db(cur):
            emp_raw = request.args.get("empresa_id", "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
    finally:
        cur.close()
        conn.close()
    producto_id = request.args.get("producto_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    hoy = date.today()
    inicio = request.args.get("inicio", "").strip() or hoy.replace(day=1).strftime("%Y-%m-%d")
    fin = request.args.get("fin", "").strip() or hoy.strftime("%Y-%m-%d")
    db2 = ConexionDB()
    conn2 = psycopg2.connect(**db2.config)
    cur2 = conn2.cursor()
    try:
        prod_id = int(producto_id) if producto_id.isdigit() else None
        suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
        filas_raw = inventario_reports_repo.listar_movimientos_global(cur2, emp_id, producto_id=prod_id, sucursal_id=suc_id, fecha_inicio=inicio, fecha_fin=fin)
        filas = []
        for row in filas_raw:
            sist, ajus = None, None
            if len(row) > 8 and row[2] in ("AJUSTE_ENTRADA", "AJUSTE_SALIDA"):
                sist, ajus = _parsear_notas_ajuste(row[8] if len(row) > 8 else None, row[9] if len(row) > 9 else None)
            diff = (ajus - sist) if (sist is not None and ajus is not None) else None
            filas.append((*row, sist, ajus, diff))
    except Exception:
        filas = []
    finally:
        cur2.close()
        conn2.close()
    try:
        from openpyxl import Workbook
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500
    periodo_txt = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    cab = _datos_cabecera_inventario(periodo_txt, emp_id=emp_id)
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    r, num_cols = 1, 13
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    ws.cell(r, 1, "Movimientos por Producto").font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    ws.row_dimensions[r].height = 26
    for txt in (f"Período: {periodo_txt}", f"Empresa: {cab['empresa']}", f"Sucursal: {cab['sucursal']}", f"Generado: {cab['generado']}"):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(r, 1, txt).font = st["font_meta"]
    r += 2
    headers = ["Fecha", "Tipo", "Producto", "Código", "Cant.", "Sistema", "Físico", "Diferencia", "Origen", "Destino", "Motivo", "Notas", "Ref."]
    hdr = r + 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(hdr, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    total_cant = total_sist = total_ajus = total_diff = 0.0
    for i, row in enumerate(filas):
        rr = hdr + 1 + i
        sist = row[11] if len(row) > 11 and row[11] is not None else ""
        ajus = row[12] if len(row) > 12 and row[12] is not None else ""
        diff = row[13] if len(row) > 13 and row[13] is not None else ""
        total_cant += float(row[5] or 0)
        if isinstance(sist, (int, float)):
            total_sist += sist
        if isinstance(ajus, (int, float)):
            total_ajus += ajus
        if isinstance(diff, (int, float)):
            total_diff += diff
        diff_str = f"{diff:+.2f}" if isinstance(diff, (int, float)) and diff != 0 else ("" if diff == 0 else "—")
        motivo_txt = kardex_repo.etiqueta_motivo_ajuste(row[10] if len(row) > 10 else None)
        vals = [str(row[1]), str(row[2]), str(row[3] or "")[:35], str(row[4] or ""), float(row[5] or 0),
                f"{sist:.2f}" if isinstance(sist, (int, float)) else "—",
                f"{ajus:.2f}" if isinstance(ajus, (int, float)) else "—",
                diff_str, str(row[6] or ""), str(row[7] or ""), motivo_txt, str(row[8] or ""), str(row[9] or "")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(rr, col, val)
            cell.border = st["border_all"]
            if col in (5, 6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = st["align_right"]
            if i % 2:
                cell.fill = st["fill_alt"]
    rr_total = hdr + 1 + len(filas)
    if filas:
        ws.cell(rr_total, 1, "TOTAL").font = st["font_total"]
        ws.cell(rr_total, 5, total_cant).number_format = '#,##0.00'
        ws.cell(rr_total, 6, total_sist if total_sist else "").number_format = '#,##0.00'
        ws.cell(rr_total, 7, total_ajus if total_ajus else "").number_format = '#,##0.00'
        ws.cell(rr_total, 8, f"{total_diff:+.2f}" if total_diff else "").number_format = '#,##0.00'
        for c in range(1, num_cols + 1):
            cell = ws.cell(rr_total, c)
            cell.border = st["border_all"]
            cell.fill = st["fill_total"]
            if c >= 5 and c <= 8:
                cell.alignment = st["align_right"]
    widths = [18, 12, 28, 12, 10, 10, 10, 12, 12, 12, 20, 22, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = min(w, 50)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"movimientos_{inicio}_{fin}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/inventario/movimientos/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_movimientos_exportar_pdf():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if _es_superadmin_db(cur):
            emp_raw = request.args.get("empresa_id", "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
    finally:
        cur.close()
        conn.close()
    producto_id = request.args.get("producto_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    hoy = date.today()
    inicio = request.args.get("inicio", "").strip() or hoy.replace(day=1).strftime("%Y-%m-%d")
    fin = request.args.get("fin", "").strip() or hoy.strftime("%Y-%m-%d")
    db2 = ConexionDB()
    conn2 = psycopg2.connect(**db2.config)
    cur2 = conn2.cursor()
    try:
        prod_id = int(producto_id) if producto_id.isdigit() else None
        suc_id = int(sucursal_id) if sucursal_id.isdigit() else None
        filas_raw = inventario_reports_repo.listar_movimientos_global(cur2, emp_id, producto_id=prod_id, sucursal_id=suc_id, fecha_inicio=inicio, fecha_fin=fin)
        filas = []
        for row in filas_raw:
            sist, ajus = None, None
            if len(row) > 8 and row[2] in ("AJUSTE_ENTRADA", "AJUSTE_SALIDA"):
                sist, ajus = _parsear_notas_ajuste(row[8] if len(row) > 8 else None, row[9] if len(row) > 9 else None)
            diff = (ajus - sist) if (sist is not None and ajus is not None) else None
            filas.append((*row, sist, ajus, diff))
    except Exception:
        filas = []
    finally:
        cur2.close()
        conn2.close()
    try:
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return "<p>Instala reportlab: pip install reportlab</p>", 500
    periodo_txt = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    cab = _datos_cabecera_inventario(periodo_txt, emp_id=emp_id)
    headers = ["Fecha", "Tipo", "Producto", "Cant.", "Sist.", "Físico", "Dif.", "Origen", "Destino", "Motivo"]
    data = [headers]
    total_cant = total_sist = total_ajus = total_diff = 0.0
    for row in filas:
        cant = float(row[5] or 0)
        sist = row[11] if len(row) > 11 and row[11] is not None else None
        ajus = row[12] if len(row) > 12 and row[12] is not None else None
        diff = row[13] if len(row) > 13 and row[13] is not None else None
        total_cant += cant
        if sist is not None:
            total_sist += sist
        if ajus is not None:
            total_ajus += ajus
        if diff is not None:
            total_diff += diff
        sist_str = f"{sist:.1f}" if sist is not None else "—"
        ajus_str = f"{ajus:.1f}" if ajus is not None else "—"
        diff_str = f"{diff:+.1f}" if diff is not None else "—"
        motivo_sh = (kardex_repo.etiqueta_motivo_ajuste(row[10] if len(row) > 10 else None) or "")[:16]
        data.append([str(row[1])[:14], str(row[2])[:10], str(row[3] or "")[:20], f"{cant:.1f}", sist_str, ajus_str, diff_str, str(row[6] or "")[:10], str(row[7] or "")[:10], motivo_sh])
    if filas:
        data.append(["TOTAL", "", "", f"{total_cant:.1f}",
                    f"{total_sist:.1f}" if total_sist else "—",
                    f"{total_ajus:.1f}" if total_ajus else "—",
                    f"{total_diff:+.1f}" if total_diff else "—", "", "", ""])
    buf = _exportar_pdf_inventario("Movimientos por Producto", f"Período: {periodo_txt}", headers, data[1:], cab, [0.10, 0.08, 0.18, 0.06, 0.06, 0.06, 0.06, 0.11, 0.11, 0.12])
    return send_file(buf, as_attachment=True, download_name=f"movimientos_{inicio}_{fin}.pdf", mimetype="application/pdf")


@bp.route("/reporte/inventario/mermas-ajustes", methods=["GET"])
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_mermas_ajustes():
    """Pérdidas por merma, avería y faltantes (Kardex AJUSTE_SALIDA con motivo y costo)."""
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        emp_id = _empresa_id()
        if es_super:
            emp_raw = request.args.get("empresa_id", "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        suc_raw = request.args.get("sucursal_id", "").strip()
        suc_id = int(suc_raw) if suc_raw.isdigit() else None
        hoy = date.today()
        inicio = request.args.get("inicio", "").strip() or hoy.replace(day=1).strftime("%Y-%m-%d")
        fin = request.args.get("fin", "").strip() or hoy.strftime("%Y-%m-%d")
        solo_perdidas = request.args.get("incluir_sobrantes", "").strip() != "1"
        usr_raw = request.args.get("usuario_id", "").strip()
        usuario_filtro_id = int(usr_raw) if usr_raw.isdigit() else None
        filas, totales = inventario_reports_repo.listar_reporte_mermas_ajustes(
            cur,
            emp_id,
            inicio,
            fin,
            sucursal_id=suc_id,
            solo_perdidas=solo_perdidas,
            usuario_id=usuario_filtro_id,
        )
        empresas = empresas_repo.listar_empresas(cur) or [] if es_super else []
        sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        usuarios_empresa = usuarios_repo.listar_usuarios(cur, empresa_id=emp_id) or []
        return render_template(
            "reporte_mermas_ajustes.html",
            filas=filas,
            totales=totales,
            inicio=inicio,
            fin=fin,
            empresa_id=emp_id,
            sucursal_id=suc_id,
            usuario_id=usuario_filtro_id,
            es_superadmin=es_super,
            empresas=empresas,
            sucursales=sucursales,
            usuarios_empresa=usuarios_empresa,
            solo_perdidas=solo_perdidas,
            etiqueta_motivo=kardex_repo.etiqueta_motivo_ajuste,
        )
    except Exception as e:
        flash(
            f"Error al cargar mermas/ajustes. Si acaba de actualizar el sistema, ejecute: "
            f"python scripts/alter_kardex_motivo_ajuste.py — {e}",
            "danger",
        )
        return redirect(url_for("admin.reporte_inventario_index"))
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/inventario/f983/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR", "BODEGUERO")
def reporte_f983_exportar_pdf():
    emp_id = _empresa_id()
    ejercicio = int(request.args.get("ejercicio", date.today().year))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = inventario_reports_repo.listar_productos_para_f983(cur, emp_id, ejercicio)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    try:
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return "<p>Instala reportlab: pip install reportlab</p>", 500
    cab = _datos_cabecera_inventario(f"Ejercicio fiscal {ejercicio}")
    headers = ["Denominación", "Código", "Unid.", "Inv. Inic.", "Compras", "Ventas", "Inv. Final", "Costo Unit."]
    data = [headers]
    for row in filas:
        data.append([str(row[0] or "")[:40], str(row[1] or "")[:15], str(row[2] or "UNI"), f"{float(row[3] or 0):.1f}", f"{float(row[4] or 0):.1f}", f"{float(row[5] or 0):.1f}", f"{float(row[6] or 0):.1f}", f"${float(row[7] or 0):.2f}"])
    buf = _exportar_pdf_inventario("Informe F-983 — Inventario Físico", f"Ejercicio: {ejercicio}", headers, data[1:], cab, [0.22, 0.12, 0.06, 0.10, 0.10, 0.10, 0.10, 0.10])
    return send_file(buf, as_attachment=True, download_name=f"F983_{ejercicio}.pdf", mimetype="application/pdf")


# --- Reportes de Facturación y Ventas ---


@bp.route("/reporte/facturacion/libro_iva")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().replace(day=1).strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = ventas_reports_repo.listar_libro_iva(cur, emp_id, inicio, fin)
        except Exception:
            pass
        cab = _datos_cabecera_reporte_ventas(inicio, fin)
        total_gravado = sum(float(r[5] or 0) for r in filas)
        total_iva = sum(float(r[6] or 0) for r in filas)
        total_gen = sum(float(r[7] or 0) for r in filas)
        total_retencion_iva = sum(float(r[9] or 0) for r in filas if len(r) > 9)
        return render_template(
            "reporte_libro_iva.html",
            filas=filas,
            inicio=inicio,
            fin=fin,
            cab=cab,
            total_gravado=total_gravado,
            total_iva=total_iva,
            total_gen=total_gen,
            total_retencion_iva=total_retencion_iva,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/facturacion/ventas_producto")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_ventas_producto():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().replace(day=1).strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = ventas_reports_repo.listar_ventas_por_producto(cur, emp_id, inicio, fin)
        except Exception:
            pass
        total_ventas = sum(float(r[3] or 0) for r in filas)
        cab = _datos_cabecera_reporte_ventas(inicio, fin)
        return render_template("reporte_ventas_producto.html", filas=filas, inicio=inicio, fin=fin, cab=cab, total_ventas=total_ventas)
    finally:
        cur.close()
        conn.close()


@bp.route("/contingencia")
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def contingencia():
    """Plan de Contingencia MH: DTEs pendientes, evento y sincronización."""
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT v.id, TO_CHAR(v.fecha_registro, 'DD/MM/YYYY HH24:MI'),
                   COALESCE(v.tipo_comprobante, 'TICKET'), v.total_pagar,
                   COALESCE(v.codigo_generacion, ''), COALESCE(v.numero_control, ''),
                   COALESCE(v.estado_dte, 'RESPALDO'), COALESCE(v.causa_contingencia, 0)
            FROM ventas v
            WHERE (v.empresa_id IS NULL OR v.empresa_id = %s)
              AND COALESCE(v.estado, 'ACTIVO') = 'ACTIVO'
              AND COALESCE(v.estado_dte, 'RESPALDO') IN ('CONTINGENCIA', 'PENDIENTE_TRANSMISION')
            ORDER BY v.fecha_registro ASC
            """,
            (emp_id,),
        )
        dte_pendientes = cur.fetchall() or []
        cur.execute(
            """
            SELECT id, TO_CHAR(fecha_inicio, 'DD/MM/YYYY HH24:MI'), TO_CHAR(fecha_fin, 'DD/MM/YYYY HH24:MI'),
                   causa, descripcion_causa, estado, TO_CHAR(fecha_registro, 'DD/MM/YYYY HH24:MI')
            FROM evento_contingencia
            WHERE empresa_id = %s
            ORDER BY fecha_registro DESC
            LIMIT 20
            """,
            (emp_id,),
        )
        eventos = cur.fetchall() or []
    finally:
        cur.close()
        conn.close()
    from azdigital.utils.mh_utils import CAUSAS_CONTINGENCIA, check_mh_online
    mh_online = check_mh_online()
    return render_template(
        "contingencia.html",
        dte_pendientes=dte_pendientes,
        eventos=eventos,
        causas=CAUSAS_CONTINGENCIA,
        mh_online=mh_online,
    )


@bp.route("/contingencia/registrar_evento", methods=["POST"])
@rol_requerido("GERENTE", "CONTADOR")
def contingencia_registrar_evento():
    """Registra Evento de Contingencia (tipo 1) para enviar al MH."""
    emp_id = _empresa_id()
    fecha_inicio = (request.form.get("fecha_inicio") or "").strip()
    fecha_fin = (request.form.get("fecha_fin") or "").strip()
    causa = int(request.form.get("causa") or 1)
    descripcion = (request.form.get("descripcion") or "").strip()[:100]
    if not fecha_inicio or not fecha_fin:
        flash("Fechas de inicio y fin son requeridas.", "warning")
        return redirect(url_for("admin.contingencia"))
    try:
        from datetime import datetime as dt
        fi = dt.strptime(fecha_inicio, "%Y-%m-%dT%H:%M")
        ff = dt.strptime(fecha_fin, "%Y-%m-%dT%H:%M")
    except ValueError:
        flash("Formato de fecha inválido. Use YYYY-MM-DDTHH:MM", "warning")
        return redirect(url_for("admin.contingencia"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO evento_contingencia (empresa_id, fecha_inicio, fecha_fin, causa, descripcion_causa, estado)
            VALUES (%s, %s, %s, %s, %s, 'PENDIENTE')
            """,
            (emp_id, fi, ff, causa, descripcion or None),
        )
        conn.commit()
        flash("Evento de contingencia registrado. Debe transmitirse al MH en las próximas 24 horas.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error al registrar evento: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.contingencia"))


@bp.route("/contingencia/sincronizar", methods=["POST"])
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def contingencia_sincronizar():
    """
    Mismo proceso que AgilDTE (Configuración → «Desactivar y enviar pendientes» / procesar contingencia completa):
    POST /api/empresas/{id}/procesar-contingencia-completa/ — reporte MH + envío de facturas PendienteEnvio en el servidor central.
    """
    emp_id = _empresa_id()
    raw_tipo = (request.form.get("tipo_contingencia") or "").strip()
    motivo = (request.form.get("motivo") or "").strip() or None
    try:
        tipo_c = int(raw_tipo) if raw_tipo.isdigit() else None
        if tipo_c is not None and tipo_c not in (1, 2, 3, 4, 5):
            tipo_c = None
    except (TypeError, ValueError):
        tipo_c = None

    from azdigital.integration.agildte_contingencia import procesar_contingencia_completa_remoto

    out = procesar_contingencia_completa_remoto(
        emp_id,
        tipo_contingencia=tipo_c,
        motivo=motivo,
    )
    if out.get("ok"):
        data = out.get("respuesta") or {}
        msg = (data.get("mensaje") or "Proceso de contingencia completado en AgilDTE.").strip()
        rs = data.get("resumen_envio") or {}
        if isinstance(rs, dict) and rs.get("total") is not None:
            msg += (
                f" — Facturas: total {rs.get('total')}, aceptadas {rs.get('aceptadas')}, "
                f"rechazadas {rs.get('rechazadas')}, errores {rs.get('errores')}."
            )
        flash(msg, "success")
    else:
        det = out.get("detalle")
        if isinstance(det, dict):
            det_txt = det.get("detail") or det.get("mensaje") or str(det)
        else:
            det_txt = det or out.get("mensaje") or "Error al llamar a AgilDTE."
        flash(f"No se completó la sincronización: {det_txt}", "danger")
    return redirect(url_for("admin.contingencia"))


@bp.route("/reporte/facturacion/documentos_anulados")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_documentos_anulados():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", "")
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = ventas_reports_repo.listar_documentos_anulados(cur, emp_id, fecha_inicio=inicio or None, fecha_fin=fin or None)
        except Exception:
            pass
        cab = _datos_cabecera_reporte_ventas(inicio or "—", fin or "—")
        total = sum(float(r[2] or 0) for r in filas)
        return render_template("reporte_documentos_anulados.html", filas=filas, inicio=inicio, fin=fin, cab=cab, total=total)
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/facturacion/comprobante_invalidacion/<int:venta_id>")
@rol_requerido("GERENTE", "CONTADOR")
def imprimir_comprobante_invalidacion(venta_id: int):
    """Comprobante de Invalidación para justificar contablemente la anulación."""
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT v.id, TO_CHAR(v.fecha_registro, 'DD/MM/YYYY HH24:MI'), v.total_pagar,
                   COALESCE(v.cliente_nombre, '—'), COALESCE(v.tipo_comprobante, 'TICKET'),
                   COALESCE(v.motivo_anulacion, '—'),
                   COALESCE(u.username, '—'),
                   TO_CHAR(v.fecha_anulacion, 'DD/MM/YYYY HH24:MI'),
                   COALESCE(v.codigo_generacion, '')
            FROM ventas v
            LEFT JOIN usuarios u ON u.id = v.usuario_anulo_id
            WHERE v.id = %s AND (v.empresa_id IS NULL OR v.empresa_id = %s) AND COALESCE(v.estado, 'ACTIVO') = 'ANULADO'
            """,
            (venta_id, emp_id),
        )
        row = cur.fetchone()
        if not row:
            return "Documento no encontrado o no anulado", 404
        return render_template(
            "comprobante_invalidacion_print.html",
            venta_id=venta_id,
            fecha_registro=row[1],
            total=row[2],
            cliente_nombre=row[3],
            tipo_comprobante=row[4],
            motivo_anulacion=row[5],
            usuario_anulo=row[6],
            fecha_anulacion=row[7],
            codigo_generacion=row[8] or "—",
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/facturacion/corte_caja")
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def reporte_corte_caja():
    """Corte de Caja (X/Z): ventas por tipo DTE, método de pago e impuestos."""
    emp_id = _empresa_id()
    fecha = request.args.get("fecha", date.today().strftime("%Y-%m-%d"))
    usuario_id = request.args.get("usuario_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    uid = int(usuario_id) if usuario_id.isdigit() else None
    sid = int(sucursal_id) if sucursal_id.isdigit() else None
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        datos = cierre_caja_repo.obtener_datos_corte(cur, emp_id, fecha, usuario_id=uid, sucursal_id=sid)
        try:
            usuarios = usuarios_repo.listar_usuarios(cur, empresa_id=emp_id) or []
        except Exception:
            conn.rollback()
            usuarios = []
        try:
            sucursales = sucursales_repo.listar_sucursales(cur, empresa_id=emp_id) or []
        except Exception:
            conn.rollback()
            sucursales = []
        return render_template(
            "reporte_corte_caja.html",
            fecha=fecha,
            datos=datos,
            usuarios=usuarios,
            sucursales=sucursales,
            usuario_id=usuario_id,
            sucursal_id=sucursal_id,
        )
    finally:
        cur.close()
        conn.close()


def _get_cierre_por_usuario_fecha(cur, emp_id: int, usuario_id: int, fecha_str: str) -> dict | None:
    """Obtiene el último cierre CERRADO del cajero en la fecha. Para reporte PDF."""
    cur.execute(
        """
        SELECT c.id, c.monto_apertura, c.ventas_efectivo, c.salidas_efectivo,
               c.monto_esperado, c.monto_real, c.diferencia, c.fecha_cierre,
               (SELECT username FROM usuarios WHERE id = c.usuario_id),
               (SELECT nombre FROM sucursales WHERE id = c.sucursal_id)
        FROM cierre_caja c
        WHERE c.empresa_id = %s AND c.usuario_id = %s
          AND c.fecha_cierre::date = %s AND c.estado = 'CERRADO'
        ORDER BY c.fecha_cierre DESC LIMIT 1
        """,
        (emp_id, usuario_id, fecha_str),
    )
    r = cur.fetchone()
    if not r:
        return None
    return {
        "monto_apertura": float(r[1] or 0),
        "ventas_efectivo": float(r[2] or 0),
        "salidas_efectivo": float(r[3] or 0),
        "monto_esperado": float(r[4] or 0),
        "monto_real": float(r[5] or 0),
        "diferencia": float(r[6] or 0),
        "fecha_cierre": r[7],
        "cajero_nombre": r[8] or "—",
        "sucursal_nombre": r[9] or "",
    }


@bp.route("/reporte/facturacion/corte_caja/ticket")
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def reporte_corte_caja_ticket():
    """Vista formato ticket 80mm para impresora térmica."""
    emp_id = _empresa_id()
    fecha = request.args.get("fecha", date.today().strftime("%Y-%m-%d"))
    usuario_id = request.args.get("usuario_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    uid = int(usuario_id) if usuario_id.isdigit() else None
    sid = int(sucursal_id) if sucursal_id.isdigit() else None
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        datos = cierre_caja_repo.obtener_datos_corte(cur, emp_id, fecha, usuario_id=uid, sucursal_id=sid)
        cab = _datos_cabecera_reporte_ventas(fecha, fecha)
        cierre = _get_cierre_por_usuario_fecha(cur, emp_id, uid, fecha) if uid else None
        if uid and not cierre:
            cur.execute("SELECT username FROM usuarios WHERE id = %s", (uid,))
            r = cur.fetchone()
            cajero_nombre = r[0] if r else "—"
        elif cierre:
            cajero_nombre = cierre["cajero_nombre"] or "—"
        else:
            cajero_nombre = "Todos los cajeros"
        sucursal_nombre = cierre["sucursal_nombre"] if cierre else cab.get("sucursal", "")
        if "Todas" in (sucursal_nombre or ""):
            sucursal_nombre = ""
        # Formatear fecha para mostrar (dd/mm/yyyy)
        try:
            f = datetime.strptime(fecha, "%Y-%m-%d")
            fecha_display = f.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            fecha_display = fecha
        return render_template(
            "reporte_corte_caja_ticket.html",
            datos=datos,
            cierre=cierre,
            empresa_nombre=cab.get("empresa", "Empresa"),
            sucursal_nombre=sucursal_nombre or "",
            cajero_nombre=cajero_nombre,
            fecha=fecha,
            fecha_display=fecha_display,
            usuario_id=usuario_id,
            sucursal_id=sucursal_id,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/facturacion/corte_caja/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def reporte_corte_caja_exportar_pdf():
    emp_id = _empresa_id()
    fecha = request.args.get("fecha", date.today().strftime("%Y-%m-%d"))
    usuario_id = request.args.get("usuario_id", "").strip()
    sucursal_id = request.args.get("sucursal_id", "").strip()
    uid = int(usuario_id) if usuario_id.isdigit() else None
    sid = int(sucursal_id) if sucursal_id.isdigit() else None
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        datos = cierre_caja_repo.obtener_datos_corte(cur, emp_id, fecha, usuario_id=uid, sucursal_id=sid)
        cab = _datos_cabecera_reporte_ventas(fecha, fecha)
        cierre = _get_cierre_por_usuario_fecha(cur, emp_id, uid, fecha) if uid else None
        if uid and not cierre:
            cur.execute("SELECT username FROM usuarios WHERE id = %s", (uid,))
            r = cur.fetchone()
            cajero_nombre = r[0] if r else "—"
        else:
            cajero_nombre = cierre["cajero_nombre"] if cierre else "—"
        sucursal_nombre = cierre["sucursal_nombre"] if cierre else cab.get("sucursal", "")
        f_cierre = (cierre["fecha_cierre"] if cierre else None)
        f_cierre_str = f_cierre.strftime("%d/%m/%Y %I:%M %p") if f_cierre else "—"

        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=0.6*inch, leftMargin=0.6*inch)
        styles = getSampleStyleSheet()
        story = []

        # 1. Encabezado
        emp_nom = cab.get("empresa", "Empresa")
        titulo = f"{emp_nom}" + (f" / {sucursal_nombre}" if sucursal_nombre and "Todas" not in sucursal_nombre else "")
        story.append(Paragraph(html.escape("CORTE DE CAJA (X/Z)"), styles["Title"]))
        story.append(Paragraph(f"<b>{html.escape(titulo)}</b>", styles["Normal"]))
        story.append(Paragraph(f"<b>Nombre del Cajero:</b> {html.escape(cajero_nombre)}", styles["Normal"]))
        story.append(Paragraph(f"<b>Fecha y Hora de Cierre:</b> {f_cierre_str}", styles["Normal"]))
        story.append(Paragraph(f"<b>Estado del Turno:</b> {'Cerrado' if cierre else 'Resumen del día'}", styles["Normal"]))
        story.append(Spacer(1, 12))

        vpt_det = datos.get("ventas_por_tipo_dte_detalle", {})
        vpt = datos.get("ventas_por_tipo_dte", {})
        vpp = datos.get("ventas_por_pago", {})

        # 2. Resumen Documentos (Fiscal)
        story.append(Paragraph("<b>2. Resumen de Documentos Emitidos (Fiscal)</b>", styles["Normal"]))
        story.append(Paragraph("Cuadre con reportes Ministerio de Hacienda.", ParagraphStyle(name="Meta", parent=styles["Normal"], fontSize=9, textColor=colors.gray)))
        def _cant_tot(tipo):
            d = vpt_det.get(tipo, {}) if isinstance(vpt_det.get(tipo), dict) else {}
            c = d.get("cantidad", 0) if d else 0
            t = float(d.get("total", 0) or vpt.get(tipo, 0) or 0)
            return c, t
        rows = []
        for tipo, lbl, es_nc in [("FACTURA", "Facturas (01)", False), ("CREDITO_FISCAL", "Créditos Fiscales (03)", False),
                                  ("NOTA_CREDITO", "Notas de Crédito (05)", True), ("SUJETO_EXCLUIDO", "Sujeto Excluido (14)", False),
                                  ("TICKET", "Tickets", False)]:
            c, t = _cant_tot(tipo)
            if es_nc and t != 0:
                rows.append([lbl, f"{c} doc · Restando $ {abs(t):,.2f}"])
            else:
                rows.append([lbl, f"{c} doc · $ {t:,.2f}"])
        rows.append(["<b>TOTAL VENTA BRUTA</b>", f"<b>$ {float(datos.get('total_ventas', 0)):,.2f}</b>"])
        t1 = Table(rows, colWidths=[3*inch, 2*inch])
        t1.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black)]))
        story.append(t1)
        story.append(Spacer(1, 12))

        # 3. Desglose Método de Pago
        story.append(Paragraph("<b>3. Desglose por Método de Pago (Financiero)</b>", styles["Normal"]))
        story.append(Paragraph("Justificación de cobro al turno.", ParagraphStyle(name="Meta2", parent=styles["Normal"], fontSize=9, textColor=colors.gray)))
        rows2 = [
            ["Efectivo", f"$ {float(vpp.get('EFECTIVO', 0)):,.2f}"],
            ["Tarjeta de Crédito/Débito", f"$ {float(vpp.get('TARJETA', 0)):,.2f}"],
            ["Transferencia Bancaria", f"$ {float(vpp.get('TRANSFERENCIA', 0)):,.2f}"],
            ["Bitcoin (Chivo/Otros)", f"$ {float(vpp.get('BITCOIN', 0)):,.2f}"],
            ["<b>TOTAL RECIBIDO</b>", f"<b>$ {float(datos.get('total_ventas', 0)):,.2f}</b>"],
        ]
        t2 = Table(rows2, colWidths=[3*inch, 2*inch])
        t2.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black)]))
        story.append(t2)
        story.append(Paragraph("<i>(Debe coincidir con la Venta Bruta)</i>", ParagraphStyle(name="Meta3", parent=styles["Normal"], fontSize=8, textColor=colors.gray)))
        story.append(Spacer(1, 12))

        # 4. Arqueo (solo si hay cierre)
        if cierre:
            story.append(Paragraph("<b>4. Arqueo de Efectivo (El Cuadre)</b>", styles["Normal"]))
            story.append(Paragraph("Detectar faltantes o sobrantes.", ParagraphStyle(name="Meta4", parent=styles["Normal"], fontSize=9, textColor=colors.gray)))
            rows3 = [
                ["Fondo de Apertura (Caja Chica)", f"$ {cierre['monto_apertura']:,.2f}"],
                ["(+) Ventas en Efectivo", f"$ {cierre['ventas_efectivo']:,.2f}"],
                ["(-) Salidas de Efectivo (Gastos)", f"$ {cierre['salidas_efectivo']:,.2f}"],
                ["(=) EFECTIVO ESPERADO", f"$ {cierre['monto_esperado']:,.2f}"],
                ["( ) EFECTIVO REAL CONTADO", f"$ {cierre['monto_real']:,.2f}"],
                ["(+/-) DIFERENCIA", f"$ {cierre['diferencia']:,.2f}"],
            ]
            t3 = Table(rows3, colWidths=[3*inch, 2*inch])
            t3.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#dc2626") if cierre['diferencia'] != 0 else colors.black)]))
            story.append(t3)
            story.append(Spacer(1, 12))

        # 5. Firmas
        story.append(Paragraph("<b>5. Sección de Firmas</b>", styles["Normal"]))
        story.append(Paragraph("Firma Cajero: _________________________", styles["Normal"]))
        story.append(Paragraph("Firma Supervisor: _________________________", styles["Normal"]))

        doc.build(story)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"corte_caja_{fecha}.pdf", mimetype="application/pdf")
    finally:
        cur.close()
        conn.close()


@bp.route("/cierre_caja")
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def cierre_caja():
    """Apertura y cierre a ciegas de caja."""
    emp_id = _empresa_id()
    user_id = session.get("user_id")
    suc_sid = _sucursal_id_session()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        apertura = None
        if user_id:
            apertura_row = cierre_caja_repo.apertura_abierta(cur, int(user_id), emp_id, suc_sid)
            if apertura_row:
                apertura = {"id": apertura_row[0], "monto_apertura": apertura_row[1], "fecha_apertura": str(apertura_row[2]) if apertura_row[2] else ""}
        fecha = date.today().strftime("%Y-%m-%d")
        datos = cierre_caja_repo.obtener_datos_corte(cur, emp_id, fecha, usuario_id=int(user_id) if user_id else None, sucursal_id=suc_sid)
        ultimos_cierres = cierre_caja_repo.listar_cierres_cerrados(cur, emp_id, usuario_id=int(user_id) if user_id else None)
        return render_template("cierre_caja.html", apertura=apertura, datos=datos, ultimos_cierres=ultimos_cierres)
    finally:
        cur.close()
        conn.close()


@bp.route("/cierre_caja/apertura", methods=["POST"])
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def cierre_caja_apertura():
    emp_id = _empresa_id()
    user_id = session.get("user_id")
    suc_sid = _sucursal_id_session()
    monto = float(request.form.get("monto_apertura", 0) or 0)
    if not user_id:
        flash("Sesión inválida", "danger")
        return redirect(url_for("admin.cierre_caja"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        existente = cierre_caja_repo.apertura_abierta(cur, int(user_id), emp_id, suc_sid)
        if existente:
            flash("Ya tiene una caja abierta. Ciérrela primero.", "warning")
            return redirect(url_for("admin.cierre_caja"))
        cierre_caja_repo.crear_apertura(cur, int(user_id), emp_id, monto, suc_sid)
        conn.commit()
        flash(f"Caja abierta con $ {monto:,.2f}", "success")
    except Exception as e:
        conn.rollback()
        flash(str(e), "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.cierre_caja"))


@bp.route("/cierre_caja/cerrar", methods=["POST"])
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def cierre_caja_cerrar():
    emp_id = _empresa_id()
    user_id = session.get("user_id")
    cierre_id = int(request.form.get("cierre_id", 0) or 0)
    ventas_efectivo = float(request.form.get("ventas_efectivo", 0) or 0)
    ventas_tarjeta = float(request.form.get("ventas_tarjeta", 0) or 0)
    ventas_credito = float(request.form.get("ventas_credito", 0) or 0)
    ventas_otro = float(request.form.get("ventas_otro", 0) or 0)
    salidas = float(request.form.get("salidas_efectivo", 0) or 0)
    monto_real = float(request.form.get("monto_real", 0) or 0)
    if not cierre_id or not user_id:
        flash("Datos inválidos", "danger")
        return redirect(url_for("admin.cierre_caja"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        ok = cierre_caja_repo.cerrar_caja(cur, cierre_id, ventas_efectivo, ventas_tarjeta, ventas_credito, ventas_otro, salidas, monto_real, emp_id)
        conn.commit()
        if ok:
            return redirect(url_for("admin.cierre_caja_comprobante", cierre_id=cierre_id))
        else:
            flash("No se pudo cerrar (verifique que la caja esté abierta)", "danger")
    except Exception as e:
        conn.rollback()
        flash(str(e), "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.cierre_caja"))


@bp.route("/cierre_caja/comprobante/<int:cierre_id>")
@rol_requerido("GERENTE", "CONTADOR", "CAJERO")
def cierre_caja_comprobante(cierre_id):
    """Comprobante de Cierre de Caja — vista para imprimir en ticket 80mm o panel admin."""
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        cierre = cierre_caja_repo.get_cierre_con_cabecera(cur, cierre_id)
        if not cierre or cierre.get("empresa_id") != emp_id:
            return render_template("comprobante_cierre_caja.html", cierre=None, datos={})
        datos = cierre_caja_repo.obtener_datos_corte_por_cierre(cur, cierre_id)
        return render_template("comprobante_cierre_caja.html", cierre=cierre, datos=datos)
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/facturacion/cuentas_cobrar")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_cuentas_cobrar():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = ventas_reports_repo.listar_cuentas_por_cobrar(cur, emp_id)
        except Exception:
            pass
        total = sum(float(r[5] or 0) for r in filas)
        cab = _datos_cabecera_inventario("Cartera de clientes — Pendientes de cobro")
        return render_template("reporte_cuentas_cobrar.html", filas=filas, cab=cab, total=total)
    finally:
        cur.close()
        conn.close()


def _exportar_excel_facturacion(titulo, headers, filas, cab, inicio, fin, col_moneda=None):
    """Export Excel para reportes de facturación."""
    from openpyxl import Workbook
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    r, nc = 1, len(headers)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    ws.cell(r, 1, titulo).font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    periodo = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin) if inicio and fin else cab.get("periodo", "")
    for txt in (f"Período: {periodo}", f"Empresa: {cab['empresa']}", f"Generado: {cab['generado']}"):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        ws.cell(r, 1, txt).font = st["font_meta"]
    r += 2
    hdr = r + 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(hdr, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    for i, row in enumerate(filas):
        rr = hdr + 1 + i
        for col, val in enumerate(row, 1):
            cell = ws.cell(rr, col, val)
            cell.border = st["border_all"]
            if col_moneda and col in col_moneda and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = st["align_right"]
            if i % 2:
                cell.fill = st["fill_alt"]
    return wb


@bp.route("/reporte/facturacion/libro_iva/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_exportar_excel():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_libro_iva(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    headers = ["N°", "Fecha", "Tipo DTE", "Cliente", "Doc/NRC", "Venta gravada", "IVA", "Ret. IVA 1%", "Total"]
    rows = [[r[0], r[1], r[2], str(r[3] or "")[:40], str(r[4] or "")[:25], float(r[5] or 0), float(r[6] or 0), float(r[9] or 0) if len(r) > 9 else 0, float(r[7] or 0)] for r in filas]
    wb = _exportar_excel_facturacion("Libro de IVA", headers, rows, cab, inicio, fin, col_moneda=[6, 7, 8, 9])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"libro_iva_{inicio}_{fin}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/facturacion/libro_iva/exportar_csv")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_exportar_csv():
    """CSV para conciliación con DTE transmitidos al MH. Separador ; encoding UTF-8."""
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_libro_iva(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    out = StringIO()
    writer = csv.writer(out, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(["N°", "Fecha", "Tipo DTE", "Cliente", "Doc/NRC", "Venta gravada", "IVA", "Ret. IVA 1%", "Total"])
    for r in filas:
        writer.writerow([
            r[0], r[1], r[2], str(r[3] or "")[:40], str(r[4] or "")[:25],
            f"{float(r[5] or 0):.2f}", f"{float(r[6] or 0):.2f}",
            f"{float(r[9] or 0):.2f}" if len(r) > 9 else "0.00", f"{float(r[7] or 0):.2f}",
        ])
    buf = BytesIO(out.getvalue().encode("utf-8-sig"))
    return send_file(buf, as_attachment=True, download_name=f"libro_iva_{inicio}_{fin}.csv", mimetype="text/csv; charset=utf-8")


@bp.route("/reporte/facturacion/libro_iva/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_exportar_pdf():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_libro_iva(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    periodo = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    headers = ["N°", "Fecha", "Tipo", "Cliente", "Vta grav.", "IVA", "Ret. 1%", "Total"]
    data = [headers] + [[str(r[0]), r[1], str(r[2])[:8], str(r[3] or "")[:20], f"${float(r[5] or 0):,.2f}", f"${float(r[6] or 0):,.2f}", f"${float(r[9] or 0):,.2f}" if len(r) > 9 else "$0.00", f"${float(r[7] or 0):,.2f}"] for r in filas]
    buf = _exportar_pdf_inventario("Libro de IVA", f"Período: {periodo}", headers, data[1:] if len(data) > 1 else [], cab, [0.05, 0.09, 0.07, 0.22, 0.12, 0.11, 0.10, 0.12])
    return send_file(buf, as_attachment=True, download_name=f"libro_iva_{inicio}_{fin}.pdf", mimetype="application/pdf")


# --- Libro IVA Compras ---
@bp.route("/reporte/facturacion/libro_iva_compras")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_compras():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().replace(day=1).strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = []
        try:
            filas = compras_reports_repo.listar_libro_iva_compras(cur, emp_id, inicio, fin)
        except Exception:
            pass
        cab = _datos_cabecera_reporte_ventas(inicio, fin)
        total_gravado = sum(float(r[4] or 0) for r in filas)
        total_iva = sum(float(r[5] or 0) for r in filas)
        total_gen = sum(float(r[6] or 0) for r in filas)
        total_retencion = sum(float(r[7] or 0) for r in filas)
        ret_practicada = sum(float(r[7] or 0) for r in filas if len(r) > 8 and r[8])
        ret_sufrida = total_retencion - ret_practicada
        return render_template(
            "reporte_libro_iva_compras.html",
            filas=filas,
            inicio=inicio,
            fin=fin,
            cab=cab,
            total_gravado=total_gravado,
            total_iva=total_iva,
            total_gen=total_gen,
            total_retencion=total_retencion,
            ret_practicada=ret_practicada,
            ret_sufrida=ret_sufrida,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/reporte/facturacion/libro_iva_compras/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_compras_exportar_excel():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = compras_reports_repo.listar_libro_iva_compras(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    headers = ["N°", "Fecha", "Proveedor", "NIT/NRC", "Compra gravada", "IVA", "Total", "Ret. IVA 1%", "Tipo ret."]
    rows = [[r[0], r[1], str(r[2] or "")[:40], str(r[3] or ""), float(r[4] or 0), float(r[5] or 0), float(r[6] or 0), float(r[7] or 0), "Practicada" if len(r) > 8 and r[8] else "Sufrida"] for r in filas]
    wb = _exportar_excel_facturacion("Libro IVA Compras", headers, rows, cab, inicio, fin, col_moneda=[4, 5, 6, 7])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"libro_iva_compras_{inicio}_{fin}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/facturacion/libro_iva_compras/exportar_csv")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_compras_exportar_csv():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = compras_reports_repo.listar_libro_iva_compras(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    out = StringIO()
    writer = csv.writer(out, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(["N°", "Fecha", "Proveedor", "NIT/NRC", "Compra gravada", "IVA", "Total", "Ret. IVA 1%", "Tipo ret."])
    for r in filas:
        writer.writerow([
            r[0], r[1], str(r[2] or "")[:40], str(r[3] or ""),
            f"{float(r[4] or 0):.2f}", f"{float(r[5] or 0):.2f}", f"{float(r[6] or 0):.2f}",
            f"{float(r[7] or 0):.2f}", "Practicada" if len(r) > 8 and r[8] else "Sufrida",
        ])
    buf = BytesIO(out.getvalue().encode("utf-8-sig"))
    return send_file(buf, as_attachment=True, download_name=f"libro_iva_compras_{inicio}_{fin}.csv", mimetype="text/csv; charset=utf-8")


@bp.route("/reporte/facturacion/libro_iva_compras/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_libro_iva_compras_exportar_pdf():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = compras_reports_repo.listar_libro_iva_compras(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    periodo = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    headers = ["N°", "Fecha", "Proveedor", "Compra grav.", "IVA", "Total", "Ret. 1%", "Tipo"]
    data = [headers] + [[str(r[0]), r[1], str(r[2] or "")[:18], f"${float(r[4] or 0):,.2f}", f"${float(r[5] or 0):,.2f}", f"${float(r[6] or 0):,.2f}", f"${float(r[7] or 0):,.2f}", "Pract." if len(r) > 8 and r[8] else "Sufr."] for r in filas]
    buf = _exportar_pdf_inventario("Libro IVA Compras", f"Período: {periodo}", headers, data[1:] if len(data) > 1 else [], cab, [0.05, 0.08, 0.18, 0.12, 0.10, 0.12, 0.10, 0.08])
    return send_file(buf, as_attachment=True, download_name=f"libro_iva_compras_{inicio}_{fin}.pdf", mimetype="application/pdf")


@bp.route("/reporte/facturacion/ventas_producto/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_ventas_producto_exportar_excel():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_ventas_por_producto(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    headers = ["Producto", "Código", "Cantidad", "Subtotal", "Ventas (#)"]
    rows = [[str(r[0] or "")[:45], str(r[1] or ""), float(r[2] or 0), float(r[3] or 0), r[4]] for r in filas]
    wb = _exportar_excel_facturacion("Ventas por Producto", headers, rows, cab, inicio, fin, col_moneda=[4])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"ventas_producto_{inicio}_{fin}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/facturacion/ventas_producto/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_ventas_producto_exportar_pdf():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", date.today().strftime("%Y-%m-%d"))
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_ventas_por_producto(cur, emp_id, inicio, fin)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_reporte_ventas(inicio, fin)
    periodo = _formatear_fecha_pdf(inicio) + " al " + _formatear_fecha_pdf(fin)
    headers = ["Producto", "Código", "Cant.", "Subtotal"]
    data = [headers] + [[str(r[0] or "")[:35], str(r[1] or "")[:12], f"{float(r[2] or 0):,.1f}", f"${float(r[3] or 0):,.2f}"] for r in filas]
    buf = _exportar_pdf_inventario("Ventas por Producto", f"Período: {periodo}", headers, data[1:] if len(data) > 1 else [], cab, [0.35, 0.15, 0.12, 0.20])
    return send_file(buf, as_attachment=True, download_name=f"ventas_producto_{inicio}_{fin}.pdf", mimetype="application/pdf")


@bp.route("/reporte/facturacion/documentos_anulados/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_documentos_anulados_exportar_excel():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", "")
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_documentos_anulados(cur, emp_id, inicio or None, fin or None)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_inventario(f"{inicio or '—'} al {fin or '—'}")
    headers = ["N°", "Fecha", "Total", "Cliente", "Tipo", "Motivo", "Anuló", "Fecha Anul."]
    rows = [[r[0], r[1], float(r[2] or 0), str(r[3] or "")[:40], r[4], str(r[5] or "")[:50] if len(r) > 5 else "", r[6] if len(r) > 6 else "", r[7] if len(r) > 7 else ""] for r in filas]
    wb = _exportar_excel_facturacion("Documentos Anulados", headers, rows, cab, inicio, fin, col_moneda=[3])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="documentos_anulados.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/facturacion/documentos_anulados/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_documentos_anulados_exportar_pdf():
    emp_id = _empresa_id()
    inicio = request.args.get("inicio", "")
    fin = request.args.get("fin", date.today().strftime("%Y-%m-%d"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_documentos_anulados(cur, emp_id, inicio or None, fin or None)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_inventario(f"{inicio or '—'} al {fin or '—'}")
    headers = ["N°", "Fecha", "Total", "Cliente", "Tipo", "Motivo", "Anuló", "F.Anul"]
    data = [headers] + [[str(r[0]), r[1], f"${float(r[2] or 0):,.2f}", str(r[3] or "")[:20], str(r[4] or "")[:6], str(r[5] or "")[:25] if len(r) > 5 else "", r[6] if len(r) > 6 else "", r[7] if len(r) > 7 else ""] for r in filas]
    buf = _exportar_pdf_inventario("Documentos Anulados", "Control de invalidación DTE", headers, data[1:] if len(data) > 1 else [], cab, [0.05, 0.10, 0.08, 0.22, 0.08, 0.22, 0.10, 0.10])
    return send_file(buf, as_attachment=True, download_name="documentos_anulados.pdf", mimetype="application/pdf")


@bp.route("/reporte/facturacion/cuentas_cobrar/exportar_excel")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_cuentas_cobrar_exportar_excel():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_cuentas_por_cobrar(cur, emp_id)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_inventario("Cartera de clientes")
    headers = ["N°", "Fecha", "Cliente", "Documento", "Tipo", "Total", "Días"]
    rows = [[r[0], r[1], str(r[2] or "")[:40], str(r[3] or ""), r[4], float(r[5] or 0), r[6]] for r in filas]
    wb = _exportar_excel_facturacion("Cuentas por Cobrar", headers, rows, cab, "", "", col_moneda=[6])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="cuentas_por_cobrar.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reporte/facturacion/cuentas_cobrar/exportar_pdf")
@rol_requerido("GERENTE", "CONTADOR")
def reporte_cuentas_cobrar_exportar_pdf():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        filas = ventas_reports_repo.listar_cuentas_por_cobrar(cur, emp_id)
    except Exception:
        filas = []
    finally:
        cur.close()
        conn.close()
    cab = _datos_cabecera_inventario("Cartera de clientes")
    headers = ["N°", "Fecha", "Cliente", "Total", "Días"]
    data = [headers] + [[str(r[0]), r[1], str(r[2] or "")[:30], f"${float(r[5] or 0):,.2f}", r[6]] for r in filas]
    buf = _exportar_pdf_inventario("Cuentas por Cobrar", "Cartera pendiente", headers, data[1:] if len(data) > 1 else [], cab, [0.08, 0.12, 0.42, 0.18, 0.10])
    return send_file(buf, as_attachment=True, download_name="cuentas_por_cobrar.pdf", mimetype="application/pdf")


@bp.route("/configuracion")
@superadmin_required
def configuracion():
    if session.get("rol") not in ("ADMIN", "SUPERADMIN"):
        return redirect(url_for("core.index"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        emp_id = int(request.args.get("empresa") or 0) if es_super else _empresa_id()
        if not emp_id:
            emp_id = _empresa_id()
        if es_super:
            try:
                empresas = empresas_repo.listar_empresas(cur) or []
            except Exception:
                empresas = []
            try:
                empresas_lista = empresas_repo.listar_empresas_detalle(cur) or []
            except Exception:
                empresas_lista = [(e[0], e[1], "", "", False, None) for e in empresas] if empresas else []
        else:
            empresas = []
            empresas_lista = []
        empresa_row = db.ejecutar_sql("SELECT * FROM empresas WHERE id = %s", (emp_id,), es_select=True)
        empresa = empresa_row[0] if empresa_row else None
        actividades = []
        try:
            cur.execute("SELECT codigo, descripcion FROM actividades_economicas ORDER BY codigo")
            actividades = cur.fetchall() or []
        except Exception:
            pass
        codigo_act_emp = ""
        try:
            cur.execute("SELECT codigo_actividad_economica FROM empresas WHERE id = %s", (emp_id,))
            r = cur.fetchone()
            codigo_act_emp = (r[0] or "").strip() if r else ""
        except Exception:
            pass
        empresa_es_gran_contribuyente = empresas_repo.get_empresa_es_gran_contribuyente(cur, emp_id)
        return render_template("configuracion.html", empresa=empresa, empresas=empresas, empresas_lista=empresas_lista, emp_id=emp_id, es_super=es_super, actividades=actividades, codigo_actividad_empresa=codigo_act_emp, empresa_es_gran_contribuyente=empresa_es_gran_contribuyente)
    finally:
        cur.close()
        conn.close()


@bp.route("/crear_empresa", methods=["GET", "POST"])
@superadmin_required
def crear_empresa_route():
    from flask import redirect as _rd
    if request.method == "GET":
        return _rd("/configuracion")
    if session.get("rol") not in ("ADMIN", "SUPERADMIN"):
        return redirect(url_for("core.index"))
    form = request.form
    nombre = (form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre de la empresa es requerido.", "danger")
        return _rd("/configuracion")
    nit = (form.get("nit") or "").strip().replace("_", "")
    nrc = (form.get("nrc") or "").strip().replace("_", "")
    nit_digitos = "".join(c for c in nit if c.isdigit())
    nrc_digitos = "".join(c for c in nrc if c.isdigit())
    # Si hay dígitos pero no completos, rechazar
    if nit_digitos and (len(nit_digitos) != 14 or not nit_digitos.isdigit()):
        flash("NIT debe tener 14 dígitos completos o dejarlo vacío (formato: 0614-161386-001-4).", "danger")
        return _rd("/configuracion")
    if nrc_digitos and (len(nrc_digitos) != 7 or not nrc_digitos.isdigit()):
        flash("NRC debe tener 7 dígitos completos o dejarlo vacío (formato: 123456-7).", "danger")
        return _rd("/configuracion")
    nit_final = nit if len(nit_digitos) == 14 else ""
    nrc_final = nrc if len(nrc_digitos) == 7 else ""
    conn = None
    cur = None
    try:
        conn = psycopg2.connect(**ConexionDB().config)
        cur = conn.cursor()
        codigo_act = (form.get("codigo_actividad_economica") or "").strip()
        actividad = form.get("actividad", "") or ""
        if codigo_act:
            desc = actividades_repo.get_descripcion_por_codigo(cur, codigo_act)
            if desc:
                actividad = desc
        eid = empresas_repo.crear_empresa(
            cur, nombre,
            nit=nit_final,
            nrc=nrc_final,
            actividad=actividad,
            direccion=form.get("direccion", ""),
            telefono=form.get("telefono", ""),
            correo=form.get("correo", ""),
            suscripcion_activa=bool(form.get("suscripcion_activa")),
            fecha_vencimiento=form.get("vencimiento") or None,
            codigo_actividad_economica=codigo_act,
        )
        registrar_accion(cur, historial_usuarios_repo.EVENTO_CONFIG_EMPRESA, f"Empresa '{nombre}' creada")
        conn.commit()
        if eid:
            flash(f"Empresa '{nombre}' creada correctamente.", "success")
            return _rd(f"/configuracion?empresa={eid}")
        flash("No se pudo crear la empresa. Verifique NIT (0614-161386-001-4) y NRC (123456-7).", "danger")
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        flash(f"Error al crear: {str(e)}", "danger")
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass
    return _rd("/configuracion")


@bp.route("/guardar_configuracion", methods=["POST"])
@superadmin_required
def guardar_configuracion():
    from flask import redirect as _rd, url_for as _uf
    if session.get("rol") not in ("ADMIN", "SUPERADMIN"):
        return _rd(_uf("core.index"))
    form = request.form
    nombre = form.get("nombre", "").strip()
    nit = form.get("nit", "").strip()
    nrc = form.get("nrc", "").strip()
    codigo_actividad = form.get("codigo_actividad_economica", "").strip()
    actividad = form.get("actividad", "").strip()
    direccion = form.get("direccion", "").strip()
    tel = form.get("tel", "").strip()
    correo = form.get("correo", "").strip()
    vencimiento = form.get("vencimiento") or None
    suscripcion_activa = True if form.get("suscripcion_activa") else False
    es_gran_contribuyente = bool(form.get("es_gran_contribuyente"))
    emp_form = form.get("empresa_id")
    emp_id = int(emp_form) if emp_form and str(emp_form).isdigit() else _empresa_id()

    db = ConexionDB()
    conn = None
    cur = None
    try:
        conn = psycopg2.connect(**db.config)
        cur = conn.cursor()
        if codigo_actividad:
            desc = actividades_repo.get_descripcion_por_codigo(cur, codigo_actividad)
            if desc:
                actividad = desc
        empresas_repo.actualizar_empresa(
            cur,
            emp_id,
            nombre,
            nit,
            nrc,
            actividad,
            direccion,
            tel,
            correo,
            suscripcion_activa,
            vencimiento,
            codigo_actividad_economica=codigo_actividad,
            es_gran_contribuyente=es_gran_contribuyente,
        )
        registrar_accion(cur, historial_usuarios_repo.EVENTO_CONFIG_EMPRESA, "Configuración empresa actualizada")
        conn.commit()
        flash("Configuración guardada correctamente.", "success")
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f"Error al guardar: {str(e)}", "danger")
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass
    return _rd(f"/configuracion?empresa={emp_id}")


def _normalizar_producto(
    p,
    empresa_nombre: str | None = None,
    empresa_id_default: int | None = None,
) -> tuple:
    """Tupla extendida: + fraccionable, unidades_por_caja, unidades_por_docena, mh_codigo, texto_stock_amigable."""
    from azdigital.utils.stock_display import texto_stock_grupos

    if not p or not isinstance(p, (list, tuple)):
        en = str(empresa_nombre or "—")
        return (0, "", "", 0.0, 0, empresa_id_default, None, en, "—", 0.0, "", 0, False, None, 12, "59", "0 u.", "")
    r = list(p) if isinstance(p, tuple) else list(p)
    bid = r[0]
    cod = r[1] or ""
    nom = r[2] or ""
    precio = float(r[3]) if len(r) > 3 and r[3] is not None else 0.0
    stock = float(r[4]) if len(r) > 4 and r[4] is not None else 0.0
    en = str(empresa_nombre or "—")
    if len(r) >= 18:
        r18 = list(r[:18])
        r18[15] = normalizar_codigo_mh(str(r[15]) if r[15] is not None else None)
        return tuple(r18)
    if len(r) >= 17:
        r17 = list(r[:17])
        r17[15] = normalizar_codigo_mh(str(r[15]) if r[15] is not None else None)
        return tuple(r17 + [""])
    if len(r) >= 16:
        upc = int(r[13]) if r[13] is not None else None
        upd = int(r[14]) if r[14] is not None else 12
        mh = normalizar_codigo_mh(str(r[15]) if r[15] is not None else None)
        txt = texto_stock_grupos(
            stock,
            upc if upc and upc > 0 else None,
            upd if upd > 0 else None,
            etiqueta_umb=None,
            fraccionable=bool(r[12]) if len(r) > 12 else False,
            mh_codigo_unidad=mh,
        )
        return (
            bid,
            cod,
            nom,
            precio,
            stock,
            int(r[5]) if r[5] is not None else None,
            int(r[6]) if r[6] is not None else None,
            str(r[7] or "—"),
            str(r[8] or "—"),
            float(r[9]) if r[9] is not None else 0.0,
            (r[10] or "").strip(),
            float(r[11]) if r[11] is not None else 0.0,
            bool(r[12]),
            upc,
            upd,
            mh,
            txt,
            "",
        )
    # 12 cols: global con costo (emp_id, suc_id, nombre_emp, nombre_suc, costo, promo_tipo, promo_val)
    if len(r) >= 12:
        return (
            bid, cod, nom, precio, stock,
            int(r[5]) if r[5] is not None else None,
            int(r[6]) if r[6] is not None else None,
            str(r[7] or "—"), str(r[8] or "—"),
            float(r[9]) if r[9] is not None else 0.0,
            (r[10] or "").strip(), float(r[11]) if r[11] is not None else 0,
            False, None, 12, "59",
            texto_stock_grupos(stock, None, None),
            "",
        )
    if len(r) >= 11:
        return (
            bid, cod, nom, precio, stock,
            int(r[5]) if r[5] is not None else None,
            int(r[6]) if r[6] is not None else None,
            en, str(r[7] or "—"),
            float(r[8]) if r[8] is not None else 0.0,
            (r[9] or "").strip(), float(r[10]) if r[10] is not None else 0,
            False, None, 12, "59",
            texto_stock_grupos(stock, None, None),
            "",
        )
    if len(r) >= 10:
        return (
            bid, cod, nom, precio, stock,
            int(r[5]) if r[5] is not None else None,
            int(r[6]) if r[6] is not None else None,
            en, str(r[7] or "—"),
            0.0, (r[8] or "").strip(), float(r[9]) if r[9] is not None and isinstance(r[9], (int, float)) else 0,
            False, None, 12, "59",
            texto_stock_grupos(stock, None, None),
            "",
        )
    if len(r) >= 9:
        return (
            bid, cod, nom, precio, stock,
            int(r[5]) if r[5] is not None else None,
            int(r[6]) if r[6] is not None else None,
            str(r[7] or "—"), str(r[8] or "—"),
            0.0, "", 0.0,
            False, None, 12, "59",
            texto_stock_grupos(stock, None, None),
            "",
        )
    if len(r) >= 8:
        return (
            bid, cod, nom, precio, stock,
            int(r[5]) if r[5] is not None else None,
            int(r[6]) if r[6] is not None else None,
            en, str(r[7] or "—"),
            0.0, "", 0.0,
            False, None, 12, "59",
            texto_stock_grupos(stock, None, None),
            "",
        )
    while len(r) < 5:
        r.append(0 if len(r) == 4 else "")
    eid = empresa_id_default if empresa_id_default is not None else None
    return (
        r[0], r[1] or "", r[2] or "", float(r[3] or 0), float(r[4] or 0), eid, None, en, "—", 0.0, "", 0.0,
        False, None, 12, "59",
        texto_stock_grupos(float(r[4] or 0), None, None),
        "",
    )


def _sucursal_valida_para_empresa(cur, sucursal_id: int | None, empresa_id: int) -> bool:
    if sucursal_id is None:
        return True
    suc = sucursales_repo.get_sucursal(cur, int(sucursal_id))
    return bool(suc and len(suc) >= 6 and int(suc[5]) == int(empresa_id))


def _acceso_producto_inventario(cur, producto_id: int, emp_id: int, es_super: bool) -> bool:
    cur.execute("SELECT empresa_id FROM productos WHERE id = %s", (producto_id,))
    r = cur.fetchone()
    if not r:
        return False
    if es_super:
        return True
    return int(r[0]) == int(emp_id)


@bp.route("/inventario")
@rol_requerido("GERENTE", "BODEGUERO")
def inventario():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            raw = productos_repo.listar_inventario_global(cur, limit=500) or []
        else:
            raw = productos_repo.listar_inventario(cur, limit=500, empresa_id=emp_id) or []
        en = session.get("empresa_nombre")
        productos = [_normalizar_producto(p, empresa_nombre=en, empresa_id_default=emp_id) for p in raw]
        empresas = empresas_repo.listar_empresas(cur) or [] if es_super else []
        sucursales_todas = sucursales_repo.listar_sucursales_con_empresa(cur) or [] if es_super else []
        sucursales_empresa = (
            sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or [] if not es_super else []
        )
        rol = (session.get("rol") or "").strip().upper()
        requiere_clave_supervisor = rol == "BODEGUERO"
        catalogo_mh = mh_unidades_repo.listar_todas(cur) or []
        catalogo_mh_grupos = catalogo_para_select_optgroups(dict(catalogo_mh))
        return render_template(
            "inventario.html",
            productos=productos,
            es_superadmin=es_super,
            empresas=empresas,
            sucursales_todas=sucursales_todas,
            sucursales_empresa=sucursales_empresa,
            requiere_clave_supervisor=requiere_clave_supervisor,
            catalogo_mh=catalogo_mh,
            catalogo_mh_grupos=catalogo_mh_grupos,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/inventario/")
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_slash():
    return inventario()


@bp.route("/inventario/presentaciones/<int:producto_id>")
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_presentaciones_producto(producto_id: int):
    """UMB + presentaciones adicionales (Tira, etc.) para el modal de inventario."""
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if not _acceso_producto_inventario(cur, producto_id, emp_id, es_super):
            return jsonify({"error": "No autorizado"}), 403
        if not presentaciones_repo.tabla_existe(cur):
            return jsonify({"umb_nombre": "Unidad base", "extras": []})
        rows = presentaciones_repo.listar_por_producto(cur, producto_id)
        if not rows:
            return jsonify({"umb_nombre": "Unidad base", "extras": []})
        umb = "Unidad base"
        extras: list[dict] = []
        for r in rows:
            es_u = bool(r[3]) if len(r) > 3 else False
            nombre = str(r[1] or "").strip()
            fac = float(r[2]) if r[2] is not None else 1.0
            if es_u:
                umb = nombre or umb
                continue
            low = nombre.lower()
            if low == "docena" or low == "caja":
                continue
            extras.append({"nombre": nombre, "factor": fac})
        return jsonify({"umb_nombre": umb, "extras": extras})
    finally:
        cur.close()
        conn.close()


@bp.route("/inventario/plantilla_carga")
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_plantilla_carga():
    """Descarga plantilla Excel para carga masiva de inventario."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    headers = ["codigo", "nombre", "precio", "stock", "costo"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    ejemplos = [
        ("PROD001", "Producto ejemplo 1", 5.99, 100, 3.50),
        ("PROD002", "Producto ejemplo 2", 12.50, 50, 8.00),
    ]
    for r, row in enumerate(ejemplos, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="plantilla_inventario.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/inventario/cargar_masivo", methods=["POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_cargar_masivo():
    """Carga masiva de productos desde Excel o CSV."""
    if "archivo" not in request.files:
        flash("No se envió ningún archivo.", "danger")
        return redirect(url_for("admin.inventario"))
    f = request.files["archivo"]
    if not f or not f.filename:
        flash("Seleccione un archivo Excel (.xlsx) o CSV.", "danger")
        return redirect(url_for("admin.inventario"))
    fn = (f.filename or "").lower()
    if not (fn.endswith(".xlsx") or fn.endswith(".xls") or fn.endswith(".csv")):
        flash("Formato no soportado. Use Excel (.xlsx) o CSV.", "danger")
        return redirect(url_for("admin.inventario"))

    emp_id = _empresa_id()
    suc_raw = (request.form.get("sucursal_id") or "").strip()
    sucursal_id = int(suc_raw) if suc_raw.isdigit() else None
    if request.form.get("empresa_id") and str(request.form.get("empresa_id")).isdigit():
        emp_id = int(request.form.get("empresa_id"))

    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if not es_super:
            emp_id = _empresa_id()
        if sucursal_id and not _sucursal_valida_para_empresa(cur, sucursal_id, emp_id):
            flash("La sucursal no pertenece a la empresa.", "danger")
            return redirect(url_for("admin.inventario"))

        filas = []
        if fn.endswith(".csv"):
            import csv as csv_mod
            content = f.read().decode("utf-8-sig", errors="replace")
            reader = csv_mod.reader(StringIO(content))
            rows = list(reader)
            if not rows:
                flash("El archivo CSV está vacío.", "danger")
                return redirect(url_for("admin.inventario"))
            encabezados = [str(c).strip().lower() for c in rows[0]]
            idx_cod = _idx_col(encabezados, "codigo", "código", "codigo_barra")
            idx_nom = _idx_col(encabezados, "nombre", "producto", "descripcion")
            idx_pre = _idx_col(encabezados, "precio", "precio_unitario")
            idx_stk = _idx_col(encabezados, "stock", "cantidad", "existencia")
            idx_cos = _idx_col(encabezados, "costo")
            for r in rows[1:]:
                if len(r) <= max(idx_nom or 0, 0):
                    continue
                nom = (r[idx_nom] or "").strip() if idx_nom is not None else ""
                if not nom:
                    continue
                cod = (r[idx_cod] or "").strip() if idx_cod is not None else nom[:30]
                pre = _parse_float(r[idx_pre]) if idx_pre is not None else 0
                stk = _parse_float(r[idx_stk]) if idx_stk is not None else 0
                cos = _parse_float(r[idx_cos]) if idx_cos is not None else 0
                filas.append((cod or nom[:30], nom, pre, stk, cos))
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                flash("Instale openpyxl para cargar Excel: pip install openpyxl", "danger")
                return redirect(url_for("admin.inventario"))
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                flash("El archivo Excel está vacío.", "danger")
                return redirect(url_for("admin.inventario"))
            encabezados = [str(c or "").strip().lower() for c in rows[0]]
            idx_cod = _idx_col(encabezados, "codigo", "código", "codigo_barra")
            idx_nom = _idx_col(encabezados, "nombre", "producto", "descripcion")
            idx_pre = _idx_col(encabezados, "precio", "precio_unitario")
            idx_stk = _idx_col(encabezados, "stock", "cantidad", "existencia")
            idx_cos = _idx_col(encabezados, "costo")
            for r in rows[1:]:
                r = list(r) if r else []
                if len(r) <= max(idx_nom or 0, 0):
                    continue
                nom = (str(r[idx_nom] or "").strip()) if idx_nom is not None else ""
                if not nom:
                    continue
                cod = (str(r[idx_cod] or "").strip()) if idx_cod is not None else nom[:30]
                pre = _parse_float(r[idx_pre]) if idx_pre is not None else 0
                stk = _parse_float(r[idx_stk]) if idx_stk is not None else 0
                cos = _parse_float(r[idx_cos]) if idx_cos is not None else 0
                filas.append((cod or nom[:30], nom, pre, stk, cos))

        if not filas:
            flash("No se encontraron filas válidas (nombre requerido). Revise el formato.", "warning")
            return redirect(url_for("admin.inventario"))

        creados = 0
        errores = []
        for cod, nom, pre, stk, cos in filas:
            try:
                nid = productos_repo.crear_producto(
                    cur, cod, nom, pre, stk,
                    empresa_id=emp_id, sucursal_id=sucursal_id,
                    costo_unitario=cos,
                )
                suc_para_stock = sucursal_id
                if suc_para_stock is None:
                    primera = kardex_repo.primera_sucursal_empresa(cur, emp_id)
                    if primera is None:
                        sucursales_repo.crear_sucursal(cur, "Principal", "0001", "", "", empresa_id=emp_id)
                        cur.execute("SELECT id FROM sucursales WHERE empresa_id = %s ORDER BY id DESC LIMIT 1", (emp_id,))
                        r = cur.fetchone()
                        suc_para_stock = int(r[0]) if r else None
                    else:
                        suc_para_stock = primera
                kardex_repo.reemplazar_stock_unificado(cur, nid, suc_para_stock, stk, registrar_entrada=True)
                creados += 1
            except Exception as ex:
                err_msg = str(ex)
                if "duplicad" in err_msg.lower() or "unique" in err_msg.lower() or "codigo" in err_msg.lower():
                    errores.append(f"'{nom}' (código duplicado)")
                else:
                    errores.append(f"'{nom}': {err_msg[:80]}")

        conn.commit()
        if creados > 0:
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PRODUCTO_CREADO, f"Carga masiva: {creados} producto(s) creados")
            flash(f"Carga completada: {creados} producto(s) creados con sus cantidades.", "success")
        if errores:
            for e in errores[:10]:
                flash(e, "warning")
            if len(errores) > 10:
                flash(f"... y {len(errores) - 10} error(es) más (códigos duplicados).", "warning")
    except Exception as e:
        conn.rollback()
        flash(f"Error al procesar el archivo: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.inventario"))


def _idx_col(encabezados: list, *nombres: str) -> int | None:
    for n in nombres:
        try:
            return encabezados.index(n)
        except ValueError:
            continue
    return None


def _parse_float(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(",", ".")
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def _rol_ajuste_inventario_sin_supervisor() -> bool:
    """Gerente/Admin pueden aplicar ajustes grandes sin segunda firma."""
    r = (session.get("rol") or "").strip().upper()
    return r in ("GERENTE", "ADMIN", "SUPERADMIN")


def _validar_supervisor_ajuste_inventario(cur, username: str, password: str) -> bool:
    username = (username or "").strip()
    password = (password or "").strip()
    if not username or not password:
        return False
    u = usuarios_repo.get_usuario_login(cur, username)
    if not u:
        return False
    rol = (u[3] or "").strip().upper()
    if rol not in ("GERENTE", "ADMIN", "SUPERADMIN"):
        return False
    return verificar_password(u[2], password)


@bp.route("/inventario/kardex/<int:producto_id>")
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_kardex(producto_id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if not _acceso_producto_inventario(cur, producto_id, emp_id, es_super):
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("admin.inventario"))
        prod = productos_repo.get_producto(cur, producto_id, None if es_super else emp_id)
        if not prod:
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("admin.inventario"))
        cod = prod[1] or ""
        nom = prod[2] or ""
        prod_emp = int(prod[5]) if len(prod) > 5 and prod[5] is not None else emp_id
        tiene_kardex = kardex_repo.tabla_existe(cur, kardex_repo.TABLA_KARDEX) and kardex_repo.tabla_existe(
            cur, kardex_repo.TABLA_STOCK
        )
        sucursales_movimiento = sucursales_repo.listar_sucursales_min(cur, empresa_id=prod_emp) or []
        if not sucursales_movimiento and tiene_kardex:
            sucursales_repo.crear_sucursal(cur, "Principal", "0001", "", "", empresa_id=prod_emp)
            conn.commit()
            sucursales_movimiento = sucursales_repo.listar_sucursales_min(cur, empresa_id=prod_emp) or []
        stock_suc = kardex_repo.listar_stock_por_sucursal(cur, producto_id) or []
        if not stock_suc and tiene_kardex:
            stock_actual = float(prod[4] or 0) if len(prod) > 4 else 0
            if stock_actual > 0:
                sid = kardex_repo.primera_sucursal_empresa(cur, prod_emp)
                if sid:
                    kardex_repo.reemplazar_stock_unificado(cur, producto_id, sid, stock_actual, registrar_entrada=True)
                    conn.commit()
                    stock_suc = kardex_repo.listar_stock_por_sucursal(cur, producto_id) or []
        kardex_rows = kardex_repo.listar_kardex_producto(cur, producto_id, limit=250) or []
        umb_nom = presentaciones_repo.nombre_umb_producto(cur, producto_id)
        unidades_por_caja_k = None
        try:
            cur.execute(
                "SELECT unidades_por_caja FROM productos WHERE id = %s",
                (producto_id,),
            )
            rxc = cur.fetchone()
            if rxc and rxc[0] is not None:
                nxc = int(rxc[0])
                if nxc > 0:
                    unidades_por_caja_k = nxc
        except Exception:
            pass
        return render_template(
            "inventario_kardex.html",
            producto_id=producto_id,
            codigo=cod,
            nombre=nom,
            umb_nombre=umb_nom,
            unidades_por_caja=unidades_por_caja_k,
            empresa_id_prod=prod_emp,
            kardex_rows=kardex_rows,
            stock_suc=stock_suc,
            tiene_kardex=tiene_kardex,
            es_superadmin=es_super,
            sucursales_movimiento=sucursales_movimiento,
            etiqueta_motivo=kardex_repo.etiqueta_motivo_ajuste,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/inventario/movimiento", methods=["POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_registrar_movimiento():
    emp_id = _empresa_id()
    form = request.form
    pid_raw = form.get("producto_id") or ""
    tipo = (form.get("tipo") or "").strip().upper()
    notas = (form.get("notas") or "").strip()
    if not pid_raw.isdigit():
        flash("Producto inválido.", "danger")
        return redirect(url_for("admin.inventario"))
    producto_id = int(pid_raw)
    try:
        cant = float((form.get("cantidad") or "0").replace(",", "."))
    except ValueError:
        flash("Cantidad inválida.", "danger")
        return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))
    if cant <= 0:
        flash("La cantidad debe ser mayor a cero.", "warning")
        return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))

    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    uid = session.get("user_id")
    try:
        es_super = _es_superadmin_db(cur)
        if not _acceso_producto_inventario(cur, producto_id, emp_id, es_super):
            flash("Sin permiso.", "danger")
            return redirect(url_for("admin.inventario"))
        modo_cant = (form.get("cantidad_en") or "UMB").strip().upper()
        if modo_cant == "CAJA":
            cur.execute(
                "SELECT COALESCE(unidades_por_caja, 0) FROM productos WHERE id = %s",
                (producto_id,),
            )
            rup = cur.fetchone()
            fac_caja = int(rup[0]) if rup and rup[0] is not None else 0
            if fac_caja <= 0:
                flash(
                    "Defina «unidades por caja» en Inventario para registrar movimientos por caja.",
                    "warning",
                )
                return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))
            cant = float(cant * fac_caja)
        elif modo_cant not in ("", "UMB", "BASE"):
            flash("Modo de cantidad no válido.", "danger")
            return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))
        if not kardex_repo.tabla_existe(cur, kardex_repo.TABLA_STOCK):
            flash("Ejecute el script de base de datos: scripts/alter_inventario_kardex.py", "danger")
            return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))

        s_orig = (form.get("sucursal_id") or "").strip()
        if not s_orig.isdigit():
            flash("Seleccione sucursal de origen.", "warning")
            return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))
        suc_o = int(s_orig)

        if tipo == "ENTRADA":
            kardex_repo.registrar_entrada(cur, producto_id, suc_o, cant, uid, notas or None)
        elif tipo == "SALIDA":
            kardex_repo.registrar_salida(cur, producto_id, suc_o, cant, uid, notas or None)
        elif tipo == "TRASLADO":
            s_dst = (form.get("sucursal_destino_id") or "").strip()
            if not s_dst.isdigit():
                flash("Seleccione sucursal de destino.", "warning")
                return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))
            suc_d = int(s_dst)
            kardex_repo.registrar_traslado(cur, producto_id, suc_o, suc_d, cant, uid, notas or None)
        else:
            flash("Tipo de movimiento no válido.", "danger")
            return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))
        registrar_accion(cur, historial_usuarios_repo.EVENTO_AJUSTE_INVENTARIO, f"Kardex producto #{producto_id}: {tipo}")
        conn.commit()
        flash("Movimiento registrado en kardex.", "success")
    except ValueError as e:
        conn.rollback()
        flash(str(e), "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.inventario_kardex", producto_id=producto_id))


@bp.route("/inventario/conteo-fisico", methods=["GET", "POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_conteo_fisico():
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        emp_id = _empresa_id()
        if es_super:
            emp_raw = request.args.get("empresa_id") if request.method == "GET" else (request.form.get("empresa_id") or "").strip()
            if emp_raw and emp_raw.isdigit():
                emp_id = int(emp_raw)
        else:
            emp_id = _empresa_id()

        sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        tiene_kardex = kardex_repo.tabla_existe(cur, kardex_repo.TABLA_KARDEX) and kardex_repo.tabla_existe(
            cur, kardex_repo.TABLA_STOCK
        )
        if not tiene_kardex:
            flash("Ejecute: python scripts/alter_inventario_kardex.py", "danger")
            return redirect(url_for("admin.inventario"))

        if request.method == "POST":
            if es_super:
                emp_raw = (request.form.get("empresa_id") or "").strip()
                if emp_raw.isdigit():
                    emp_id = int(emp_raw)
            suc_raw = (request.form.get("sucursal_id") or "").strip()
            sucursal_id = int(suc_raw) if suc_raw.isdigit() else None
            fecha_txt = (request.form.get("fecha_conteo") or "").strip()[:10]
            referencia = f"Conteo físico {fecha_txt or date.today().isoformat()}"
            uid = session.get("user_id")
            justificacion = (request.form.get("justificacion_conteo") or "").strip()
            supervisor_u = (request.form.get("supervisor_inventario_usuario") or "").strip()
            supervisor_p = (request.form.get("supervisor_inventario_clave") or "").strip()
            productos = inventario_reports_repo.listar_productos_para_conteo(cur, emp_id, sucursal_id)
            candidatos: list[dict] = []
            errores: list[str] = []
            for p in productos:
                pid = p[0]
                stock_sist = float(p[3] or 0)
                fisico_raw = (request.form.get(f"fisico_{pid}") or "").strip().replace(",", ".")
                if not fisico_raw:
                    continue
                try:
                    fisico = float(fisico_raw)
                except ValueError:
                    continue
                if abs(fisico - stock_sist) < 0.0001:
                    continue
                if fisico < 0:
                    errores.append(f"{p[2]}: conteo no puede ser negativo")
                    continue
                motivo_key = None
                if fisico < stock_sist:
                    motivo_key = (request.form.get(f"motivo_{pid}") or "").strip().upper()
                    if motivo_key not in kardex_repo.MOTIVOS_AJUSTE_SALIDA:
                        errores.append(
                            f"{p[2]}: indique motivo del faltante (merma, avería, ajuste de inventario o faltante)."
                        )
                        continue
                costo_u = float(p[5] or 0) if len(p) > 5 else 0.0
                absdiff = abs(fisico - stock_sist)
                impacto_est = absdiff * costo_u if costo_u > 0 else 0.0
                candidatos.append(
                    {
                        "pid": pid,
                        "nombre": p[2],
                        "stock_sist": stock_sist,
                        "fisico": fisico,
                        "motivo_key": motivo_key,
                        "absdiff": absdiff,
                        "impacto": impacto_est,
                    }
                )
            if errores:
                for e in errores:
                    flash(e, "warning")
            redir = url_for("admin.inventario_conteo_fisico", sucursal_id=suc_raw or None)
            if es_super and emp_id:
                from urllib.parse import urlencode
                qs = {"empresa_id": emp_id}
                if suc_raw:
                    qs["sucursal_id"] = suc_raw
                redir = url_for("admin.inventario_conteo_fisico") + "?" + urlencode(qs)
            if not candidatos:
                if not errores:
                    flash("No hay diferencias que ajustar.", "info")
                return redirect(redir)
            min_j = kardex_repo.MIN_CARACTERES_JUSTIFICACION_AJUSTE
            if len(justificacion) < min_j:
                flash(
                    f"Justificación obligatoria del conteo (mín. {min_j} caracteres): "
                    "explique el acta, responsable o hallazgos.",
                    "warning",
                )
                return redirect(redir)
            max_imp = max(c["impacto"] for c in candidatos)
            max_qty = max(c["absdiff"] for c in candidatos)
            need_sup = max_imp > kardex_repo.UMBRAL_IMPACTO_AJUSTE_USD or max_qty > kardex_repo.UMBRAL_CANTIDAD_AJUSTE_UMB
            if need_sup and not _rol_ajuste_inventario_sin_supervisor():
                if not _validar_supervisor_ajuste_inventario(cur, supervisor_u, supervisor_p):
                    flash(
                        "Ajuste elevado: supera $" + f"{kardex_repo.UMBRAL_IMPACTO_AJUSTE_USD:.0f} "
                        "o " + f"{kardex_repo.UMBRAL_CANTIDAD_AJUSTE_UMB:.0f} UMB. "
                        "Ingrese usuario y clave de Gerente o Administrador.",
                        "danger",
                    )
                    return redirect(redir)
            aplicados = 0
            errores2: list[str] = []
            for c in candidatos:
                try:
                    kardex_repo.ajustar_por_conteo_fisico(
                        cur,
                        c["pid"],
                        c["stock_sist"],
                        c["fisico"],
                        sucursal_id,
                        uid,
                        referencia,
                        motivo_ajuste=c["motivo_key"],
                        comentario_justificacion=justificacion,
                    )
                    aplicados += 1
                except ValueError as e:
                    errores2.append(f"{c['nombre']}: {e}")
            if errores2:
                for e in errores2:
                    flash(e, "warning")
            if aplicados > 0:
                registrar_accion(
                    cur,
                    historial_usuarios_repo.EVENTO_CONTEO_FISICO,
                    f"Conteo físico: {aplicados} producto(s). Ref: {referencia[:80]}",
                )
                conn.commit()
                flash(f"Ajustes aplicados: {aplicados} producto(s). Quedaron registrados en Kardex con justificación.", "success")
            elif candidatos:
                conn.rollback()
                flash(
                    "No se aplicó ningún ajuste. Revise los mensajes (p. ej. stock insuficiente en sucursal).",
                    "warning",
                )
            redir = url_for("admin.inventario_conteo_fisico", sucursal_id=suc_raw or None)
            if es_super and emp_id:
                from urllib.parse import urlencode
                qs = {"empresa_id": emp_id}
                if suc_raw: qs["sucursal_id"] = suc_raw
                redir = url_for("admin.inventario_conteo_fisico") + "?" + urlencode(qs)
            return redirect(redir)

        suc_raw = request.args.get("sucursal_id", "").strip()
        sucursal_id = int(suc_raw) if suc_raw.isdigit() else None
        productos = inventario_reports_repo.listar_productos_para_conteo(cur, emp_id, sucursal_id)
        sucursal_nombre = None
        if sucursal_id:
            s = sucursales_repo.get_sucursal(cur, sucursal_id)
            sucursal_nombre = s[1] if s else None
        empresas = empresas_repo.listar_empresas(cur) or [] if es_super else []
        return render_template(
            "inventario_conteo_fisico.html",
            productos=productos,
            sucursales=sucursales,
            sucursal_id=sucursal_id,
            sucursal_nombre=sucursal_nombre,
            fecha_hoy=date.today().isoformat(),
            es_superadmin=es_super,
            empresas=empresas,
            empresa_id=emp_id,
            umbral_impacto_ajuste_usd=kardex_repo.UMBRAL_IMPACTO_AJUSTE_USD,
            umbral_cantidad_ajuste_umb=kardex_repo.UMBRAL_CANTIDAD_AJUSTE_UMB,
            min_caracteres_justificacion_ajuste=kardex_repo.MIN_CARACTERES_JUSTIFICACION_AJUSTE,
        )
    finally:
        cur.close()
        conn.close()


def _construir_datos_reporte_conteo(cur, emp_id: int, sucursal_id, fecha_txt: str, form_data: dict) -> tuple:
    """Retorna (filas_detalle, filas_grandes_deficiencias, resumen, sucursal_nombre)."""
    productos = inventario_reports_repo.listar_productos_para_conteo(cur, emp_id, sucursal_id)
    sucursal_nombre = "Todas (Global)" if not sucursal_id else ""
    if sucursal_id:
        s = sucursales_repo.get_sucursal(cur, sucursal_id)
        sucursal_nombre = s[1] if s else "—"
    filas = []
    grandes = []
    total_productos = len(productos)
    contados = 0
    con_diferencias = 0
    total_grandes = 0
    valor_ajuste_pos = 0.0
    valor_ajuste_neg = 0.0
    UMBRAL_PCT = 10
    UMBRAL_UNIDADES = 5
    for p in productos:
        pid, codigo, nombre, stock_sist, unidad, costo = p[0], p[1], p[2], float(p[3] or 0), p[4], float(p[5] if len(p) > 5 else 0)
        fisico_raw = (form_data.get(f"fisico_{pid}") or "").strip().replace(",", ".")
        fisico = float(fisico_raw) if fisico_raw else None
        if fisico is None:
            filas.append((codigo, nombre, stock_sist, "—", "—", "—", "Sin conteo", 0.0))
            continue
        contados += 1
        diff = fisico - stock_sist
        pct = (100 * diff / stock_sist) if stock_sist > 0 else (100 if diff > 0 else 0)
        es_grande = abs(pct) >= UMBRAL_PCT or abs(diff) >= UMBRAL_UNIDADES
        if abs(diff) >= 0.001:
            con_diferencias += 1
        if es_grande and abs(diff) >= 0.001:
            total_grandes += 1
            valor_ajuste = abs(diff) * costo
            if diff > 0:
                valor_ajuste_pos += valor_ajuste
            else:
                valor_ajuste_neg += valor_ajuste
            grandes.append((codigo, nombre, stock_sist, fisico, diff, pct, "GRAN DEFICIENCIA", valor_ajuste))
        estado = "Gran deficiencia" if es_grande and abs(diff) >= 0.001 else ("Deficiencia" if abs(diff) >= 0.001 else "OK")
        impacto = abs(diff) * costo if abs(diff) >= 0.001 else 0.0
        filas.append((codigo, nombre, stock_sist, fisico, diff, round(pct, 1), estado, round(impacto, 2)))
    no_contados = total_productos - contados
    resumen = {
        "total_productos": total_productos,
        "contados": contados,
        "no_contados": no_contados,
        "con_diferencias": con_diferencias,
        "grandes_deficiencias": total_grandes,
        "valor_faltantes": round(valor_ajuste_neg, 2),
        "valor_sobrantes": round(valor_ajuste_pos, 2),
    }
    return filas, grandes, resumen, sucursal_nombre


@bp.route("/inventario/conteo-fisico/exportar-excel", methods=["POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_conteo_fisico_exportar_excel():
    form = request.form
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if _es_superadmin_db(cur):
            emp_raw = (form.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
    finally:
        cur.close()
        conn.close()
    db2 = ConexionDB()
    conn2 = psycopg2.connect(**db2.config)
    cur2 = conn2.cursor()
    try:
        suc_raw = (form.get("sucursal_id") or "").strip()
        sucursal_id = int(suc_raw) if suc_raw.isdigit() else None
        fecha_txt = (form.get("fecha_conteo") or date.today().isoformat())[:10]
        filas, grandes, resumen, suc_nom = _construir_datos_reporte_conteo(cur2, emp_id, sucursal_id, fecha_txt, form)
    finally:
        cur2.close()
        conn2.close()
    cab = _datos_cabecera_inventario(f"Conteo físico {fecha_txt}", suc_nom, emp_id)
    cab["sucursal"] = suc_nom
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return "<p>Instala openpyxl: pip install openpyxl</p>", 500
    st = _estilos_excel_elegante()
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo Fisico"
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1, "Reporte de Conteo Físico — Sistema vs. Bodega").font = st["font_title"]
    ws.cell(r, 1).alignment = st["align_center"]
    r += 1
    for txt in (f"Fecha conteo: {fecha_txt}", f"Empresa: {cab['empresa']}", f"Sucursal: {cab['sucursal']}", f"Generado: {cab['generado']}"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(r, 1, txt).font = st["font_meta"]
        r += 1
    r += 1
    headers = ["Código", "Producto", "Stock sistema", "Conteo físico", "Diferencia", "% Var", "Estado", "Impacto $"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(r, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    hdr = r
    r += 1
    for row in filas:
        for col, val in enumerate(row, 1):
            cell = ws.cell(r, col, val)
            cell.border = st["border_all"]
            if col in (3, 4, 5, 6, 8) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00' if col != 8 else '"$"#,##0.00'
                cell.alignment = st["align_right"]
            if "Gran deficiencia" in str(row[6]):
                cell.fill = PatternFill(start_color="FFFEE2E2", end_color="FFFEE2E2", fill_type="solid")
        r += 1
    r += 2
    ws.cell(r, 1, "RESUMEN").font = Font(bold=True, size=12)
    r += 1
    ws.cell(r, 1, f"Productos en el sistema: {resumen['total_productos']}").font = st["font_meta"]
    r += 1
    ws.cell(r, 1, f"Productos contados: {resumen['contados']}").font = st["font_meta"]
    r += 1
    ws.cell(r, 1, f"Productos NO contados: {resumen['no_contados']}").font = Font(bold=resumen['no_contados'] > 0, color="FFDC2626" if resumen['no_contados'] > 0 else None)
    r += 1
    ws.cell(r, 1, f"Productos con diferencias: {resumen['con_diferencias']}").font = st["font_meta"]
    r += 1
    ws.cell(r, 1, f"Grandes deficiencias (≥{10}% o ≥5 und): {resumen['grandes_deficiencias']}").font = Font(bold=True, color="FFDC2626")
    r += 1
    ws.cell(r, 1, f"Impacto faltantes (físico < sistema): ${resumen['valor_faltantes']:,.2f}").font = st["font_meta"]
    r += 1
    ws.cell(r, 1, f"Impacto sobrantes (físico > sistema): ${resumen['valor_sobrantes']:,.2f}").font = st["font_meta"]
    r += 2
    no_contados_filas = [f for f in filas if f[6] == "Sin conteo"]
    if no_contados_filas:
        ws.cell(r, 1, "PRODUCTOS NO CONTADOS (sin valor físico)").font = Font(bold=True, size=11, color="FFDC2626")
        r += 1
        for f in no_contados_filas[:100]:
            ws.cell(r, 1, f"{f[0]} — {f[1][:40]}").font = st["font_meta"]
            r += 1
        if len(no_contados_filas) > 100:
            ws.cell(r, 1, f"... y {len(no_contados_filas) - 100} más").font = st["font_meta"]
            r += 1
        r += 1
    ws.cell(r, 1, "GRANDES DEFICIENCIAS (detalle)").font = Font(bold=True, size=11, color="FFDC2626")
    r += 1
    for col, h in enumerate(headers, 1):
        c = ws.cell(r, col, h)
        c.font, c.fill, c.alignment, c.border = st["font_header"], st["fill_header"], st["align_center"], st["border_all"]
    r += 1
    for row in grandes:
        for col, val in enumerate(row, 1):
            cell = ws.cell(r, col, val)
            cell.border = st["border_all"]
            cell.fill = PatternFill(start_color="FFFEE2E2", end_color="FFFEE2E2", fill_type="solid")
            if col in (3, 4, 5, 6, 8):
                cell.number_format = '#,##0.00' if col != 8 else '"$"#,##0.00'
                cell.alignment = st["align_right"]
            r += 1
    r += 2
    recom = [
        "• Revise causas de diferencias (merma, robos, errores de registro).",
        "• Documente acta de conteo firmada por responsable.",
        "• Grandes deficiencias requieren investigación y medidas correctivas.",
    ]
    ws.cell(r, 1, "RECOMENDACIONES").font = Font(bold=True, size=11)
    r += 1
    for txt in recom:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(r, 1, txt).font = st["font_meta"]
        r += 1
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 14
    ws.column_dimensions["B"].width = 28
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"conteo_fisico_{fecha_txt}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/inventario/conteo-fisico/exportar-pdf", methods=["POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def inventario_conteo_fisico_exportar_pdf():
    form = request.form
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if _es_superadmin_db(cur):
            emp_raw = (form.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
    finally:
        cur.close()
        conn.close()
    suc_raw = (form.get("sucursal_id") or "").strip()
    sucursal_id = int(suc_raw) if suc_raw.isdigit() else None
    fecha_txt = (form.get("fecha_conteo") or date.today().isoformat())[:10]
    db2 = ConexionDB()
    conn2 = psycopg2.connect(**db2.config)
    cur2 = conn2.cursor()
    try:
        filas, grandes, resumen, suc_nom = _construir_datos_reporte_conteo(cur2, emp_id, sucursal_id, fecha_txt, form)
    finally:
        cur2.close()
        conn2.close()
    cab = _datos_cabecera_inventario(f"Conteo físico {fecha_txt}", suc_nom, emp_id)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table
    except ImportError:
        return "<p>Instala reportlab: pip install reportlab</p>", 500
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="RptMeta", parent=styles["Normal"], fontSize=9, leading=12, textColor=colors.HexColor("#475569")))
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=0.6*inch, leftMargin=0.6*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = [
        Paragraph(html.escape("Reporte de Conteo Físico — Sistema vs. Bodega"), ParagraphStyle(name="Title", fontName="Helvetica-Bold", fontSize=14, spaceAfter=6)),
        Paragraph(f"<b>Fecha:</b> {html.escape(fecha_txt)} | <b>Sucursal:</b> {html.escape(suc_nom)} | <b>Empresa:</b> {html.escape(cab['empresa'])}", styles["RptMeta"]),
        Spacer(1, 0.15*inch),
        Paragraph(f"<b>Resumen:</b> {resumen['total_productos']} en sistema | {resumen['contados']} contados | <b>{resumen['no_contados']} NO contados</b> | {resumen['con_diferencias']} con diferencias | {resumen['grandes_deficiencias']} grandes deficiencias | Faltantes: ${resumen['valor_faltantes']:,.2f} | Sobrantes: ${resumen['valor_sobrantes']:,.2f}", styles["RptMeta"]),
        Spacer(1, 0.2*inch),
    ]
    headers = ["Código", "Producto", "Sist.", "Fís.", "Dif.", "%", "Estado"]
    data = [headers] + [[str(x) for x in (r[0], r[1][:25], r[2], r[3], r[4], r[5], r[6])] for r in filas[:50]]
    if len(filas) > 50:
        data.append(["...", f"({len(filas)-50} productos más)", "", "", "", "", ""])
    t = Table(data, colWidths=[0.9*inch, 2.2*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 1.2*inch], repeatRows=1)
    t.setStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1"))])
    elements.append(t)
    if grandes:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph("<b>Grandes deficiencias (≥10% o ≥5 unidades)</b>", ParagraphStyle(fontName="Helvetica-Bold", fontSize=10, textColor=colors.HexColor("#dc2626"))))
        elements.append(Spacer(1, 0.08*inch))
        data2 = [headers] + [[str(x) for x in (r[0], r[1][:25], r[2], r[3], r[4], r[5], r[6])] for r in grandes]
        t2 = Table(data2, colWidths=[0.9*inch, 2.2*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 1.2*inch], repeatRows=1)
        t2.setStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dc2626")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fef2f2")])])
        elements.append(t2)
    elements.append(Spacer(1, 0.15*inch))
    elements.append(Paragraph("<i>Recomendaciones: Documente acta de conteo firmada. Investigue grandes deficiencias. AZ DIGITAL</i>", styles["RptMeta"]))
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"conteo_fisico_{fecha_txt}.pdf", mimetype="application/pdf")


def _entero_positivo_desde_formulario(raw: str | None) -> int | None:
    """Acepta '60', '60.0', etc. (isdigit() falla con decimales y no aplicaba precio UMB desde caja)."""
    s = (raw or "").strip().replace(",", ".")
    if not s:
        return None
    try:
        v = float(s)
        if v <= 0:
            return None
        return int(round(v))
    except ValueError:
        return None


@bp.route("/guardar_producto", methods=["POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def guardar_producto():
    form = request.form
    producto_id = form.get("producto_id")
    codigo = (form.get("codigo") or "").strip()
    nombre = (form.get("nombre") or "").strip()
    precio_str_strip = (form.get("precio") or "").strip()
    if not precio_str_strip:
        precio_str_strip = (form.get("precio_umb_mirror") or "").strip()
    costo_str_strip = (form.get("costo") or "").strip()
    precio_sin_valor_form = not precio_str_strip
    costo_sin_valor_form = not costo_str_strip
    precio = precio_str_strip.replace(",", ".")
    costo = costo_str_strip.replace(",", ".")
    stock = form.get("stock", "0").replace(",", ".")
    promo_tipo = (form.get("promocion_tipo") or "").strip().upper()
    promo_val_raw = (form.get("promocion_valor") or "").strip()
    promo_val = float(promo_val_raw) if promo_val_raw else None
    if promo_tipo and promo_tipo not in ("2X1", "PORCENTAJE"):
        promo_tipo = ""
    fraccionable = form.get("fraccionable") in ("1", "on", "true", "True", "yes", "Yes")
    upcaja_raw = (form.get("unidades_por_caja") or "").strip()
    unidades_por_caja = _entero_positivo_desde_formulario(upcaja_raw)
    updoc_raw = (form.get("unidades_por_docena") or "12").strip()
    unidades_por_docena = _entero_positivo_desde_formulario(updoc_raw) or 12
    if unidades_por_docena < 1:
        unidades_por_docena = 12
    mh_cod = normalizar_codigo_mh(form.get("mh_codigo_unidad"))
    umb_nombre = (form.get("umb_nombre") or "Unidad base").strip()[:80] or "Unidad base"
    raw_extras = (form.get("presentaciones_extra_json") or "").strip() or "[]"
    try:
        extra_list = json.loads(raw_extras)
    except json.JSONDecodeError:
        extra_list = []
    extras: list[tuple[str, float]] = []
    for x in extra_list:
        if isinstance(x, dict) and x.get("nombre") and x.get("factor") is not None:
            try:
                extras.append((str(x["nombre"]).strip()[:80], float(x["factor"])))
            except (TypeError, ValueError):
                pass
    emp_id = _empresa_id()
    suc_raw = (form.get("sucursal_id") or "").strip()
    sucursal_id = int(suc_raw) if suc_raw.isdigit() else None
    try:
        precio_f = float(precio) if precio else 0
        costo_f = float(costo) if costo else 0
        stock_f = float(stock) if stock else 0
    except ValueError:
        flash("Precio, costo y stock deben ser números.", "danger")
        return redirect(url_for("admin.inventario"))
    fp_raw = (form.get("factor_presentacion_comercial") or "").strip().replace(",", ".")
    factor_pres: float | None = None
    if fp_raw:
        try:
            factor_pres = float(fp_raw)
            if factor_pres <= 0:
                factor_pres = None
        except ValueError:
            factor_pres = None
    if factor_pres is None and unidades_por_caja and unidades_por_caja > 0:
        factor_pres = float(unidades_por_caja)
    pp_pres = form.get("precio_por_presentacion") or form.get("precio_por_caja")
    cp_pres = form.get("costo_por_presentacion") or form.get("costo_por_caja")
    # Sin factor (ej. queso por libra, sin caja): no exigir «unidades por caja»; los montos
    # opcionales de «precio/costo caja» se ignoran y mandan el precio/costo por UMB del formulario.
    if factor_pres is None:
        pp_pres = None
        cp_pres = None
    precio_f, costo_f = aplicar_derivacion_desde_presentacion(
        pp_pres,
        cp_pres,
        factor_pres,
        precio_f,
        costo_f,
    )
    if not nombre:
        flash("El nombre del producto es requerido.", "danger")
        return redirect(url_for("admin.inventario"))

    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if producto_id and producto_id.isdigit():
            prod = (
                productos_repo.get_producto(cur, int(producto_id), None)
                if es_super
                else productos_repo.get_producto(cur, int(producto_id), emp_id)
            )
            if not prod:
                flash("Producto no encontrado.", "danger")
                return redirect(url_for("admin.inventario"))
            if precio_sin_valor_form and not presentacion_tiene_monto_derivable(pp_pres) and len(prod) > 3:
                precio_f = float(prod[3] or 0)
            if costo_sin_valor_form and not presentacion_tiene_monto_derivable(cp_pres):
                try:
                    cur.execute(
                        "SELECT COALESCE(costo_unitario, 0) FROM productos WHERE id = %s",
                        (int(producto_id),),
                    )
                    rc = cur.fetchone()
                    if rc:
                        costo_f = float(rc[0] or 0)
                except Exception:
                    pass
            stock_actual = float(prod[4] or 0) if len(prod) > 4 else 0
            rol = (session.get("rol") or "").strip().upper()
            if rol == "BODEGUERO" and (
                stock_f < stock_actual or (stock_actual > 0 and stock_f == 0)
            ):
                sup_user = (form.get("supervisor_user") or "").strip()
                sup_pass = form.get("supervisor_password") or ""
                if not sup_user or not sup_pass:
                    flash("Se requiere autorización de Gerente para reducir stock o dejarlo en cero.", "warning")
                    return redirect(url_for("admin.inventario"))
                u = usuarios_repo.get_usuario_login(cur, sup_user)
                if not u:
                    flash("Usuario o contraseña de gerente incorrectos.", "danger")
                    return redirect(url_for("admin.inventario"))
                u_rol = (u[3] or "").strip().upper() if len(u) > 3 else ""
                if u_rol not in ("GERENTE", "ADMIN", "SUPERADMIN"):
                    flash("Solo Gerente o Administrador puede autorizar cambios de stock.", "danger")
                    return redirect(url_for("admin.inventario"))
                if not verificar_password(u[2], sup_pass):
                    flash("Usuario o contraseña de gerente incorrectos.", "danger")
                    return redirect(url_for("admin.inventario"))
            prod_emp = int(prod[5]) if len(prod) > 5 and prod[5] is not None else emp_id
            target_emp = (
                _empresa_id_desde_form_super(cur, es_super, form.get("empresa_id"), prod_emp)
                if es_super
                else emp_id
            )
            if not _sucursal_valida_para_empresa(cur, sucursal_id, target_emp):
                flash("La sucursal no pertenece a la empresa indicada.", "danger")
                return redirect(url_for("admin.inventario"))
            nueva_emp = target_emp if es_super and target_emp != prod_emp else None
            if es_super:
                productos_repo.actualizar_producto(
                    cur,
                    int(producto_id),
                    codigo,
                    nombre,
                    precio_f,
                    stock_f,
                    sucursal_id,
                    empresa_scope=None,
                    nueva_empresa_id=nueva_emp,
                    promocion_tipo=promo_tipo or None,
                    promocion_valor=promo_val,
                    costo_unitario=costo_f,
                    fraccionable=fraccionable,
                    unidades_por_caja=unidades_por_caja,
                    unidades_por_docena=unidades_por_docena,
                    mh_codigo_unidad=mh_cod,
                )
            else:
                productos_repo.actualizar_producto(
                    cur,
                    int(producto_id),
                    codigo,
                    nombre,
                    precio_f,
                    stock_f,
                    sucursal_id,
                    empresa_scope=emp_id,
                    nueva_empresa_id=None,
                    promocion_tipo=promo_tipo or None,
                    promocion_valor=promo_val,
                    costo_unitario=costo_f,
                    fraccionable=fraccionable,
                    unidades_por_caja=unidades_por_caja,
                    unidades_por_docena=unidades_por_docena,
                    mh_codigo_unidad=mh_cod,
                )
            suc_para_stock = sucursal_id
            if suc_para_stock is None:
                primera = kardex_repo.primera_sucursal_empresa(cur, target_emp)
                if primera is None:
                    sucursales_repo.crear_sucursal(
                        cur, "Principal", "0001", "", "", empresa_id=target_emp
                    )
                    cur.execute("SELECT id FROM sucursales WHERE empresa_id = %s ORDER BY id DESC LIMIT 1", (target_emp,))
                    r = cur.fetchone()
                    suc_para_stock = int(r[0]) if r else None
                else:
                    suc_para_stock = primera
            if presentaciones_repo.tabla_existe(cur):
                filas = presentaciones_repo.construir_filas_desde_legacy(
                    umb_nombre, unidades_por_docena, unidades_por_caja, extras if extras else None
                )
                presentaciones_repo.reemplazar_todas(cur, int(producto_id), filas)
            kardex_repo.reemplazar_stock_unificado(cur, int(producto_id), suc_para_stock, stock_f)
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PRODUCTO_EDITADO, f"Producto #{producto_id} actualizado")
            conn.commit()
            flash("Producto actualizado.", "success")
        else:
            target_emp = _empresa_id_desde_form_super(cur, es_super, form.get("empresa_id"), emp_id)
            if not _sucursal_valida_para_empresa(cur, sucursal_id, target_emp):
                flash("La sucursal no pertenece a la empresa indicada.", "danger")
                return redirect(url_for("admin.inventario"))
            nid = productos_repo.crear_producto(
                cur, codigo or nombre[:20], nombre, precio_f, stock_f,
                empresa_id=target_emp, sucursal_id=sucursal_id,
                promocion_tipo=promo_tipo or None, promocion_valor=promo_val,
                costo_unitario=costo_f,
                fraccionable=fraccionable,
                unidades_por_caja=unidades_por_caja,
                unidades_por_docena=unidades_por_docena,
                mh_codigo_unidad=mh_cod,
            )
            # Si crear_producto usó INSERT reducido (BD antigua), rellenar conversión/MH en un UPDATE aparte.
            productos_repo._actualizar_conversion_producto(
                cur,
                nid,
                fraccionable,
                unidades_por_caja,
                unidades_por_docena,
                mh_cod,
                empresa_scope=None if es_super else target_emp,
            )
            suc_para_stock = sucursal_id
            if suc_para_stock is None:
                primera = kardex_repo.primera_sucursal_empresa(cur, target_emp)
                if primera is None:
                    sucursales_repo.crear_sucursal(
                        cur, "Principal", "0001", "", "", empresa_id=target_emp
                    )
                    cur.execute("SELECT id FROM sucursales WHERE empresa_id = %s ORDER BY id DESC LIMIT 1", (target_emp,))
                    r = cur.fetchone()
                    suc_para_stock = int(r[0]) if r else None
                else:
                    suc_para_stock = primera
            if presentaciones_repo.tabla_existe(cur):
                filas = presentaciones_repo.construir_filas_desde_legacy(
                    umb_nombre, unidades_por_docena, unidades_por_caja, extras if extras else None
                )
                presentaciones_repo.reemplazar_todas(cur, int(nid), filas)
            kardex_repo.reemplazar_stock_unificado(cur, nid, suc_para_stock, stock_f, registrar_entrada=True)
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PRODUCTO_CREADO, f"Producto #{nid} registrado")
            conn.commit()
            flash("Producto registrado.", "success")
        return redirect(url_for("admin.inventario"))
    except Exception as e:
        conn.rollback()
        err = str(e)
        if "productos_codigo_barra_key" in err or "productos_empresa_codigo_uniq" in err or ("codigo_barra" in err and "duplicad" in err.lower()):
            flash("Ya existe un producto con ese código. Use un código distinto o edite el existente.", "danger")
        else:
            flash(f"Error al guardar: {err}", "danger")
        return redirect(url_for("admin.inventario"))
    finally:
        cur.close()
        conn.close()


@bp.route("/eliminar_producto/<int:producto_id>", methods=["GET", "POST"])
@rol_requerido("GERENTE", "BODEGUERO")
def eliminar_producto(producto_id):
    if request.method == "GET":
        return redirect(url_for("admin.inventario"))
    rol = (session.get("rol") or "").strip().upper()
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if rol == "BODEGUERO":
            form = request.form
            sup_user = (form.get("supervisor_user") or "").strip()
            sup_pass = form.get("supervisor_password") or ""
            if not sup_user or not sup_pass:
                flash("Se requiere autorización de Gerente para eliminar productos.", "warning")
                return redirect(url_for("admin.inventario"))
            u = usuarios_repo.get_usuario_login(cur, sup_user)
            if not u:
                flash("Usuario o contraseña de supervisor incorrectos.", "danger")
                return redirect(url_for("admin.inventario"))
            u_rol = (u[3] or "").strip().upper() if len(u) > 3 else ""
            if u_rol not in ("GERENTE", "ADMIN", "SUPERADMIN"):
                flash("Solo Gerente o Administrador puede autorizar la eliminación.", "danger")
                return redirect(url_for("admin.inventario"))
            if not verificar_password(u[2], sup_pass):
                flash("Usuario o contraseña de supervisor incorrectos.", "danger")
                return redirect(url_for("admin.inventario"))
        es_super = _es_superadmin_db(cur)
        scope = None if es_super else emp_id
        deleted = productos_repo.eliminar_producto(cur, producto_id, scope)
        if deleted:
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PRODUCTO_ELIMINADO, f"Producto #{producto_id} eliminado")
            conn.commit()
            flash("Producto eliminado.", "success")
        else:
            conn.rollback()
            flash("Producto no encontrado o no pertenece a su empresa.", "warning")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.inventario"))


# --- PROMOCIONES (por fechas) ---
@bp.route("/promociones")
@admin_required
def promociones():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        rows = promociones_repo.listar_promociones(cur, empresa_id=emp_id)
        return render_template("promociones.html", promociones=rows, es_superadmin=es_super)
    finally:
        cur.close()
        conn.close()


@bp.route("/promociones/nueva", methods=["GET", "POST"])
@admin_required
def promociones_nueva():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        productos = productos_repo.listar_inventario(cur, limit=500, empresa_id=emp_id) or []
        if request.method == "POST":
            nombre = (request.form.get("nombre") or "").strip()
            tipo = (request.form.get("tipo") or "").strip().upper()
            valor_raw = (request.form.get("valor") or "").strip().replace(",", ".")
            fi = request.form.get("fecha_inicio") or ""
            ff = request.form.get("fecha_fin") or ""
            activa = request.form.get("activa") == "1"
            prod_ids = [int(x) for x in request.form.getlist("producto_id") if str(x).isdigit()]
            vc = request.form.get("valor_comprar") or ""
            vp = request.form.get("valor_pagar") or ""
            dm = request.form.get("descuento_monto") or ""
            pr_id = (request.form.get("producto_regalo_id") or "").strip()
            cmin = request.form.get("cantidad_min_compra") or ""
            creg = request.form.get("cantidad_regalo") or ""
            if not nombre:
                flash("El nombre es requerido.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=None)
            try:
                valor = float(valor_raw) if valor_raw else (2 if tipo in ("2X1", "3X2", "VOLUMEN") else 0)
            except ValueError:
                valor = 2
            if tipo == "PORCENTAJE" and (valor <= 0 or valor >= 100):
                flash("Para descuento %, el valor debe estar entre 1 y 99.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=None)
            if tipo == "PRECIO_FIJO" and (not valor_raw or valor <= 0):
                flash("Para precio fijo, indica un precio mayor a 0.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=None)
            if tipo == "DESCUENTO_CANTIDAD" and (valor <= 0 or valor >= 100):
                flash("Para descuento por cantidad, el % debe estar entre 1 y 99.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=None)
            if tipo == "2X1":
                valor, vc, vp = 2, 2, 1
            elif tipo == "3X2":
                valor, vc, vp = 3, 3, 2
            valor_comprar = float(vc) if vc else (2 if tipo in ("2X1", "VOLUMEN") else None)
            valor_pagar = float(vp) if vp else (1 if tipo in ("2X1", "VOLUMEN") else None)
            descuento_monto = float(dm) if dm else None
            producto_regalo_id = int(pr_id) if pr_id.isdigit() else None
            cantidad_min_compra = float(cmin) if cmin else 1
            cantidad_regalo = float(creg) if creg else 1
            fecha_inicio = date.fromisoformat(fi) if fi and len(fi) >= 10 else None
            fecha_fin = date.fromisoformat(ff) if ff and len(ff) >= 10 else None
            promociones_repo.crear_promocion(
                cur, emp_id, nombre, tipo, valor, fecha_inicio, fecha_fin, activa, producto_ids=prod_ids or None,
                valor_comprar=valor_comprar, valor_pagar=valor_pagar, descuento_monto=descuento_monto,
                producto_regalo_id=producto_regalo_id, cantidad_min_compra=cantidad_min_compra, cantidad_regalo=cantidad_regalo,
            )
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PROMOCION_CREADA, "Promoción creada")
            conn.commit()
            flash("Promoción creada.", "success")
            return redirect(url_for("admin.promociones"))
        return render_template("promocion_form.html", productos=productos, promocion=None)
    finally:
        cur.close()
        conn.close()


@bp.route("/promociones/editar/<int:promocion_id>", methods=["GET", "POST"])
@admin_required
def promociones_editar(promocion_id: int):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        promocion = promociones_repo.get_promocion(cur, promocion_id, empresa_id=emp_id)
        if not promocion:
            flash("Promoción no encontrada.", "danger")
            return redirect(url_for("admin.promociones"))
        prod_ids = promociones_repo.get_productos_promocion(cur, promocion_id)
        productos = productos_repo.listar_inventario(cur, limit=500, empresa_id=promocion[1] or emp_id) or []
        if request.method == "POST":
            nombre = (request.form.get("nombre") or "").strip()
            tipo = (request.form.get("tipo") or "").strip().upper()
            valor_raw = (request.form.get("valor") or "").strip().replace(",", ".")
            fi = request.form.get("fecha_inicio") or ""
            ff = request.form.get("fecha_fin") or ""
            activa = request.form.get("activa") == "1"
            prod_ids = [int(x) for x in request.form.getlist("producto_id") if str(x).isdigit()]
            vc = request.form.get("valor_comprar") or ""
            vp = request.form.get("valor_pagar") or ""
            dm = request.form.get("descuento_monto") or ""
            pr_id = (request.form.get("producto_regalo_id") or "").strip()
            cmin = request.form.get("cantidad_min_compra") or ""
            creg = request.form.get("cantidad_regalo") or ""
            if not nombre:
                flash("El nombre es requerido.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=promocion, prod_ids=prod_ids)
            try:
                valor = float(valor_raw) if valor_raw else (2 if tipo in ("2X1", "3X2", "VOLUMEN") else 0)
            except ValueError:
                valor = 2
            if tipo == "PORCENTAJE" and (valor <= 0 or valor >= 100):
                flash("Para descuento %, el valor debe estar entre 1 y 99.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=promocion, prod_ids=prod_ids)
            if tipo == "PRECIO_FIJO" and (not valor_raw or valor <= 0):
                flash("Para precio fijo, indica un precio mayor a 0.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=promocion, prod_ids=prod_ids)
            if tipo == "DESCUENTO_CANTIDAD" and (valor <= 0 or valor >= 100):
                flash("Para descuento por cantidad, el % debe estar entre 1 y 99.", "danger")
                return render_template("promocion_form.html", productos=productos, promocion=promocion, prod_ids=prod_ids)
            if tipo == "2X1":
                valor, vc, vp = 2, 2, 1
            elif tipo == "3X2":
                valor, vc, vp = 3, 3, 2
            valor_comprar = float(vc) if vc else (2 if tipo in ("2X1", "VOLUMEN") else None)
            valor_pagar = float(vp) if vp else (1 if tipo in ("2X1", "VOLUMEN") else None)
            descuento_monto = float(dm) if dm else None
            producto_regalo_id = int(pr_id) if pr_id.isdigit() else None
            cantidad_min_compra = float(cmin) if cmin else 1
            cantidad_regalo = float(creg) if creg else 1
            fecha_inicio = date.fromisoformat(fi) if fi and len(fi) >= 10 else None
            fecha_fin = date.fromisoformat(ff) if ff and len(ff) >= 10 else None
            promociones_repo.actualizar_promocion(
                cur, promocion_id, nombre, tipo, valor, fecha_inicio, fecha_fin, activa, producto_ids=prod_ids, empresa_id=emp_id,
                valor_comprar=valor_comprar, valor_pagar=valor_pagar, descuento_monto=descuento_monto,
                producto_regalo_id=producto_regalo_id, cantidad_min_compra=cantidad_min_compra, cantidad_regalo=cantidad_regalo,
            )
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PROMOCION_EDITADA, f"Promoción #{promocion_id} actualizada")
            conn.commit()
            flash("Promoción actualizada.", "success")
            return redirect(url_for("admin.promociones"))
        return render_template("promocion_form.html", productos=productos, promocion=promocion, prod_ids=prod_ids)
    finally:
        cur.close()
        conn.close()


@bp.route("/promociones/eliminar/<int:promocion_id>", methods=["POST"])
@admin_required
def promociones_eliminar(promocion_id: int):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if promociones_repo.eliminar_promocion(cur, promocion_id, empresa_id=emp_id):
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PROMOCION_ELIMINADA, f"Promoción #{promocion_id} eliminada")
            conn.commit()
            flash("Promoción eliminada.", "success")
        else:
            flash("Promoción no encontrada.", "warning")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.promociones"))


def _es_superadmin():
    """ADMIN y SUPERADMIN = superusuario con control total."""
    r = (session.get("rol") or "").strip().upper()
    return r in ("ADMIN", "SUPERADMIN")


def _es_superadmin_db(cur):
    """Verifica en BD si el usuario actual es superusuario (ADMIN o SUPERADMIN)."""
    uid = session.get("user_id")
    if not uid:
        return False
    cur.execute("SELECT rol FROM usuarios WHERE id = %s AND activo = TRUE", (uid,))
    r = cur.fetchone()
    rol = str(r[0] or "").strip().upper() if r else ""
    return rol in ("ADMIN", "SUPERADMIN")


@bp.route("/usuarios")
@admin_required
def usuarios():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            usuarios_lista = usuarios_repo.listar_usuarios_todos(cur) or []
            empresas = empresas_repo.listar_empresas(cur) or []
            sucursales_todas = sucursales_repo.listar_sucursales_con_empresa(cur) or []
            sucursales = []
        else:
            usuarios_lista = usuarios_repo.listar_usuarios(cur, empresa_id=emp_id) or []
            empresas = []
            sucursales_todas = []
            sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        return render_template(
            "usuarios.html",
            usuarios=usuarios_lista,
            sucursales=sucursales,
            sucursales_todas=sucursales_todas,
            empresas=empresas,
            usuario=None,
            es_superadmin=es_super,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/usuarios/")
@admin_required
def usuarios_slash():
    return usuarios()


@bp.route("/usuarios/historial")
@admin_required
def historial_usuarios():
    """Muestra el historial de accesos (login, logout, cambios de contraseña)."""
    emp_id = _empresa_id() if not _es_superadmin() else None
    evento = request.args.get("evento", "").strip() or None
    usuario_id = request.args.get("usuario_id", "").strip()
    usuario_id = int(usuario_id) if usuario_id.isdigit() else None
    offset = max(0, int(request.args.get("offset", 0) or 0))
    limite = min(500, max(50, int(request.args.get("limite", 100) or 100)))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if not historial_usuarios_repo.tabla_existe(cur):
            flash("La tabla de historial no existe. Ejecute: python scripts/alter_historial_usuarios.py", "warning")
            return redirect(url_for("admin.usuarios"))
        registros = historial_usuarios_repo.listar(
            cur, empresa_id=emp_id, usuario_id=usuario_id, evento=evento, limite=limite, offset=offset
        )
        return render_template(
            "historial_usuarios.html",
            registros=registros,
            evento_actual=evento,
            usuario_id_filtro=usuario_id,
            offset=offset,
            limite=limite,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/usuarios/editar/<int:id>")
@admin_required
def editar_usuario(id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            usuarios_lista = usuarios_repo.listar_usuarios_todos(cur) or []
            empresas = empresas_repo.listar_empresas(cur) or []
            sucursales_todas = sucursales_repo.listar_sucursales_con_empresa(cur) or []
            sucursales = []
        else:
            usuarios_lista = usuarios_repo.listar_usuarios(cur, empresa_id=emp_id) or []
            ids_permitidos = [u[0] for u in usuarios_lista]
            if id not in ids_permitidos:
                flash("Usuario no encontrado o no pertenece a su empresa.", "danger")
                return redirect(url_for("admin.usuarios"))
            empresas = []
            sucursales_todas = []
            sucursales = sucursales_repo.listar_sucursales_min(cur, empresa_id=emp_id) or []
        usuario_edit = usuarios_repo.get_usuario(cur, id)
        if not usuario_edit:
            flash("Usuario no encontrado.", "danger")
            return redirect(url_for("admin.usuarios"))
        rol_editado = str(usuario_edit[2] or "").strip().upper()
        if not es_super and rol_editado in ("ADMIN", "SUPERADMIN"):
            flash("Solo un superusuario puede editar otros superusuarios.", "danger")
            return redirect(url_for("admin.usuarios"))
        return render_template(
            "usuarios.html",
            usuarios=usuarios_lista,
            sucursales=sucursales,
            sucursales_todas=sucursales_todas,
            empresas=empresas,
            usuario=usuario_edit,
            es_superadmin=_es_superadmin(),
        )
    except Exception as e:
        flash(f"Error al cargar usuario: {str(e)}", "danger")
        return redirect(url_for("admin.usuarios"))
    finally:
        cur.close()
        conn.close()


@bp.route("/guardar_usuario", methods=["POST"])
@admin_required
def guardar_usuario():
    form = request.form
    usuario_id = form.get("usuario_id")
    username = (form.get("user") or "").strip()
    password = (form.get("pass") or "").strip()
    rol = (form.get("rol") or "CAJERO").strip().upper()
    sucursal_id = form.get("sucursal")
    # Verificar superadmin desde BD
    db_check = ConexionDB()
    conn_check = psycopg2.connect(**db_check.config)
    cur_check = conn_check.cursor()
    cur_check.execute("SELECT rol FROM usuarios WHERE id = %s", (session.get("user_id"),))
    r_rol = cur_check.fetchone()
    cur_check.close()
    conn_check.close()
    r_rol_str = str(r_rol[0] or "").strip().upper() if r_rol else ""
    es_super = r_rol_str in ("ADMIN", "SUPERADMIN")

    if es_super:
        emp_form = form.get("empresa")
        rol_guardar = (form.get("rol") or "CAJERO").strip().upper()
        if rol_guardar in ("ADMIN", "SUPERADMIN") and emp_form in ("0", "", "none"):
            empresa_id = None  # Superusuario sin empresa
        elif emp_form and str(emp_form).isdigit() and int(emp_form) > 0:
            empresa_id = int(emp_form)
        else:
            empresa_id = _empresa_id()
    else:
        empresa_id = _empresa_id()
    if not username:
        flash("El usuario es requerido.", "danger")
        return redirect(url_for("admin.usuarios"))
    if not usuario_id and not password:
        flash("La contraseña es requerida al crear usuario.", "danger")
        return redirect(url_for("admin.usuarios"))
    if not es_super:
        if rol in ("SUPERADMIN", "ADMIN"):
            rol = "CAJERO"
            flash("Solo un superusuario puede crear otros superusuarios.", "warning")

    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if usuario_id and usuario_id.isdigit():
            pw_hash = generate_password_hash(password) if password else None
            usuarios_repo.actualizar_usuario(cur, int(usuario_id), username, pw_hash, rol, sucursal_id, empresa_id)
            registrar_accion(cur, historial_usuarios_repo.EVENTO_USUARIO_EDITADO, f"Usuario {username} actualizado")
            conn.commit()
            flash("Usuario actualizado correctamente.", "success")
        else:
            pw_hash = generate_password_hash(password)
            usuarios_repo.crear_usuario(cur, username, pw_hash, rol, sucursal_id, empresa_id)
            registrar_accion(cur, historial_usuarios_repo.EVENTO_USUARIO_CREADO, f"Usuario {username} creado")
            conn.commit()
            flash("Usuario creado correctamente.", "success")
        return redirect(url_for("admin.usuarios"))
    except Exception as e:
        conn.rollback()
        flash(f"Error al guardar: {str(e)}", "danger")
        return redirect(url_for("admin.usuarios"))
    finally:
        cur.close()
        conn.close()


@bp.route("/eliminar_usuario/<int:usuario_id>", methods=["GET", "POST"])
@admin_required
def eliminar_usuario(usuario_id):
    if request.method == "GET":
        return redirect(url_for("admin.usuarios"))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        cur.execute("SELECT rol FROM usuarios WHERE id = %s", (session.get("user_id"),))
        r_rol2 = cur.fetchone()
        r_rol2_str = str(r_rol2[0] or "").strip().upper() if r_rol2 else ""
        es_super_del = r_rol2_str in ("ADMIN", "SUPERADMIN")
        if not es_super_del:
            emp_id = _empresa_id()
            usuarios_lista = usuarios_repo.listar_usuarios(cur, empresa_id=emp_id) or []
            if usuario_id not in [u[0] for u in usuarios_lista]:
                flash("Usuario no encontrado o no pertenece a su empresa.", "danger")
                return redirect(url_for("admin.usuarios"))
            u_target = usuarios_repo.get_usuario(cur, usuario_id)
            if u_target and str(u_target[2] or "").strip().upper() in ("ADMIN", "SUPERADMIN"):
                flash("Solo un superusuario puede desactivar otros superusuarios.", "danger")
                return redirect(url_for("admin.usuarios"))
        usuarios_repo.eliminar_usuario(cur, usuario_id)
        registrar_accion(cur, historial_usuarios_repo.EVENTO_USUARIO_ELIMINADO, f"Usuario #{usuario_id} desactivado")
        conn.commit()
        flash("Usuario desactivado correctamente.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.usuarios"))


# ========== PROVEEDORES Y COMPRAS ==========

@bp.route("/proveedores")
@rol_requerido("GERENTE")
def proveedores():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        lista = proveedores_repo.listar(cur, emp_id, incluir_inactivos=True)
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        actividades = []
        try:
            cur.execute("SELECT codigo, descripcion FROM actividades_economicas ORDER BY codigo")
            actividades = cur.fetchall() or []
        except Exception:
            pass
        return render_template("proveedores.html", proveedores=lista, empresas=empresas, emp_id=emp_id, es_super=es_super, proveedor=None, actividades=actividades)
    finally:
        cur.close()
        conn.close()


@bp.route("/proveedores/editar/<int:id>")
@rol_requerido("GERENTE")
def proveedores_editar(id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        p = proveedores_repo.get(cur, id, emp_id)
        if not p:
            flash("Proveedor no encontrado.", "danger")
            return redirect(url_for("admin.proveedores"))
        lista = proveedores_repo.listar(cur, emp_id, incluir_inactivos=True)
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        actividades = []
        try:
            cur.execute("SELECT codigo, descripcion FROM actividades_economicas ORDER BY codigo")
            actividades = cur.fetchall() or []
        except Exception:
            pass
        proveedor = {"id": p[0], "empresa_id": p[1], "nombre": p[2], "nit": p[3] or "", "nrc": p[4] or "", "direccion": p[5] or "", "telefono": p[6] or "", "correo": p[7] or "", "contacto": p[8] or "", "activo": p[9], "tipo_documento": (p[10] or "NIT").strip().upper() if len(p) > 10 else "NIT", "giro_actividad": p[11] or "" if len(p) > 11 else "", "clasificacion_contribuyente": (p[12] or "PEQUEÑO").strip().upper() if len(p) > 12 else "PEQUEÑO", "es_gran_contribuyente": bool(p[13]) if len(p) > 13 else False}
        return render_template("proveedores.html", proveedores=lista, empresas=empresas, emp_id=emp_id, es_super=es_super, proveedor=proveedor, actividades=actividades)
    finally:
        cur.close()
        conn.close()


@bp.route("/guardar_proveedor", methods=["POST"])
@rol_requerido("GERENTE")
def guardar_proveedor():
    from azdigital.utils.validar_documentos import validar_nit_dui, validar_nrc

    form = request.form
    emp_id = _empresa_id()
    es_super = False
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super and form.get("empresa_id"):
            emp_id = int(form.get("empresa_id"))
        prov_id = form.get("proveedor_id", "").strip()
        nombre = (form.get("nombre") or "").strip()
        if not nombre:
            flash("El nombre es requerido.", "danger")
            return redirect(url_for("admin.proveedores"))
        nit = (form.get("nit") or "").strip()
        nrc = (form.get("nrc") or "").strip()
        tipo_doc = (form.get("tipo_documento") or "NIT").strip().upper() or "NIT"
        if tipo_doc not in ("NIT", "DUI"):
            tipo_doc = "NIT"
        if nit:
            ok, msg = validar_nit_dui(tipo_doc, nit)
            if not ok:
                flash(msg, "danger")
                return redirect(url_for("admin.proveedores") + (f"?empresa_id={emp_id}" if es_super and emp_id else ""))
        if nrc:
            ok, msg = validar_nrc(nrc)
            if not ok:
                flash(msg, "danger")
                return redirect(url_for("admin.proveedores") + (f"?empresa_id={emp_id}" if es_super and emp_id else ""))
        giro = (form.get("giro_actividad") or "").strip()
        cla = (form.get("clasificacion_contribuyente") or "PEQUEÑO").strip().upper() or "PEQUEÑO"
        if cla not in ("GRANDE", "MEDIANO", "PEQUEÑO"):
            cla = "PEQUEÑO"
        es_gran = bool(form.get("es_gran_contribuyente"))
        if prov_id.isdigit():
            proveedores_repo.actualizar(cur, int(prov_id), emp_id, nombre, nit, nrc,
                form.get("direccion", ""), form.get("telefono", ""), form.get("correo", ""), form.get("contacto", ""),
                activo=bool(form.get("activo")), tipo_documento=tipo_doc, giro_actividad=giro,
                clasificacion_contribuyente=cla, es_gran_contribuyente=es_gran)
        else:
            proveedores_repo.crear(cur, emp_id, nombre, nit, nrc, form.get("direccion", ""),
                form.get("telefono", ""), form.get("correo", ""), form.get("contacto", ""),
                tipo_documento=tipo_doc, giro_actividad=giro, clasificacion_contribuyente=cla, es_gran_contribuyente=es_gran)
        registrar_accion(cur, historial_usuarios_repo.EVENTO_PROVEEDOR_GUARDADO, f"Proveedor {nombre} guardado")
        conn.commit()
        flash("Proveedor guardado.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.proveedores") + (f"?empresa_id={emp_id}" if es_super and emp_id else ""))


@bp.route("/eliminar_proveedor/<int:id>", methods=["POST"])
@rol_requerido("GERENTE")
def eliminar_proveedor(id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super and request.form.get("empresa_id"):
            emp_id = int(request.form.get("empresa_id"))
        if proveedores_repo.eliminar(cur, id, emp_id):
            registrar_accion(cur, historial_usuarios_repo.EVENTO_PROVEEDOR_ELIMINADO, f"Proveedor #{id} eliminado")
            conn.commit()
            flash("Proveedor eliminado (o marcado inactivo si tenía compras).", "success")
        else:
            flash("Proveedor no encontrado.", "warning")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.proveedores"))


@bp.route("/compras/cargar_dte", methods=["GET", "POST"])
@rol_requerido("GERENTE")
def compras_cargar_dte():
    """Carga JSON DTE del proveedor, extrae codigoGeneracion y selloRecepcion."""
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or request.form.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        if request.method == "POST" and "archivo_dte" in request.files:
            f = request.files["archivo_dte"]
            if f and f.filename and f.filename.lower().endswith(".json"):
                try:
                    contenido = f.read().decode("utf-8", errors="replace")
                    from azdigital.services.dte_proveedor_service import extraer_dte_proveedor, DTEProveedorData
                    dte = extraer_dte_proveedor(contenido)
                    session["dte_proveedor_temp"] = {
                        "codigo_generacion": dte.codigo_generacion,
                        "sello_recepcion": dte.sello_recepcion,
                        "numero_control": dte.numero_control,
                        "tipo_dte": dte.tipo_dte,
                        "numero_documento": dte.numero_documento,
                        "fecha_emision": dte.fecha_emision,
                        "emisor_nombre": dte.emisor_nombre,
                        "emisor_nit": dte.emisor_nit,
                        "emisor_nrc": dte.emisor_nrc,
                        "receptor_nit": dte.receptor_nit,
                        "receptor_nrc": dte.receptor_nrc,
                        "total_gravado": dte.total_gravado,
                        "total_iva": dte.total_iva,
                        "total": dte.total,
                        "items": dte.items,
                        "valido": dte.valido,
                        "error": dte.error,
                    }
                    return render_template("compras_cargar_dte.html", dte_data=dte, empresas=empresas, emp_id=emp_id, es_super=es_super)
                except Exception as e:
                    flash(f"Error al procesar JSON: {e}", "danger")
            else:
                flash("Seleccione un archivo .json válido.", "warning")
        dte_temp = session.get("dte_proveedor_temp")
        dte_data = None
        if dte_temp:
            from azdigital.services.dte_proveedor_service import DTEProveedorData
            dte_data = DTEProveedorData(
                codigo_generacion=dte_temp.get("codigo_generacion", ""),
                sello_recepcion=dte_temp.get("sello_recepcion", ""),
                numero_control=dte_temp.get("numero_control", ""),
                tipo_dte=dte_temp.get("tipo_dte", ""),
                numero_documento=dte_temp.get("numero_documento", ""),
                fecha_emision=dte_temp.get("fecha_emision", ""),
                emisor_nombre=dte_temp.get("emisor_nombre", ""),
                emisor_nit=dte_temp.get("emisor_nit", ""),
                emisor_nrc=dte_temp.get("emisor_nrc", ""),
                receptor_nit=dte_temp.get("receptor_nit", ""),
                receptor_nrc=dte_temp.get("receptor_nrc", ""),
                total_gravado=float(dte_temp.get("total_gravado", 0)),
                total_iva=float(dte_temp.get("total_iva", 0)),
                total=float(dte_temp.get("total", 0)),
                items=dte_temp.get("items", []),
                valido=dte_temp.get("valido", False),
                error=dte_temp.get("error", ""),
            )
        return render_template("compras_cargar_dte.html", dte_data=dte_data, empresas=empresas, emp_id=emp_id, es_super=es_super)
    finally:
        cur.close()
        conn.close()


@bp.route("/compras/registrar_desde_dte", methods=["POST"])
@rol_requerido("GERENTE")
def compras_registrar_desde_dte():
    """Registra compra desde DTE cargado en sesión."""
    emp_id = _empresa_id()
    if request.form.get("empresa_id") and str(session.get("rol", "")).upper() in ("ADMIN", "SUPERADMIN"):
        try:
            emp_id = int(request.form.get("empresa_id"))
        except ValueError:
            pass
    dte_temp = session.get("dte_proveedor_temp")
    if not dte_temp or not dte_temp.get("valido") or not dte_temp.get("emisor_nit"):
        flash("No hay DTE válido en sesión. Cargue el archivo JSON nuevamente.", "warning")
        return redirect(url_for("admin.compras_cargar_dte") + (f"?empresa_id={emp_id}" if emp_id else ""))
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        prov = proveedores_repo.buscar_por_nit(cur, dte_temp["emisor_nit"], emp_id)
        if prov:
            prov_id = int(prov[0])
        else:
            prov_id = proveedores_repo.crear(cur, emp_id, dte_temp["emisor_nombre"] or "Proveedor DTE", dte_temp["emisor_nit"], dte_temp.get("emisor_nrc", ""))
        num_fact = dte_temp.get("numero_documento", "") or dte_temp.get("numero_control", "") or "DTE"
        fecha = dte_temp.get("fecha_emision", "")[:10] if dte_temp.get("fecha_emision") else date.today().isoformat()
        notas = f"Importado desde DTE. Código: {dte_temp.get('codigo_generacion', '')}"
        compra_id = compras_repo.crear(cur, emp_id, prov_id, num_fact, fecha, session.get("user_id"), notas, dte_temp.get("codigo_generacion", ""), dte_temp.get("sello_recepcion", ""))
        total_compra = 0.0
        items_dte = dte_temp.get("items", [])
        suc_id = None
        try:
            suc_id = kardex_repo.primera_sucursal_empresa(cur, emp_id)
        except Exception:
            pass
        for it in items_dte:
            desc = (it.get("descripcion") or "").strip()
            cod = (it.get("codigo") or "").strip()
            cant = float(it.get("cantidad", 1) or 1)
            precio = float(it.get("precio_unitario", 0) or 0)
            if precio <= 0:
                subt = float(it.get("subtotal", 0) or 0)
                precio = subt / cant if cant else 0
            prod = productos_repo.buscar_por_codigo(cur, cod, empresa_id=emp_id) if cod else None
            if not prod and desc:
                prods = productos_repo.buscar_por_nombre(cur, desc[:30], limit=1, empresa_id=emp_id)
                prod = prods[0] if prods else None
            if prod:
                pid = int(prod[0])
                total_compra += cant * precio
                compras_repo.agregar_detalle(cur, compra_id, pid, cant, precio)
                lista_compras_repo.registrar_costo_compra(cur, pid, precio, cant, session.get("user_id"), f"Compra #{compra_id} DTE")
                lista_compras_repo.actualizar_costo_y_precio_producto(cur, pid, emp_id, precio, None)
                try:
                    if suc_id:
                        kardex_repo.registrar_entrada(cur, pid, suc_id, cant, session.get("user_id"), f"Compra #{compra_id} DTE")
                    else:
                        productos_repo.incrementar_stock(cur, pid, cant)
                except Exception:
                    productos_repo.incrementar_stock(cur, pid, cant)
        emp_gran = empresas_repo.get_empresa_es_gran_contribuyente(cur, emp_id)
        prov_data = proveedores_repo.get(cur, prov_id, emp_id)
        prov_gran = bool(prov_data[13]) if prov_data and len(prov_data) > 13 else False
        from azdigital.services.compras_service import calcular_retencion_iva_compras
        ret_iva = calcular_retencion_iva_compras(emp_gran, prov_gran, total_compra)
        compras_repo.actualizar_total(cur, compra_id, retencion_iva=ret_iva)
        registrar_accion(cur, historial_usuarios_repo.EVENTO_COMPRA_REGISTRADA, f"Compra #{compra_id} desde DTE")
        conn.commit()
        session.pop("dte_proveedor_temp", None)
        flash(f"Compra #{compra_id} registrada desde DTE. Proveedor: {dte_temp.get('emisor_nombre', '')}.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.compras") + (f"?empresa_id={emp_id}" if str(session.get("rol", "")).upper() in ("ADMIN", "SUPERADMIN") and emp_id else ""))


@bp.route("/compras")
@rol_requerido("GERENTE")
def compras():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        lista = compras_repo.listar(cur, emp_id)
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        return render_template("compras.html", compras=lista, empresas=empresas, emp_id=emp_id, es_super=es_super)
    finally:
        cur.close()
        conn.close()


@bp.route("/compras/nueva", methods=["GET", "POST"])
@rol_requerido("GERENTE")
def compras_nueva():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or request.form.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        proveedores_lista = proveedores_repo.listar(cur, emp_id)
        productos_lista = productos_repo.listar_inventario(cur, limit=500, empresa_id=emp_id)
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        from azdigital.services.compras_service import (
            calcular_retencion_iva_compras,
            linea_compra_a_unidad_base,
            opciones_presentacion_compra,
        )

        compra_pres_opts = {}
        for prod in productos_lista or []:
            try:
                pid_o = int(prod[0])
            except (TypeError, ValueError):
                continue
            compra_pres_opts[str(pid_o)] = opciones_presentacion_compra(cur, pid_o, prod)
        _tpl_compra_nueva = lambda **kw: render_template(
            "compras_nueva.html",
            proveedores=proveedores_lista,
            productos=productos_lista,
            empresas=empresas,
            emp_id=emp_id,
            es_super=es_super,
            today=date.today().isoformat(),
            compra_presentaciones_opts=compra_pres_opts,
            **kw,
        )
        if request.method == "POST":
            prov_id = int((request.form.get("proveedor_id") or "0"))
            num_fact = (request.form.get("numero_factura") or "").strip() or "S/N"
            fecha = (request.form.get("fecha") or date.today().isoformat())[:10]
            notas = (request.form.get("notas") or "").strip()
            if not prov_id:
                flash("Seleccione un proveedor.", "danger")
                return _tpl_compra_nueva()
            compra_id = compras_repo.crear(cur, emp_id, prov_id, num_fact, fecha, session.get("user_id"), notas)
            pids = request.form.getlist("producto_id")
            cants = request.form.getlist("cantidad")
            costos = request.form.getlist("costo")
            factors = request.form.getlist("factor_compra")
            pres_noms = request.form.getlist("presentacion_nombre")
            total_compra = 0.0
            for i, pid_raw in enumerate(pids):
                try:
                    pid = int(pid_raw)
                    cant = float((cants[i] if i < len(cants) else "0").replace(",", "."))
                    costo = float((costos[i] if i < len(costos) else "0").replace(",", "."))
                    fac_s = (factors[i] if i < len(factors) else "1").strip().replace(",", ".")
                    try:
                        factor = float(fac_s)
                    except ValueError:
                        factor = 1.0
                    if factor <= 0:
                        factor = 1.0
                    pres_nom = (pres_noms[i] if i < len(pres_noms) else "").strip() or None
                    if pid and cant > 0 and costo >= 0:
                        cant_umb, costo_umb, subtotal = linea_compra_a_unidad_base(cant, costo, factor)
                        total_compra += subtotal
                        compras_repo.agregar_detalle(
                            cur,
                            compra_id,
                            pid,
                            cant_umb,
                            costo_umb,
                            cantidad_recibida_presentacion=cant,
                            factor_conversion=factor,
                            presentacion_nombre=pres_nom,
                        )
                        lista_compras_repo.registrar_costo_compra(
                            cur, pid, costo_umb, cant_umb, session.get("user_id"), f"Compra #{compra_id}"
                        )
                        lista_compras_repo.actualizar_costo_y_precio_producto(cur, pid, emp_id, costo_umb, None)
                        notas_k = f"Compra #{compra_id}"
                        if abs(factor - 1.0) > 1e-9:
                            pn = pres_nom or "empaque"
                            notas_k += f": {cant:g} {pn} × {factor:g} = +{cant_umb:g} UMB"
                        try:
                            cur.execute("SELECT sucursal_id FROM productos WHERE id = %s", (pid,))
                            rp = cur.fetchone()
                            suc_id = int(rp[0]) if rp and rp[0] else kardex_repo.primera_sucursal_empresa(cur, emp_id)
                            if suc_id:
                                kardex_repo.registrar_entrada(
                                    cur, pid, suc_id, cant_umb, session.get("user_id"), notas_k
                                )
                            else:
                                productos_repo.incrementar_stock(cur, pid, cant_umb)
                        except Exception:
                            productos_repo.incrementar_stock(cur, pid, cant_umb)
                except (ValueError, Exception):
                    pass
            emp_gran = empresas_repo.get_empresa_es_gran_contribuyente(cur, emp_id)
            prov = proveedores_repo.get(cur, prov_id, emp_id)
            prov_gran = bool(prov[13]) if prov and len(prov) > 13 else False
            ret_iva = calcular_retencion_iva_compras(emp_gran, prov_gran, total_compra)
            compras_repo.actualizar_total(cur, compra_id, retencion_iva=ret_iva)
            registrar_accion(cur, historial_usuarios_repo.EVENTO_COMPRA_REGISTRADA, f"Compra #{compra_id} registrada")
            conn.commit()
            msg = f"Compra #{compra_id} registrada. Stock y costos actualizados."
            if ret_iva > 0:
                msg += f" Retención IVA 1%: ${ret_iva:,.2f}."
            flash(msg, "success")
            return redirect(url_for("admin.compras") + (f"?empresa_id={emp_id}" if es_super and emp_id else ""))
        return _tpl_compra_nueva()
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.compras"))


@bp.route("/compras/ver/<int:id>")
@rol_requerido("GERENTE")
def compras_ver(id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            emp_raw = (request.args.get("empresa_id") or "").strip()
            if emp_raw.isdigit():
                emp_id = int(emp_raw)
        c = compras_repo.get(cur, id, emp_id)
        if not c:
            flash("Compra no encontrada.", "danger")
            return redirect(url_for("admin.compras"))
        detalles = compras_repo.get_detalles(cur, id)
        cur.execute("SELECT nombre FROM proveedores WHERE id = %s", (c[1],))
        prov_nom = (cur.fetchone() or ("—",))[0]
        return render_template("compras_ver.html", compra=c, detalles=detalles, proveedor_nombre=prov_nom, emp_id=emp_id)
    finally:
        cur.close()
        conn.close()


@bp.route("/clientes")
@rol_requerido("GERENTE", "CAJERO")
def clientes():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        rol_str = str(session.get("rol") or "").strip().upper()
        es_super = rol_str in ("ADMIN", "SUPERADMIN") or _es_superadmin_db(cur)
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        sucursales_todas = (sucursales_repo.listar_sucursales_con_empresa(cur) or []) if es_super else []
        empresas_map = {e[0]: e[1] for e in empresas} if empresas else {}
        clientes_lista = clientes_repo.listar_clientes(cur, empresa_id=emp_id) or []
        actividades = []
        try:
            cur.execute("SELECT codigo, descripcion FROM actividades_economicas ORDER BY codigo")
            actividades = cur.fetchall() or []
        except Exception:
            pass
        return render_template("clientes.html", clientes=clientes_lista, cliente=None, cliente_dict=None, empresas=empresas, sucursales_todas=sucursales_todas, empresas_map=empresas_map, es_superadmin=es_super, actividades=actividades)
    finally:
        cur.close()
        conn.close()


@bp.route("/clientes/")
@rol_requerido("GERENTE", "CAJERO")
def clientes_slash():
    return redirect(url_for("admin.clientes"))


@bp.route("/clientes/editar/<int:id>")
@rol_requerido("GERENTE", "CAJERO")
def editar_cliente(id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        rol_str = str(session.get("rol") or "").strip().upper()
        es_super = rol_str in ("ADMIN", "SUPERADMIN") or _es_superadmin_db(cur)
        cliente_edit = clientes_repo.get_cliente(cur, id)
        if not cliente_edit:
            flash("Cliente no encontrado.", "danger")
            return redirect(url_for("admin.clientes"))
        if not es_super and (len(cliente_edit) > 1 and cliente_edit[1] != emp_id):
            flash("Cliente no pertenece a su empresa.", "danger")
            return redirect(url_for("admin.clientes"))
        cl = cliente_edit
        if len(cl) >= 11:
            cliente_dict = {"id": cl[0], "empresa_id": cl[1], "sucursal_id": cl[2], "nombre": cl[3], "tipo_documento": cl[4], "numero_documento": cl[5], "correo": cl[6], "es_contribuyente": cl[7], "es_gran_contribuyente": cl[8], "direccion": cl[9], "telefono": cl[10], "codigo_actividad_economica": cl[11] if len(cl) > 11 else ""}
        elif len(cl) >= 10:
            cliente_dict = {"id": cl[0], "empresa_id": cl[1], "sucursal_id": cl[2], "nombre": cl[3], "tipo_documento": cl[4], "numero_documento": cl[5], "correo": cl[6], "es_contribuyente": cl[7], "es_gran_contribuyente": False, "direccion": cl[8], "telefono": cl[9], "codigo_actividad_economica": cl[10] if len(cl) > 10 else ""}
        else:
            cliente_dict = {"id": cl[0], "empresa_id": cl[1], "sucursal_id": None, "nombre": cl[2], "tipo_documento": cl[3], "numero_documento": cl[4], "correo": cl[5], "es_contribuyente": cl[6], "es_gran_contribuyente": False, "direccion": cl[7], "telefono": cl[8], "codigo_actividad_economica": ""}
        empresas = (empresas_repo.listar_empresas(cur) or []) if es_super else []
        sucursales_todas = (sucursales_repo.listar_sucursales_con_empresa(cur) or []) if es_super else []
        empresas_map = {e[0]: e[1] for e in empresas} if empresas else {}
        clientes_lista = clientes_repo.listar_clientes(cur, empresa_id=emp_id) or []
        actividades = []
        try:
            cur.execute("SELECT codigo, descripcion FROM actividades_economicas ORDER BY codigo")
            actividades = cur.fetchall() or []
        except Exception:
            pass
        return render_template("clientes.html", clientes=clientes_lista, cliente=cliente_edit, cliente_dict=cliente_dict, empresas=empresas, sucursales_todas=sucursales_todas, empresas_map=empresas_map, es_superadmin=es_super, actividades=actividades)
    finally:
        cur.close()
        conn.close()


@bp.route("/guardar_cliente", methods=["POST"])
@rol_requerido("GERENTE", "CAJERO")
def guardar_cliente():
    form = request.form
    cliente_id = form.get("cliente_id")
    nombre_cliente = (form.get("nombre_cliente") or "").strip()
    tipo_documento = (form.get("tipo_documento") or "").strip()
    numero_documento = (form.get("numero_documento") or "").strip()
    correo = (form.get("correo") or "").strip()
    es_contribuyente = bool(form.get("es_contribuyente"))
    es_gran_contribuyente = bool(form.get("es_gran_contribuyente"))
    direccion = (form.get("direccion") or "").strip()
    telefono = (form.get("telefono") or "").strip()
    codigo_actividad = (form.get("codigo_actividad_economica") or "").strip()

    if not nombre_cliente:
        return redirect(url_for("admin.clientes"))

    tipo_doc = (tipo_documento or "").strip().upper()
    if tipo_doc == "NIT" and numero_documento:
        from azdigital.utils.validar_documentos import validar_nit
        ok_nit, msg = validar_nit(numero_documento)
        if not ok_nit:
            flash(msg, "danger")
            if cliente_id and str(cliente_id).isdigit():
                return redirect(url_for("admin.editar_cliente", id=cliente_id))
            return redirect(url_for("admin.clientes"))
    if tipo_doc == "NRC" and numero_documento:
        from azdigital.utils.validar_documentos import validar_nrc
        ok_nrc, msg = validar_nrc(numero_documento)
        if not ok_nrc:
            flash(msg, "danger")
            if cliente_id and str(cliente_id).isdigit():
                return redirect(url_for("admin.editar_cliente", id=cliente_id))
            return redirect(url_for("admin.clientes"))

    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        rol_str = str(session.get("rol") or "").strip().upper()
        es_super = rol_str in ("ADMIN", "SUPERADMIN") or _es_superadmin_db(cur)
        emp_form = form.get("empresa_id")
        empresa_id = int(emp_form) if es_super and emp_form and str(emp_form).isdigit() else _empresa_id()
        suc_form = form.get("sucursal_id")
        sucursal_id = int(suc_form) if suc_form and str(suc_form).isdigit() else None

        if cliente_id and cliente_id.isdigit():
            try:
                clientes_repo.actualizar_cliente(
                    cur,
                    int(cliente_id),
                    nombre_cliente,
                    tipo_documento,
                    numero_documento,
                    correo,
                    es_contribuyente,
                    direccion,
                    telefono,
                    empresa_id=empresa_id if es_super else None,
                    sucursal_id=sucursal_id,
                    actualizar_sucursal=es_super,
                    codigo_actividad_economica=codigo_actividad,
                    es_gran_contribuyente=es_gran_contribuyente,
                )
            except Exception:
                conn.rollback()
                clientes_repo.actualizar_cliente(
                    cur,
                    int(cliente_id),
                    nombre_cliente,
                    tipo_documento,
                    numero_documento,
                    correo,
                    es_contribuyente,
                    direccion,
                    telefono,
                    empresa_id=empresa_id if es_super else None,
                    sucursal_id=None,
                    actualizar_sucursal=False,
                    codigo_actividad_economica=codigo_actividad,
                    es_gran_contribuyente=es_gran_contribuyente,
                )
        else:
            clientes_repo.crear_cliente(
                cur,
                empresa_id=empresa_id,
                nombre_cliente=nombre_cliente,
                tipo_documento=tipo_documento,
                numero_documento=numero_documento,
                correo=correo,
                es_contribuyente=es_contribuyente,
                direccion=direccion,
                telefono=telefono,
                sucursal_id=sucursal_id if es_super else None,
                codigo_actividad_economica=codigo_actividad,
                es_gran_contribuyente=es_gran_contribuyente,
            )
        registrar_accion(cur, historial_usuarios_repo.EVENTO_CLIENTE_EDITADO if (cliente_id and cliente_id.isdigit()) else historial_usuarios_repo.EVENTO_CLIENTE_CREADO, f"Cliente {nombre_cliente} guardado")
        conn.commit()
        flash("Cliente guardado correctamente.", "success")
        return redirect(url_for("admin.clientes"))
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f"Error al guardar: {str(e)}", "danger")
        return redirect(url_for("admin.clientes"))
    finally:
        cur.close()
        conn.close()


@bp.route("/eliminar_cliente/<int:cliente_id>", methods=["GET", "POST"])
@rol_requerido("GERENTE", "CAJERO")
def eliminar_cliente(cliente_id):
    if request.method == "GET":
        return redirect(url_for("admin.clientes"))
    emp_id = _empresa_id()
    rol_str = str(session.get("rol") or "").strip().upper()
    es_super = rol_str in ("ADMIN", "SUPERADMIN")
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if not es_super:
            es_super = _es_superadmin_db(cur)
        c = clientes_repo.get_cliente(cur, cliente_id)
        if not c:
            flash("Cliente no encontrado.", "danger")
            return redirect(url_for("admin.clientes"))
        if not es_super and len(c) > 1 and c[1] != emp_id:
            flash("Cliente no pertenece a su empresa.", "danger")
            return redirect(url_for("admin.clientes"))
        clientes_repo.eliminar_cliente(cur, cliente_id)
        registrar_accion(cur, historial_usuarios_repo.EVENTO_CLIENTE_ELIMINADO, f"Cliente #{cliente_id} eliminado")
        conn.commit()
        flash("Cliente eliminado.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.clientes"))


def _empresa_id_desde_form_super(cur, es_super: bool, form_empresa_id: str | None, fallback: int) -> int:
    """Si es superusuario, toma empresa del formulario si es válida; si no, fallback."""
    if not es_super:
        return fallback
    raw = (form_empresa_id or "").strip()
    if raw.isdigit():
        eid = int(raw)
        if empresas_repo.get_empresa(cur, eid):
            return eid
    return fallback


@bp.route("/sucursales")
@admin_required
def sucursales():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        if es_super:
            sucursales_lista = sucursales_repo.listar_sucursales_todas_con_empresa(cur) or []
            empresas = empresas_repo.listar_empresas(cur) or []
        else:
            sucursales_lista = sucursales_repo.listar_sucursales(cur, empresa_id=emp_id) or []
            empresas = []
        return render_template(
            "sucursales.html",
            sucursales=sucursales_lista,
            sucursal=None,
            empresas=empresas,
            es_superadmin=es_super,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/sucursales/")
@admin_required
def sucursales_slash():
    return sucursales()


@bp.route("/sucursales/editar/<int:id>")
@admin_required
def editar_sucursal(id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        suc = sucursales_repo.get_sucursal(cur, id)
        if not suc:
            flash("Sucursal no encontrada.", "danger")
            return redirect(url_for("admin.sucursales"))
        if not es_super and (len(suc) < 6 or suc[5] != emp_id):
            flash("Sucursal no pertenece a su empresa.", "danger")
            return redirect(url_for("admin.sucursales"))
        if es_super:
            sucursales_lista = sucursales_repo.listar_sucursales_todas_con_empresa(cur) or []
            empresas = empresas_repo.listar_empresas(cur) or []
        else:
            sucursales_lista = sucursales_repo.listar_sucursales(cur, empresa_id=emp_id) or []
            empresas = []
        return render_template(
            "sucursales.html",
            sucursales=sucursales_lista,
            sucursal=suc,
            empresas=empresas,
            es_superadmin=es_super,
        )
    finally:
        cur.close()
        conn.close()


@bp.route("/guardar_sucursal", methods=["POST"])
@admin_required
def guardar_sucursal():
    nombre = (request.form.get("nombre") or "").strip()
    codigo = (request.form.get("codigo") or "").strip()
    direccion = (request.form.get("direccion") or "").strip()
    telefono = (request.form.get("telefono") or "").strip()
    sucursal_id = request.form.get("sucursal_id")
    if not nombre or not codigo:
        flash("Nombre y código son requeridos.", "danger")
        return redirect(url_for("admin.sucursales"))

    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    emp_id = _empresa_id()
    try:
        es_super = _es_superadmin_db(cur)
        if sucursal_id and sucursal_id.isdigit():
            suc = sucursales_repo.get_sucursal(cur, int(sucursal_id))
            if suc and (es_super or (len(suc) >= 6 and suc[5] == emp_id)):
                nueva_emp = None
                if es_super:
                    fb = int(suc[5]) if len(suc) >= 6 and suc[5] is not None else emp_id
                    nueva_emp = _empresa_id_desde_form_super(cur, True, request.form.get("empresa_id"), fb)
                sucursales_repo.actualizar_sucursal(
                    cur, int(sucursal_id), nombre, codigo, direccion, telefono, empresa_id=nueva_emp if es_super else None
                )
                registrar_accion(cur, historial_usuarios_repo.EVENTO_SUCURSAL_EDITADA, f"Sucursal {nombre} actualizada")
                conn.commit()
                flash("Sucursal actualizada correctamente.", "success")
            else:
                flash("Sucursal no encontrada.", "danger")
        else:
            target_emp = _empresa_id_desde_form_super(cur, es_super, request.form.get("empresa_id"), emp_id)
            sucursales_repo.crear_sucursal(cur, nombre, codigo, direccion, telefono, empresa_id=target_emp)
            registrar_accion(cur, historial_usuarios_repo.EVENTO_SUCURSAL_CREADA, f"Sucursal {nombre} registrada")
            conn.commit()
            flash("Sucursal registrada correctamente.", "success")
        return redirect(url_for("admin.sucursales"))
    except Exception as e:
        conn.rollback()
        flash(f"Error al guardar: {str(e)}", "danger")
        return redirect(url_for("admin.sucursales"))
    finally:
        cur.close()
        conn.close()


@bp.route("/eliminar_sucursal/<int:sucursal_id>", methods=["GET", "POST"])
@admin_required
def eliminar_sucursal(sucursal_id):
    if request.method == "GET":
        return redirect(url_for("admin.sucursales"))
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        es_super = _es_superadmin_db(cur)
        suc = sucursales_repo.get_sucursal(cur, sucursal_id)
        if not suc:
            flash("Sucursal no encontrada.", "danger")
            return redirect(url_for("admin.sucursales"))
        if not es_super and (len(suc) < 6 or suc[5] != emp_id):
            flash("Sucursal no pertenece a su empresa.", "danger")
            return redirect(url_for("admin.sucursales"))
        sucursales_repo.eliminar_sucursal(cur, sucursal_id)
        registrar_accion(cur, historial_usuarios_repo.EVENTO_SUCURSAL_ELIMINADA, f"Sucursal #{sucursal_id} eliminada")
        conn.commit()
        flash("Sucursal eliminada.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.sucursales"))


@bp.route("/gestion_ventas")
@rol_requerido("GERENTE")
def gestion_ventas():
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        rows = ventas_repo.listar_ventas_recientes(cur, empresa_id=emp_id, limit=150) or []
        return render_template("gestion_ventas.html", ventas=rows)
    finally:
        cur.close()
        conn.close()


@bp.route("/gestion_ventas/editar/<int:venta_id>", methods=["GET", "POST"])
@rol_requerido("GERENTE")
def gestion_ventas_editar(venta_id):
    emp_id = _empresa_id()
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        v = ventas_repo.get_venta(cur, venta_id, empresa_id=emp_id)
        if not v:
            flash("Venta no encontrada.", "danger")
            return redirect(url_for("admin.gestion_ventas"))
        if request.method == "POST":
            cn = (request.form.get("cliente_nombre") or "").strip() or "Consumidor Final"
            tp = (request.form.get("tipo_pago") or "EFECTIVO").strip().upper()
            tc = (request.form.get("tipo_comprobante") or "TICKET").strip().upper()
            if tc not in ("TICKET", "FACTURA", "CREDITO_FISCAL"):
                tc = "TICKET"
            cr = (request.form.get("cliente_id") or "").strip()
            cid = int(cr) if cr.isdigit() else None
            if cid:
                snap = clientes_repo.snapshot_cliente_venta_por_id(cur, cid, emp_id)
                if snap:
                    cn = snap
            if tc == "CREDITO_FISCAL":
                if not cid:
                    flash("Crédito fiscal requiere cliente del catálogo (ID).", "danger")
                    return redirect(url_for("admin.gestion_ventas_editar", venta_id=venta_id))
                cur.execute(
                    "SELECT COALESCE(es_contribuyente, FALSE) FROM clientes WHERE id = %s AND empresa_id = %s",
                    (cid, emp_id),
                )
                r = cur.fetchone()
                if not r or not r[0]:
                    flash("Crédito fiscal: el cliente debe ser contribuyente.", "danger")
                    return redirect(url_for("admin.gestion_ventas_editar", venta_id=venta_id))
            ec = (request.form.get("estado_cobro") or "COBRADO").strip().upper()
            total_venta = float(v[2]) if v and len(v) > 2 else 0
            from azdigital.services.ventas_service import calcular_retencion_iva
            ret_iva = calcular_retencion_iva(cur, total_venta, tc, cid, emp_id)
            ok = ventas_repo.actualizar_venta_cabecera(cur, venta_id, emp_id, cn, tp, tc, cid, estado_cobro=ec if ec in ("COBRADO", "PENDIENTE") else "COBRADO", retencion_iva=ret_iva)
            if ok:
                registrar_accion(cur, historial_usuarios_repo.EVENTO_VENTA_EDITADA, f"Venta #{venta_id} actualizada")
                conn.commit()
                flash("Venta actualizada.", "success")
            else:
                conn.rollback()
                flash("No se pudo actualizar.", "warning")
            return redirect(url_for("admin.gestion_ventas"))
        detalles = ventas_repo.get_detalles(cur, venta_id) or []
        return render_template("gestion_venta_editar.html", venta=v, detalles=detalles, venta_id=venta_id)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f"Error: {e}", "danger")
        return redirect(url_for("admin.gestion_ventas"))
    finally:
        cur.close()
        conn.close()


@bp.route("/gestion_ventas/eliminar/<int:venta_id>", methods=["POST"])
@rol_requerido("GERENTE")
def gestion_ventas_eliminar(venta_id):
    emp_id = _empresa_id()
    motivo = (request.form.get("motivo_anulacion") or "").strip() or "Anulación por gestión"
    usuario_id = session.get("user_id")
    db = ConexionDB()
    conn = psycopg2.connect(**db.config)
    cur = conn.cursor()
    try:
        if ventas_repo.eliminar_venta_restaurar_stock(cur, venta_id, emp_id, motivo_anulacion=motivo, usuario_anulo_id=usuario_id):
            registrar_accion(cur, historial_usuarios_repo.EVENTO_VENTA_ANULADA, f"Venta #{venta_id} anulada. Motivo: {motivo[:80]}")
            conn.commit()
            flash("Venta anulada. Stock repuesto. Motivo registrado para auditoría.", "success")
        else:
            conn.rollback()
            flash("Venta no encontrada.", "warning")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin.gestion_ventas"))

